"""
Генерация Excel-отчёта «Анализ жалоб».
Листы:
- "Сводная"    : итоговая сводка, топ проблемных ресторанов, сводная по жалобам
                 (+ % долей), 10 расширенных категорий, позитив, средний рейтинг;
- "Жалобы"     : детальные жалобы;
- "Дубликаты"  : дубли по номеру телефона;
- "Удалённые"  : удалённые отзывы геосервисов;
- "По источникам" : сводные по каждому источнику отдельно (ОС и Тюмень объединены);
- "Компенсации"   : сводная по типам решений (без сумм), необработанные обращения
                    и обращения с решением «без компенсации».
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from . import constants as c
except Exception:
    try:
        from report.complaints import constants as c
    except Exception:
        c = None


def _const(name: str, default):
    if c is not None and hasattr(c, name):
        return getattr(c, name)
    return default


SHEET_SUMMARY = _const("EXCEL_SHEET_SUMMARY", "Сводная")
SHEET_DETAIL = _const("EXCEL_SHEET_DETAIL", "Жалобы")
SHEET_DUPLICATES = _const("EXCEL_SHEET_DUPLICATES", "Дубликаты")
SHEET_DELETED = _const("EXCEL_SHEET_DELETED", "Удалённые")
SHEET_BY_SOURCE = "По источникам"
SHEET_COMPENSATIONS = "Компенсации"

# ----------------------------------------------------------
# СТИЛИ
# ----------------------------------------------------------
THIN_SIDE = Side(style="thin")
DOUBLE_SIDE = Side(style="double")
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE
)
TOTAL_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=DOUBLE_SIDE, bottom=DOUBLE_SIDE
)

TITLE_FONT = Font(name="Calibri", size=14, bold=True)
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True)
DATA_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)

GREEN_FILL = PatternFill(fill_type="solid", start_color="D9EAD3", end_color="D9EAD3")
YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
BLUE_FILL = PatternFill(fill_type="solid", start_color="D6EAF8", end_color="D6EAF8")

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ----------------------------------------------------------
# ВСПОМОГАТЕЛЬНЫЕ
# ----------------------------------------------------------
def _clean(val):
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except Exception:
        pass
    return val


def _get(obj, names, default=None):
    """Терпимо достаёт поле из dataclass / dict / None."""
    if obj is None:
        return default
    if isinstance(names, str):
        names = [names]
    if isinstance(obj, dict):
        for n in names:
            if n in obj:
                return obj[n]
        return default
    for n in names:
        if hasattr(obj, n):
            return getattr(obj, n)
    return default


def _to_df(obj) -> pd.DataFrame:
    if obj is None:
        return pd.DataFrame()
    if isinstance(obj, pd.DataFrame):
        return obj
    try:
        return pd.DataFrame(obj)
    except Exception:
        return pd.DataFrame()


def _get_df(stats, prepared, stats_names, prepared_names):
    """
    Безопасно достаёт DataFrame: сначала из stats, потом из prepared.
    НЕ использует оператор `or` между DataFrame.
    """
    val = _get(stats, stats_names)
    if val is None:
        val = _get(prepared, prepared_names)
    return _to_df(val)


def _set_widths(ws, columns, sample_rows) -> None:
    for j, colname in enumerate(columns, start=1):
        name = str(colname)
        width = len(name) + 2
        for row in sample_rows:
            try:
                width = max(width, len(str(_clean(row.iloc[j - 1]))) + 2)
            except Exception:
                pass
        if j == 1:
            width = max(width, 20)
        if "Текст" in name or "Комментарий" in name:
            width = 60
        ws.column_dimensions[get_column_letter(j)].width = min(width, 70)


def _apply_sum_formulas(ws, first_data_row: int, last_data_row: int, columns) -> None:
    """Заменяет посчитанные в Python итоги формулами =SUM, чтобы при ручной правке
    цифр в Excel итоги пересчитывались сами. Работает с таблицами вида
    "ресторан x категория" со строками "Всего ...:" — определяет их по первому
    столбцу, а группы (СПб/Тюмень) — по границам между такими строками.
    Столбец "Всего:" (если есть) — сумма категорий текущей строки."""
    last_col_idx = len(columns)
    if last_col_idx < 2:
        return
    first_cat_col_idx = 2
    has_total_col = str(columns[-1]).strip() == "Всего:"
    cat_last_col_idx = (last_col_idx - 1) if has_total_col else last_col_idx

    group_start = first_data_row
    for r in range(first_data_row, last_data_row + 1):
        first_value = str(ws.cell(row=r, column=1).value or "")
        is_group_total = first_value.startswith("Всего") or first_value.startswith("Итого")
        if is_group_total:
            if r - 1 >= group_start:
                for col_idx in range(first_cat_col_idx, last_col_idx + 1):
                    col_letter = get_column_letter(col_idx)
                    ws.cell(row=r, column=col_idx).value = f"=SUM({col_letter}{group_start}:{col_letter}{r - 1})"
            group_start = r + 1
        elif has_total_col:
            first_letter = get_column_letter(first_cat_col_idx)
            last_letter = get_column_letter(cat_last_col_idx)
            ws.cell(row=r, column=last_col_idx).value = f"=SUM({first_letter}{r}:{last_letter}{r})"


def _write_df_block(ws, start_row: int, title: str, df: pd.DataFrame, formula_totals: bool = False) -> int:
    """Записывает заголовок + таблицу. Возвращает следующую свободную строку.

    formula_totals=True — для таблиц "ресторан x категория" (сводная по жалобам,
    по позитиву, по решениям и т.п.): столбец "Всего:" и строки "Всего ...:"
    пишутся формулами =SUM вместо готовых чисел. Не подходит для таблиц с %
    и средними — там сумма ячеек не равна показанному итогу."""
    if df is None or len(df) == 0 or len(df.columns) == 0:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.font = TITLE_FONT
        cell = ws.cell(row=start_row + 1, column=1, value="Нет данных")
        cell.font = DATA_FONT
        return start_row + 3

    max_col = len(df.columns)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
    tcell = ws.cell(row=start_row, column=1, value=title)
    tcell.font = TITLE_FONT
    tcell.alignment = CENTER_ALIGN

    hrow = start_row + 1
    for j, colname in enumerate(df.columns, start=1):
        cell = ws.cell(row=hrow, column=j, value=str(colname))
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = GREEN_FILL
        cell.border = THIN_BORDER

    r = hrow + 1
    for _, row in df.iterrows():
        first_value = str(_clean(row.iloc[0]))
        is_total = first_value.startswith("Всего") or first_value.startswith("Итого")
        for j, colname in enumerate(df.columns, start=1):
            val = _clean(row[colname])
            cell = ws.cell(row=r, column=j, value=val)
            if is_total:
                cell.font = TOTAL_FONT
                cell.fill = YELLOW_FILL
                cell.border = TOTAL_BORDER
                cell.alignment = CENTER_ALIGN if j > 1 else LEFT_ALIGN
            else:
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN if j > 1 else LEFT_ALIGN
        r += 1

    last_data_row = r - 1
    if formula_totals:
        _apply_sum_formulas(ws, hrow + 1, last_data_row, list(df.columns))

    _set_widths(ws, df.columns, [x[1] for x in df.head(50).iterrows()])
    return r + 2


# ----------------------------------------------------------
# СБОРКА КНИГИ
# ----------------------------------------------------------
def build_complaints_excel(prepared=None, stats=None, settings=None, **kwargs) -> io.BytesIO:
    """
    Собирает книгу «Анализ жалоб».
    Принимает prepared (dataclass/dict), stats (dict) и опционально settings.
    period_label можно передать через kwargs.
    """
    if prepared is None:
        prepared = kwargs.get("data") or kwargs.get("report_data")
    if stats is None:
        stats = kwargs.get("stats_dict") or {}

    # period_label: из settings, или из kwargs, или пустая строка
    period_label = _get(settings, "period_label", "")
    if not period_label:
        period_label = kwargs.get("period_label", "")
    period_label = period_label or ""

    # --- Все DataFrame достаём через _get_df (без оператора `or` между DF) ---
    complaints_summary = _get_df(
        stats, prepared,
        ["complaint_summary", "complaints_summary"],
        ["complaint_summary", "complaints_summary"],
    )
    positives_summary = _get_df(
        stats, prepared,
        ["positive_summary", "positives_summary"],
        ["positive_summary", "positives_summary"],
    )
    # НОВОЕ: 10 расширенных категорий вместо 5 основных
    extended_categories_summary = _get_df(
        stats, prepared,
        ["extended_categories_summary", "extended_summary"],
        ["extended_categories_summary"],
    )
    complaints_detail = _get_df(
        stats, prepared,
        ["complaints", "complaints_detail"],
        ["complaints", "complaints_detail", "detail"],
    )
    duplicates = _get_df(
        stats, prepared,
        ["duplicates", "duplicates_df"],
        ["duplicates", "duplicates_df"],
    )
    deleted_geo = _to_df(
        _get(prepared, ["deleted_geo", "deleted_geo_df", "deleted"])
    )

    # НОВОЕ: итоговая сводка, средний рейтинг, % долей, топ ресторанов, компенсации
    report_overview = _get_df(stats, prepared, ["report_overview"], ["report_overview"])
    avg_rating_summary = _get_df(stats, prepared, ["avg_rating_summary"], ["avg_rating_summary"])
    complaint_share_summary = _get_df(stats, prepared, ["complaint_share_summary"], ["complaint_share_summary"])
    top_restaurants = _get_df(stats, prepared, ["top_restaurants"], ["top_restaurants"])
    resolution_summary = _get_df(stats, prepared, ["resolution_summary"], ["resolution_summary"])
    unresolved_complaints = _get_df(stats, prepared, ["unresolved_complaints"], ["unresolved_complaints"])
    no_compensation_complaints = _get_df(stats, prepared, ["no_compensation_complaints"], ["no_compensation_complaints"])

    # by_source — это dict, не DataFrame, поэтому `or {}` безопасен
    by_source_stats = _get(stats, ["by_source", "by_source_stats"])
    if by_source_stats is None:
        by_source_stats = {}

    wb = Workbook()

    # ------------------------------------------------------
    # ЛИСТ 1: СВОДНАЯ
    # ------------------------------------------------------
    ws = wb.active
    ws.title = SHEET_SUMMARY
    row = 1
    if period_label:
        ws.cell(row=row, column=1, value=period_label).font = TITLE_FONT
        row += 2

    row = _write_df_block(ws, row, "Итоговая сводка", report_overview)
    row = _write_df_block(ws, row, "Топ проблемных ресторанов", top_restaurants)
    row = _write_df_block(ws, row, "Сводная по жалобам", complaints_summary, formula_totals=True)
    row = _write_df_block(ws, row, "Доля категорий от общего числа жалоб ресторана, %", complaint_share_summary)
    # ИЗМЕНЕНО: теперь пишем 10 расширенных категорий вместо 5 основных
    row = _write_df_block(ws, row, "Сводная по жалобам (10 расширенных категорий)", extended_categories_summary, formula_totals=True)
    row = _write_df_block(ws, row, "Сводная по положительным моментам", positives_summary, formula_totals=True)
    row = _write_df_block(ws, row, "Средний рейтинг по ресторанам", avg_rating_summary)

    # ------------------------------------------------------
    # ЛИСТ 2: ЖАЛОБЫ (ДЕТАЛИ)
    # ------------------------------------------------------
    ws2 = wb.create_sheet(SHEET_DETAIL)
    r2 = 1
    if period_label:
        ws2.cell(row=r2, column=1, value=period_label).font = TITLE_FONT
        r2 += 2
    _write_df_block(ws2, r2, "Жалобы (без дублей)", complaints_detail)

    # ------------------------------------------------------
    # ЛИСТ 3: ДУБЛИКАТЫ
    # ------------------------------------------------------
    ws3 = wb.create_sheet(SHEET_DUPLICATES)
    _write_df_block(ws3, 1, "Дубликаты по номеру телефона", duplicates)

    # ------------------------------------------------------
    # ЛИСТ 4: УДАЛЁННЫЕ (ГЕО)
    # ------------------------------------------------------
    ws4 = wb.create_sheet(SHEET_DELETED)
    _write_df_block(ws4, 1, "Удалённые отзывы геосервисов", deleted_geo)

    # ------------------------------------------------------
    # ЛИСТ 5: ПО ИСТОЧНИКАМ (8 таблиц)
    # ------------------------------------------------------
    ws5 = wb.create_sheet(SHEET_BY_SOURCE)
    r5 = 1
    if period_label:
        ws5.cell(row=r5, column=1, value=period_label).font = TITLE_FONT
        r5 += 2

    # «ОС и компенсации» объединяет ОС (СПб) и ОС Тюмень — это один канал,
    # разделять их в отчёте не нужно. Позитива для ОС не считаем — это не отзывы.
    source_order = ["ОС и компенсации", "Сайт", "Агрегаторы", "Геосервисы"]
    for source in source_order:
        source_data = by_source_stats.get(source, {})
        complaint_df = _to_df(source_data.get("complaint_summary"))
        r5 = _write_df_block(ws5, r5, f"Жалобы — {source}", complaint_df, formula_totals=True)
        positive_raw = source_data.get("positive_summary")
        if positive_raw is not None:
            r5 = _write_df_block(ws5, r5, f"Позитив — {source}", _to_df(positive_raw), formula_totals=True)

    # ------------------------------------------------------
    # ЛИСТ 6: КОМПЕНСАЦИИ (только количество — без сумм)
    # ------------------------------------------------------
    ws6 = wb.create_sheet(SHEET_COMPENSATIONS)
    r6 = 1
    if period_label:
        ws6.cell(row=r6, column=1, value=period_label).font = TITLE_FONT
        r6 += 2
    r6 = _write_df_block(ws6, r6, "Сводная по решениям (кол-во обращений)", resolution_summary, formula_totals=True)
    r6 = _write_df_block(ws6, r6, "Необработанные обращения", unresolved_complaints)
    _write_df_block(ws6, r6, "Решение «без компенсации»", no_compensation_complaints)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


build_excel = build_complaints_excel