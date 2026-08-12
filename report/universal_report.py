"""
Логика универсального отчёта.

Файл содержит:
- чтение исходного Excel-файла;
- нормализацию названий;
- определение города и периода;
- агрегацию данных;
- сборку Excel-отчёта.

Streamlit здесь не используется.
"""

from __future__ import annotations

import re
import warnings
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# ==========================================================
# КОНСТАНТЫ
# ==========================================================
CITY_SPB = "СПБ"
CITY_TYUMEN = "Тюмень"
CITY_ORDER = [CITY_SPB, CITY_TYUMEN]

PIZZA_SIZES = ["15", "23", "30", "35", "40"]

REQUIRED_COLUMNS = [
    "Юридическое лицо",
    "Блюдо",
    "Количество блюд",
    "Сумма со скидкой, р.",
]

LATIN_TO_CYRILLIC = {
    "a": "а",
    "e": "е",
    "o": "о",
    "p": "р",
    "c": "с",
    "y": "у",
    "x": "х",
    "A": "А",
    "E": "Е",
    "O": "О",
    "P": "Р",
    "C": "С",
    "Y": "У",
    "X": "Х",
    "ё": "е",
    "Ё": "Е",
}

NUMBER_WORDS = {
    "четыре": "4",
    "пять": "5",
    "шесть": "6",
    "семь": "7",
    "восемь": "8",
    "девять": "9",
    "десять": "10",
}


# ==========================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ==========================================================
def _parse_rules(rules_text: str) -> list[str]:
    """
    Разбирает текстовое поле со списком позиций.
    Одна строка = одна позиция.
    """
    return [line.strip() for line in rules_text.splitlines() if line.strip()]


def normalize_text(text) -> str:
    """
    Нормализует название блюда.

    Правила:
    - латиница частично переводится в кириллицу;
    - скобки НЕ удаляются вместе с содержимым, а заменяются пробелами
      (чтобы сохранить вес/размер вида "(150 гр)");
    - единицы приводятся к единому виду: "150гр" -> "150 гр",
      "150 г" -> "150 гр", "150 грамм" -> "150 гр";
    - "см" убирается (размер пиццы остаётся числом);
    - слова-паразиты убираются;
    - числа-слова унифицируются ("четыре" -> "4").
    """
    if not isinstance(text, str):
        return ""

    # Удаляем англоязычные служебные слова до замены латиницы
    text = re.sub(r"(?i)\b(pizza|new)\b", "", text)

    # Латиница в кириллицу
    for lat, cyr in LATIN_TO_CYRILLIC.items():
        text = text.replace(lat, cyr)

    # Базовая очистка
    text = text.lower().replace('"', "").replace("'", "")

    # Скобки заменяем пробелами, содержимое сохраняем
    text = text.replace("(", " ").replace(")", " ")

    # Разделяем число и единицу, унифицируем единицы
    text = re.sub(r"(\d+)\s*(см|грамм|гр|г)\b", r"\1 \2", text)
    text = re.sub(r"\bграмм\b|\bг\b", "гр", text)

    # "см" убираем: размер пиццы остаётся просто числом
    text = re.sub(r"\bсм\b", " ", text)

    # Убираем слова-паразиты для пицц
    text = re.sub(r"\bпицца\b", "", text)
    text = re.sub(
        r"\bтонкое\b|\bтрад\b|\bтрадиционное\b|\bтесто\b",
        "",
        text,
    )

    # Унификация чисел
    for word, digit in NUMBER_WORDS.items():
        text = text.replace(word, digit)

    # Убираем лишние пробелы
    text = re.sub(r"\s+", " ", text).strip()

    return text


def extract_size(text) -> str | None:
    """
    Извлекает размер пиццы: 15, 23, 30, 35 или 40.
    """
    if not isinstance(text, str):
        return None

    match = re.search(
        r"\b(15|23|30|35|40)\s*см\b",
        str(text),
        flags=re.IGNORECASE,
    )

    return match.group(1) if match else None


def extract_pizza_name(text) -> str:
    """
    Извлекает название пиццы без размера и типа теста.
    """
    if not isinstance(text, str):
        return ""

    name = str(text)

    # Убираем размер
    name = re.sub(
        r"\s*\b(15|23|30|35|40)\s*см\b",
        "",
        name,
        flags=re.IGNORECASE,
    )

    # Убираем тип теста в скобках
    name = re.sub(
        r"\s*\([^)]*(?:тонкое|трад|традиционное|тесто)[^)]*\)",
        "",
        name,
        flags=re.IGNORECASE,
    )

    # Убираем служебные слова
    name = re.sub(
        r"\b(?:пицца|pizza|тонкое|трад|традиционное|тесто|см|new)\b",
        "",
        name,
        flags=re.IGNORECASE,
    )

    # Скобки заменяем пробелами
    name = name.replace("(", " ").replace(")", " ")

    name = re.sub(r"\s+", " ", name).strip()

    return name


def map_city(legal_entity) -> str | None:
    """
    Определяет город по юридическому лицу.
    """
    le = str(legal_entity).upper()

    if "СПБ" in le:
        return CITY_SPB

    if "ПД" in le and "СПБ" not in le:
        return CITY_TYUMEN

    return None


# ==========================================================
# ЧТЕНИЕ ФАЙЛА
# ==========================================================
def extract_period(uploaded_file) -> str:
    """
    Ищет период в первых строках файла.
    Пример:
        Период 01.05.2026 по 31.05.2026
    """
    uploaded_file.seek(0)

    df = pd.read_excel(
        uploaded_file,
        header=None,
        nrows=10,
    )

    uploaded_file.seek(0)

    for _, row in df.iterrows():
        for val in row.values:
            if pd.notna(val) and "Период" in str(val):
                match = re.search(
                    r"(\d{2}\.\d{2}\.\d{4})\s+по\s+(\d{2}\.\d{2}\.\d{4})",
                    str(val),
                )

                if match:
                    start = datetime.strptime(match.group(1), "%d.%m.%Y")
                    end = datetime.strptime(match.group(2), "%d.%m.%Y")

                    return f"{start.strftime('%d.%m')}-{end.strftime('%d.%m.%Y')}"

    return f"01.{datetime.now().strftime('%m')}.{datetime.now().year}"


def read_file(uploaded_file) -> pd.DataFrame:
    """
    Читает Excel-файл, находит заголовки, очищает данные.
    """
    uploaded_file.seek(0)

    df = pd.read_excel(
        uploaded_file,
        header=None,
    )

    header_row = None

    for idx, row in df.iterrows():
        vals = [str(v) for v in row.values if pd.notna(v)]
        joined = " ".join(vals)

        if "Юридическое лицо" in joined and "Блюдо" in joined:
            header_row = idx
            break

    if header_row is None:
        raise ValueError("Не найдены заголовки!")

    headers = [
        str(h).strip() if pd.notna(h) else f"column_{i}"
        for i, h in enumerate(df.iloc[header_row].values)
    ]

    df = df[header_row + 1:].copy()
    df.columns = headers
    df = df.reset_index(drop=True)

    missing_columns = [
        column
        for column in REQUIRED_COLUMNS
        if column not in df.columns
    ]

    if missing_columns:
        raise ValueError(
            "В файле отсутствуют обязательные колонки: "
            + ", ".join(missing_columns)
        )

    # Убираем технические строки и итоги
    df = df[
        ~df["Юридическое лицо"]
        .astype(str)
        .str.contains("OLAP|всего|Итого", case=False, na=False)
    ]

    df = df.dropna(how="all")

    # Заполняем пропуски после группировок
    df["Юридическое лицо"] = df["Юридическое лицо"].ffill()

    if "Категория блюда" in df.columns:
        df["Категория блюда"] = df["Категория блюда"].ffill()

    # Убираем служебные блюда
    df = df[
        ~df["Блюдо"]
        .astype(str)
        .str.contains("+", na=False, regex=False)
    ]

    df = df[
        ~df["Блюдо"]
        .astype(str)
        .str.contains("персонал", case=False, na=False)
    ]

    df = df[df["Блюдо"].notna()]

    # Числовые поля
    df["Количество блюд"] = (
        pd.to_numeric(df["Количество блюд"], errors="coerce")
        .fillna(0)
    )

    df["Сумма со скидкой, р."] = (
        pd.to_numeric(df["Сумма со скидкой, р."], errors="coerce")
        .fillna(0)
    )

    uploaded_file.seek(0)

    return df


# ==========================================================
# АГРЕГАЦИЯ ДАННЫХ
# ==========================================================
def aggregate_data(df: pd.DataFrame, rules_text: str) -> dict:
    """
    Стандартная агрегация:
    - количество;
    - сумма;
    - по каждому правилу;
    - по СПб и Тюмени.

    Логика размеров:
    - правило БЕЗ размера -> суммируются все размеры/веса позиции;
    - правило С размером ("150 гр", "30 см", "30") -> только совпадающие.
    """
    df = df.copy()

    df["Город"] = df["Юридическое лицо"].apply(map_city)
    df = df[df["Город"].notna()]

    df["Блюдо_norm"] = df["Блюдо"].apply(normalize_text)
    df["Размер"] = df["Блюдо"].apply(extract_size)

    rules = _parse_rules(rules_text)

    result = {
        CITY_SPB: {},
        CITY_TYUMEN: {},
    }

    for city in CITY_ORDER:
        city_data = df[df["Город"] == city]

        for rule in rules:
            rule_norm = normalize_text(rule)
            rule_size = extract_size(rule)
            rule_tokens = set(rule_norm.split())

            if rule_size:
                subset = city_data[city_data["Размер"] == rule_size]
            else:
                subset = city_data.copy()

            if rule_tokens:
                matches = subset[
                    subset["Блюдо_norm"].apply(
                        lambda dish_norm: rule_tokens.issubset(
                            set(dish_norm.split())
                        )
                    )
                ]
            else:
                matches = subset.iloc[0:0]

            qty = int(matches["Количество блюд"].sum())
            total_sum = round(
                float(matches["Сумма со скидкой, р."].sum()),
                2,
            )

            result[city][rule] = {
                "qty": qty,
                "sum": total_sum,
            }

    return result


def aggregate_pizza_by_size(df: pd.DataFrame, rules_text: str) -> dict:
    """
    Агрегация пицц по размерам:
    15, 23, 30, 35, 40 см.
    """
    df = df.copy()

    df["Город"] = df["Юридическое лицо"].apply(map_city)
    df = df[df["Город"].notna()]

    df["Блюдо_norm"] = df["Блюдо"].apply(normalize_text)
    df["Размер"] = df["Блюдо"].apply(extract_size)
    df["Название_пиццы"] = df["Блюдо"].apply(extract_pizza_name)
    df["Название_пиццы_norm"] = df["Название_пиццы"].apply(normalize_text)

    rules = _parse_rules(rules_text)

    result = {
        CITY_SPB: {},
        CITY_TYUMEN: {},
    }

    for city in CITY_ORDER:
        city_data = df[df["Город"] == city]

        for rule in rules:
            rule_norm = normalize_text(rule)
            rule_tokens = set(rule_norm.split())

            if rule_tokens:
                matches = city_data[
                    city_data["Название_пиццы_norm"].apply(
                        lambda pizza_norm: rule_tokens.issubset(
                            set(pizza_norm.split())
                        )
                    )
                ]
            else:
                matches = city_data.iloc[0:0]

            size_data = {}

            for size in PIZZA_SIZES:
                size_matches = matches[matches["Размер"] == size]

                qty = int(size_matches["Количество блюд"].sum())
                total_sum = round(
                    float(size_matches["Сумма со скидкой, р."].sum()),
                    2,
                )

                size_data[size] = {
                    "qty": qty,
                    "sum": total_sum,
                }

            total_qty = sum(
                size_data[size]["qty"]
                for size in PIZZA_SIZES
            )

            total_sum = round(
                sum(
                    size_data[size]["sum"]
                    for size in PIZZA_SIZES
                ),
                2,
            )

            size_data["Всего"] = {
                "qty": total_qty,
                "sum": total_sum,
            }

            result[city][rule] = size_data

    return result


# ==========================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ EXCEL
# ==========================================================
def _get_standard_values(data: dict, city: str, rule: str) -> dict:
    """
    Безопасно достаёт данные для стандартного отчёта.
    """
    return (
        data.get(city, {})
        .get(rule, {"qty": 0, "sum": 0.0})
    )


def _get_pizza_size_values(rule_data: dict, size: str) -> dict:
    """
    Безопасно достаёт данные по размеру пиццы.
    """
    return rule_data.get(size, {"qty": 0, "sum": 0.0})


# ==========================================================
# EXCEL: СТАНДАРТНЫЙ ОТЧЁТ
# ==========================================================
def create_excel(data: dict, period_str: str, rules: list[str]) -> Workbook:
    """
    Создаёт Excel для стандартного отчёта.
    """
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws = wb.create_sheet("Отчёт")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    yellow_fill = PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type="solid",
    )

    blue_fill = PatternFill(
        start_color="00BFFF",
        end_color="00BFFF",
        fill_type="solid",
    )

    # Период
    ws.merge_cells("A1:H1")
    c = ws.cell(1, 1, period_str)
    c.font = Font(bold=True, size=16)
    c.alignment = Alignment(horizontal="center")
    c.fill = yellow_fill

    # Города
    ws.merge_cells("A2:D2")
    c = ws.cell(2, 1, "СПб сейчас")
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="center")
    c.fill = blue_fill

    ws.merge_cells("E2:H2")
    c = ws.cell(2, 5, "Тюмень сейчас")
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="center")
    c.fill = blue_fill

    # Заголовки
    headers = [
        (1, "Наименование"),
        (2, "Количество"),
        (3, "Сумма, ₽"),
        (5, "Наименование"),
        (6, "Количество"),
        (7, "Сумма, ₽"),
    ]

    for col, header in headers:
        c = ws.cell(3, col, header)
        c.font = Font(bold=True, size=10)
        c.border = border
        c.alignment = Alignment(horizontal="center")

    if not rules:
        return wb

    # Данные
    row = 4

    for rule in rules:
        spb = _get_standard_values(data, CITY_SPB, rule)
        tyumen = _get_standard_values(data, CITY_TYUMEN, rule)

        # СПб
        ws.cell(row, 1, rule).border = border

        c = ws.cell(row, 2, spb["qty"])
        c.border = border
        c.alignment = Alignment(horizontal="center")

        c = ws.cell(row, 3, spb["sum"])
        c.border = border
        c.number_format = "#,##0.00"

        # Тюмень
        ws.cell(row, 5, rule).border = border

        c = ws.cell(row, 6, tyumen["qty"])
        c.border = border
        c.alignment = Alignment(horizontal="center")

        c = ws.cell(row, 7, tyumen["sum"])
        c.border = border
        c.number_format = "#,##0.00"

        row += 1

    last_data_row = row - 1
    total_row = row + 1

    # Итого
    for col in [1, 5]:
        c = ws.cell(total_row, col, "ИТОГО")
        c.font = Font(bold=True, size=11)
        c.border = border

    for col in [2, 6]:
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}4:{col_letter}{last_data_row})"

        c = ws.cell(total_row, col, formula)
        c.font = Font(bold=True, size=11)
        c.border = border
        c.alignment = Alignment(horizontal="center")

    for col in [3, 7]:
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}4:{col_letter}{last_data_row})"

        c = ws.cell(total_row, col, formula)
        c.font = Font(bold=True, size=11)
        c.border = border
        c.number_format = "#,##0.00"

    # Ширина колонок
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15

    return wb


# ==========================================================
# EXCEL: ПИЦЦЫ ПО РАЗМЕРАМ
# ==========================================================
def create_pizza_excel(data: dict, period_str: str, rules: list[str]) -> Workbook:
    """
    Создаёт Excel с разбивкой пицц по размерам.
    """
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws = wb.create_sheet("Отчёт")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    thick_border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )

    yellow_fill = PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type="solid",
    )

    blue_fill = PatternFill(
        start_color="00BFFF",
        end_color="00BFFF",
        fill_type="solid",
    )

    gray_fill = PatternFill(
        start_color="D3D3D3",
        end_color="D3D3D3",
        fill_type="solid",
    )

    # Заголовок периода
    ws.merge_cells("A1:Q1")
    c = ws.cell(1, 1, period_str)
    c.font = Font(bold=True, size=16)
    c.alignment = Alignment(horizontal="center")
    c.fill = yellow_fill

    # Заголовок города: СПБ
    ws.merge_cells("A2:G2")
    c = ws.cell(2, 1, "СПБ")
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")
    c.fill = blue_fill

    # Заголовок города: Тюмень
    ws.merge_cells("K2:Q2")
    c = ws.cell(2, 11, "Тюмень")
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")
    c.fill = blue_fill

    # Заголовки колонок
    headers = [
        "Пиццы",
        "15 см",
        "23 см",
        "30 см",
        "35 см",
        "40 см",
        "Всего",
    ]

    # СПБ
    for col, header in enumerate(headers, 1):
        c = ws.cell(3, col, header)
        c.font = Font(bold=True, size=11)
        c.border = thick_border
        c.alignment = Alignment(horizontal="center")
        c.fill = gray_fill

    # Тюмень
    for col, header in enumerate(headers, 11):
        c = ws.cell(3, col, header)
        c.font = Font(bold=True, size=11)
        c.border = thick_border
        c.alignment = Alignment(horizontal="center")
        c.fill = gray_fill

    if not rules:
        return wb

    # Данные
    row = 4

    for rule in rules:
        spb_rule = data.get(CITY_SPB, {}).get(rule, {})
        tyumen_rule = data.get(CITY_TYUMEN, {}).get(rule, {})

        # СПБ: название
        c = ws.cell(row, 1, rule)
        c.border = border
        c.alignment = Alignment(horizontal="left")

        # СПБ: размеры
        for col_idx, size in enumerate(PIZZA_SIZES, 2):
            qty = _get_pizza_size_values(spb_rule, size)["qty"]

            c = ws.cell(row, col_idx, qty)
            c.border = border
            c.alignment = Alignment(horizontal="center")

        # СПБ: всего
        total_qty_spb = spb_rule.get("Всего", {}).get(
            "qty",
            sum(
                _get_pizza_size_values(spb_rule, size)["qty"]
                for size in PIZZA_SIZES
            ),
        )

        c = ws.cell(row, 7, total_qty_spb)
        c.font = Font(bold=True)
        c.border = thick_border
        c.alignment = Alignment(horizontal="center")

        # Тюмень: название
        c = ws.cell(row, 11, rule)
        c.border = border
        c.alignment = Alignment(horizontal="left")

        # Тюмень: размеры
        for col_idx, size in enumerate(PIZZA_SIZES, 12):
            qty = _get_pizza_size_values(tyumen_rule, size)["qty"]

            c = ws.cell(row, col_idx, qty)
            c.border = border
            c.alignment = Alignment(horizontal="center")

        # Тюмень: всего
        total_qty_tyumen = tyumen_rule.get("Всего", {}).get(
            "qty",
            sum(
                _get_pizza_size_values(tyumen_rule, size)["qty"]
                for size in PIZZA_SIZES
            ),
        )

        c = ws.cell(row, 17, total_qty_tyumen)
        c.font = Font(bold=True)
        c.border = thick_border
        c.alignment = Alignment(horizontal="center")

        row += 1

    last_data_row = row - 1
    total_row = row

    # Итого: СПБ
    c = ws.cell(total_row, 1, "Итого")
    c.font = Font(bold=True, size=11)
    c.border = thick_border

    for col_idx in range(2, 7):
        col_letter = get_column_letter(col_idx)
        formula = f"=SUM({col_letter}4:{col_letter}{last_data_row})"

        c = ws.cell(total_row, col_idx, formula)
        c.font = Font(bold=True, size=11)
        c.border = thick_border
        c.alignment = Alignment(horizontal="center")

    c = ws.cell(total_row, 7, f"=SUM(G4:G{last_data_row})")
    c.font = Font(bold=True, size=11)
    c.border = thick_border
    c.alignment = Alignment(horizontal="center")

    # Итого: Тюмень
    c = ws.cell(total_row, 11, "Итого")
    c.font = Font(bold=True, size=11)
    c.border = thick_border

    for col_idx in range(12, 17):
        col_letter = get_column_letter(col_idx)
        formula = f"=SUM({col_letter}4:{col_letter}{last_data_row})"

        c = ws.cell(total_row, col_idx, formula)
        c.font = Font(bold=True, size=11)
        c.border = thick_border
        c.alignment = Alignment(horizontal="center")

    c = ws.cell(total_row, 17, f"=SUM(Q4:Q{last_data_row})")
    c.font = Font(bold=True, size=11)
    c.border = thick_border
    c.alignment = Alignment(horizontal="center")

    # Ширина колонок
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10

    # Пустые колонки между городами
    ws.column_dimensions["H"].width = 3
    ws.column_dimensions["I"].width = 3
    ws.column_dimensions["J"].width = 3

    ws.column_dimensions["K"].width = 45
    ws.column_dimensions["L"].width = 10
    ws.column_dimensions["M"].width = 10
    ws.column_dimensions["N"].width = 10
    ws.column_dimensions["O"].width = 10
    ws.column_dimensions["P"].width = 10
    ws.column_dimensions["Q"].width = 10

    return wb