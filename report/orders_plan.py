# report/orders_plan.py

"""
Расчёт плана по количеству заказов на произвольный целевой месяц/период,
по каждому ресторану сети.

Источник данных — помесячная история заказов по ресторанам (см. load_history).

Логика расчёта (для каждого целевого календарного месяца):
1. Основной метод — сезонный коэффициент:
     коэфф = заказы(тот же месяц, год назад) / заказы(предыдущий месяц, год назад)
     прогноз = заказы(последний доступный факт перед целевым месяцем) × коэфф
   Если данных "год назад" больше одного года — коэффициент считается как
   среднее по всем доступным годам для этого календарного месяца.
2. Проверочный метод — линейный тренд (OLS) по всей доступной истории,
   экстраполированный на целевой месяц.
3. Если для сезонного коэффициента нет данных (мало истории, новый ресторан,
   слишком длинный горизонт вперёд) — план строится по тренду (fallback).
4. Отклонение прогноза (метод 1) от тренда (метод 2) в % — при |отклонении|
   больше DEVIATION_WARNING_THRESHOLD ресторан помечается needs_review=True.

Модуль не зависит от Streamlit и не привязан к конкретному способу
получения исходного файла — на вход годится путь, file-like объект или
уже загруженный DataFrame (см. load_history).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from typing import Any, Iterable, List, Optional, Sequence, Union

import pandas as pd

# ==========================================================
# КОНСТАНТЫ
# ==========================================================

MONTHS_RU = [
    "январь", "февраль", "март", "апрель", "май", "июнь",
    "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
]

# "09 (Сентябрь) 2025", "08 (Август) 2025г", "9(сентябрь)2025" и т.п.
_HEADER_RE = re.compile(
    r"(?P<month_num>\d{1,2})\s*\(\s*[А-Яа-яЁё]+\s*\)\s*(?P<year>\d{4})"
)

RESTAURANT_COLUMN_LABELS = {"торговое предприятие", "ресторан", "точка"}
TOTAL_MARKERS = {"итого", "итог", "всего"}

# группа №1 — явный номер после "№"; группа №2 — число в самом названии
# (напр. "Тюмень 2. Орджоникидзе"); группа №3 — номера нет вовсе
_NUM_MARK_RE = re.compile(r"№\s*0*(\d+)")
_NUM_FALLBACK_RE = re.compile(r"(?<!\d)0*(\d+)(?!\d)")

DEVIATION_WARNING_THRESHOLD = 0.15  # 15%

TargetPeriod = Union[str, date, tuple, Sequence]


# ==========================================================
# ЗАГРУЗКА И НОРМАЛИЗАЦИЯ ИСТОРИИ
# ==========================================================

def parse_month_header(label: Any) -> Optional[date]:
    """
    Парсит подпись столбца-месяца (например "09 (Сентябрь) 2025" или
    "08 (Август) 2025г") в дату первого числа месяца.

    Год и номер месяца берутся из самой подписи, а не из положения
    столбца — порядок столбцов в файле может быть произвольным.
    """
    if not isinstance(label, str):
        return None
    match = _HEADER_RE.search(label)
    if not match:
        return None
    month_num = int(match.group("month_num"))
    year = int(match.group("year"))
    if not 1 <= month_num <= 12:
        return None
    return date(year, month_num, 1)


def _find_header_row(raw: pd.DataFrame) -> int:
    """Строка с наибольшим числом распознанных подписей-месяцев — шапка таблицы."""
    best_row, best_score = None, 0
    for idx in range(len(raw)):
        score = sum(parse_month_header(v) is not None for v in raw.iloc[idx])
        if score > best_score:
            best_row, best_score = idx, score
    if best_row is None:
        raise ValueError("Не найдена строка-шапка с подписями месяцев вида '09 (Сентябрь) 2025'.")
    return best_row


def load_history(source: Union[str, pd.DataFrame, Any], sheet_name: Union[int, str] = 0) -> pd.DataFrame:
    """
    Читает файл истории заказов (путь/буфер/DataFrame) и приводит его
    к длинному формату: restaurant | period (date, 1-е число месяца) | orders.

    Строки/столбцы "Итого" отбрасываются. Месяцы сортируются по реальной
    дате, а не по исходному порядку столбцов.
    """
    if isinstance(source, pd.DataFrame):
        raw = source
    else:
        raw = pd.read_excel(source, sheet_name=sheet_name, header=None)

    header_row = _find_header_row(raw)
    header = raw.iloc[header_row]

    month_columns: List[tuple] = []
    for col_idx, label in header.items():
        period = parse_month_header(label)
        if period is not None:
            month_columns.append((col_idx, period))
    if not month_columns:
        raise ValueError("В шапке файла не найдено ни одного столбца с месяцем.")

    restaurant_col = header.index[0]

    body = raw.iloc[header_row + 1:]
    records = []
    for _, row in body.iterrows():
        name = row[restaurant_col]
        if not isinstance(name, str) or not name.strip():
            continue
        if name.strip().lower() in TOTAL_MARKERS:
            continue
        for col_idx, period in month_columns:
            value = row[col_idx]
            if pd.isna(value):
                continue
            records.append({"restaurant": name.strip(), "period": period, "orders": float(value)})

    history = pd.DataFrame.from_records(records, columns=["restaurant", "period", "orders"])
    history = history.sort_values(["restaurant", "period"]).reset_index(drop=True)
    return history


# ==========================================================
# СОРТИРОВКА РЕСТОРАНОВ
# ==========================================================

def restaurant_sort_key(name: str):
    """
    Сначала рестораны с явным номером "№N" (по возрастанию N), затем
    рестораны без "№", но с числом в названии (напр. "Тюмень 2..."),
    тоже по возрастанию числа; всё остальное — в конец, по алфавиту.

    Правило легко расширяется: новый город/формат с числом в названии
    автоматически попадёт во вторую группу и встанет в ней по номеру.
    """
    match = _NUM_MARK_RE.search(name)
    if match:
        return (0, int(match.group(1)), name)
    match = _NUM_FALLBACK_RE.search(name)
    if match:
        return (1, int(match.group(1)), name)
    return (2, 0, name)


def sort_restaurants(names: Iterable[str]) -> List[str]:
    return sorted(set(names), key=restaurant_sort_key)


# ==========================================================
# ЦЕЛЕВОЙ ПЕРИОД
# ==========================================================

def shift_months(d: date, months: int) -> date:
    """Возвращает 1-е число месяца, отстоящего от d на `months` (может быть отрицательным)."""
    total = (d.year * 12 + (d.month - 1)) + months
    year, month = divmod(total, 12)
    return date(year, month + 1, 1)


_shift_months = shift_months


def _to_month_date(value: Any) -> date:
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    if isinstance(value, str):
        text = value.strip()
        if re.fullmatch(r"\d{4}-\d{1,2}", text):
            year, month = text.split("-")
            return date(int(year), int(month), 1)
        parsed = pd.to_datetime(text)
        return date(parsed.year, parsed.month, 1)
    if isinstance(value, tuple) and len(value) == 2:
        year, month = value
        return date(int(year), int(month), 1)
    raise ValueError(f"Не удалось разобрать целевой период: {value!r}")


def resolve_target_months(target_period: TargetPeriod) -> List[date]:
    """
    Приводит целевой период к списку дат (1-е число каждого месяца).

    Принимает: "2026-09", date(2026, 9, 1), (2026, 9), диапазон
    "2026-09:2026-11", (start, end) или список/множество месяцев.
    """
    if isinstance(target_period, str) and ":" in target_period:
        start_raw, end_raw = target_period.split(":", 1)
        start, end = _to_month_date(start_raw), _to_month_date(end_raw)
        months, cur = [], start
        while cur <= end:
            months.append(cur)
            cur = _shift_months(cur, 1)
        return months

    if isinstance(target_period, (list, set)):
        return sorted({_to_month_date(v) for v in target_period})

    if isinstance(target_period, tuple) and len(target_period) == 2 and all(
        isinstance(v, (date, str)) for v in target_period
    ):
        start, end = _to_month_date(target_period[0]), _to_month_date(target_period[1])
        months, cur = [], start
        while cur <= end:
            months.append(cur)
            cur = _shift_months(cur, 1)
        return months

    return [_to_month_date(target_period)]


# ==========================================================
# РАСЧЁТ ПРОГНОЗА ДЛЯ ОДНОГО РЕСТОРАНА / ОДНОГО МЕСЯЦА
# ==========================================================

@dataclass
class MonthForecast:
    restaurant: str
    target_period: date
    method: str  # "seasonal" | "trend_only" | "insufficient_data"
    base_period: Optional[date] = None
    base_orders: Optional[float] = None
    seasonal_coef: Optional[float] = None
    seasonal_years_used: int = 0
    forecast_seasonal: Optional[float] = None
    forecast_trend: Optional[float] = None
    deviation_pct: Optional[float] = None
    plan: Optional[float] = None
    needs_review: bool = False
    warning: Optional[str] = None

    def to_dict(self) -> dict:
        d = dict(self.__dict__)
        d["target_period"] = self.target_period.isoformat()
        d["base_period"] = self.base_period.isoformat() if self.base_period else None
        return d


def _seasonal_coefficient(series: pd.Series, target: date) -> tuple:
    """
    series: index=period(date) -> orders, для одного ресторана.
    Возвращает (коэфф или None, число учтённых лет).
    """
    ratios = []
    k = 1
    max_k = target.year - series.index.min().year + 1 if len(series) else 0
    while k <= max_k:
        same = _shift_months(target, -12 * k)
        prev = _shift_months(same, -1)
        if same in series.index and prev in series.index and series[prev]:
            ratios.append(series[same] / series[prev])
        k += 1
    if not ratios:
        return None, 0
    return sum(ratios) / len(ratios), len(ratios)


def _linear_trend(series: pd.Series, target: date) -> Optional[float]:
    """OLS-тренд по всей истории ресторана, экстраполированный на target."""
    if len(series) < 2:
        return None
    base = series.index.min()
    x = [(d.year - base.year) * 12 + (d.month - base.month) for d in series.index]
    y = list(series.values)
    n = len(x)
    mean_x = sum(x) / n
    mean_y = sum(y) / n
    denom = sum((xi - mean_x) ** 2 for xi in x)
    if denom == 0:
        return mean_y
    slope = sum((xi - mean_x) * (yi - mean_y) for xi, yi in zip(x, y)) / denom
    intercept = mean_y - slope * mean_x
    target_x = (target.year - base.year) * 12 + (target.month - base.month)
    return intercept + slope * target_x


def forecast_restaurant_month(
    series: pd.Series,
    restaurant: str,
    target: date,
    deviation_threshold: float = DEVIATION_WARNING_THRESHOLD,
) -> MonthForecast:
    """series: index=period(date) -> orders, только для одного ресторана, отсортирован."""
    past = series[series.index < target]
    base_period = past.index.max() if len(past) else None
    base_orders = float(past[base_period]) if base_period is not None else None

    trend = _linear_trend(series, target)
    coef, years_used = _seasonal_coefficient(series, target)

    if coef is not None and base_orders is not None:
        forecast_seasonal = base_orders * coef
        deviation_pct = None
        if trend:
            deviation_pct = (forecast_seasonal / trend - 1) * 100
        return MonthForecast(
            restaurant=restaurant,
            target_period=target,
            method="seasonal",
            base_period=base_period,
            base_orders=base_orders,
            seasonal_coef=coef,
            seasonal_years_used=years_used,
            forecast_seasonal=forecast_seasonal,
            forecast_trend=trend,
            deviation_pct=deviation_pct,
            plan=forecast_seasonal,
            needs_review=bool(deviation_pct is not None and abs(deviation_pct) > deviation_threshold * 100),
        )

    if trend is not None:
        return MonthForecast(
            restaurant=restaurant,
            target_period=target,
            method="trend_only",
            base_period=base_period,
            base_orders=base_orders,
            forecast_trend=trend,
            plan=trend,
            warning="Нет данных для сезонного коэффициента — план построен по тренду.",
        )

    return MonthForecast(
        restaurant=restaurant,
        target_period=target,
        method="insufficient_data",
        base_period=base_period,
        base_orders=base_orders,
        plan=base_orders,
        warning="Недостаточно истории для расчёта тренда или сезонности.",
    )


# ==========================================================
# ПУБЛИЧНЫЙ API
# ==========================================================

def build_plan(
    history: pd.DataFrame,
    target_period: TargetPeriod,
    restaurants: Optional[Iterable[str]] = None,
    deviation_threshold: float = DEVIATION_WARNING_THRESHOLD,
) -> dict:
    """
    history: DataFrame со столбцами restaurant | period(date) | orders — см. load_history().
    target_period: см. resolve_target_months() — один месяц, диапазон или список месяцев.
    restaurants: ограничить расчёт списком ресторанов (по умолчанию — все из history).

    Возвращает JSON-совместимый словарь:
      {
        "target_months": [...],
        "restaurants": [ {restaurant, months: [MonthForecast...], plan: сумма по периоду}, ... ],
        "total": {"plan": ..., "needs_review_count": ...},
      }
    """
    target_months = resolve_target_months(target_period)

    all_names = restaurants if restaurants is not None else history["restaurant"].unique()
    ordered_names = sort_restaurants(all_names)

    restaurant_results = []
    grand_total = 0.0
    needs_review_count = 0

    for name in ordered_names:
        series = (
            history.loc[history["restaurant"] == name]
            .drop_duplicates("period", keep="last")
            .set_index("period")["orders"]
            .sort_index()
        )

        month_forecasts = [
            forecast_restaurant_month(series, name, month, deviation_threshold)
            for month in target_months
        ]

        plan_sum = sum(mf.plan for mf in month_forecasts if mf.plan is not None)
        review_flag = any(mf.needs_review for mf in month_forecasts)
        needs_review_count += int(review_flag)
        grand_total += plan_sum

        restaurant_results.append({
            "restaurant": name,
            "plan": plan_sum,
            "needs_review": review_flag,
            "months": [mf.to_dict() for mf in month_forecasts],
        })

    return {
        "target_months": [m.isoformat() for m in target_months],
        "restaurants": restaurant_results,
        "total": {
            "plan": grand_total,
            "restaurant_count": len(ordered_names),
            "needs_review_count": needs_review_count,
        },
    }


def plan_to_dataframe(plan: dict) -> pd.DataFrame:
    """
    Разворачивает результат build_plan() в плоский DataFrame — по строке
    на пару (ресторан, целевой месяц). Удобно для предпросмотра/экспорта.
    """
    rows = []
    for entry in plan["restaurants"]:
        for month in entry["months"]:
            rows.append({"restaurant": entry["restaurant"], **month})
    df = pd.DataFrame(rows)
    if not df.empty:
        df["restaurant"] = pd.Categorical(
            df["restaurant"],
            categories=sort_restaurants(df["restaurant"].unique()),
            ordered=True,
        )
        df = df.sort_values(["restaurant", "target_period"]).reset_index(drop=True)
    return df


PLAN_COLUMN_LABELS = {
    "restaurant": "Ресторан",
    "target_period": "Месяц",
    "method": "Метод",
    "base_period": "База (месяц факта)",
    "base_orders": "База (факт), шт.",
    "seasonal_coef": "Коэфф. сезонности",
    "seasonal_years_used": "Лет учтено",
    "forecast_seasonal": "Прогноз по сезонности",
    "forecast_trend": "Тренд (проверка)",
    "deviation_pct": "Отклонение от тренда, %",
    "plan": "ПЛАН, шт.",
    "needs_review": "Проверить?",
    "warning": "Комментарий",
}


def export_plan_excel(plan: dict) -> bytes:
    """Сохраняет результат build_plan() в один xlsx-лист для скачивания на сайте."""
    from io import BytesIO

    df = plan_to_dataframe(plan)
    df = df.drop(columns=["restaurant"]).assign(restaurant=df["restaurant"].astype(str))
    df = df[list(PLAN_COLUMN_LABELS.keys())].rename(columns=PLAN_COLUMN_LABELS)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="План", index=False)
    buffer.seek(0)
    return buffer.getvalue()


MONTHS_RU_SHORT = [m[:3] for m in MONTHS_RU]


def _short_month_label(d: date) -> str:
    return f"{MONTHS_RU_SHORT[d.month - 1]}.{d.year % 100:02d}"


def _sheet_title(d: date) -> str:
    title = f"План_{MONTHS_RU[d.month - 1].capitalize()}_{d.year}"
    return title[:31]


def export_plan_workbook(
    history: pd.DataFrame,
    target_period: TargetPeriod,
    restaurants: Optional[Iterable[str]] = None,
    deviation_threshold: float = DEVIATION_WARNING_THRESHOLD,
) -> bytes:
    """
    Формирует xlsx-книгу с расчётом ПЛАНА, где коэффициент сезонности,
    прогноз, тренд, отклонение и сам план — реальные формулы Excel,
    ссылающиеся на исходные цифры (как в примере ручного расчёта),
    а не готовые числа. Правишь факт в листе "Данные" — план пересчитывается.

    Лист "Данные" — история заказов по ресторанам (месяц/ресторан/значение).
    По одному листу "План_<Месяц>_<Год>" на каждый целевой месяц:
      - Факт(мес-1, год) / Факт(целевой мес, год) — за каждый доступный год,
        и частный коэффициент по нему;
      - Коэфф. сезонности = среднее по всем частным коэффициентам;
      - База = последний факт перед целевым месяцем (LOOKUP по строке);
      - Прогноз по сезонности = База × Коэфф.;
      - Тренд = INTERCEPT+SLOPE×t по всей истории ресторана (проверка);
      - Отклонение = Прогноз/Тренд − 1;
      - ПЛАН = Прогноз, либо Тренд, если сезонный коэффициент посчитать
        не из чего (формула IF/IFERROR — тот же fallback, что и в build_plan()).
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    target_months = resolve_target_months(target_period)
    all_names = restaurants if restaurants is not None else history["restaurant"].unique()
    names = sort_restaurants(all_names)

    hist = history[history["restaurant"].isin(names)]
    if hist.empty:
        raise ValueError("Нет данных по выбранным ресторанам.")

    first_period = hist["period"].min()
    last_period = hist["period"].max()
    n_months = (last_period.year - first_period.year) * 12 + (last_period.month - first_period.month) + 1
    periods = [shift_months(first_period, i) for i in range(n_months)]
    period_col = {p: i + 2 for i, p in enumerate(periods)}  # col B.. в листе "Данные"

    pivot = hist.pivot_table(index="restaurant", columns="period", values="orders", aggfunc="last")

    wb = Workbook()
    bold = Font(bold=True)
    header_fill_align = Alignment(vertical="center", wrap_text=True)

    # ================= ЛИСТ "Данные" =================
    ws = wb.active
    ws.title = "Данные"

    total_col = len(periods) + 2
    ws.cell(row=1, column=1, value=(
        f"Динамика заказов по ресторанам, {_short_month_label(first_period)} — "
        f"{_short_month_label(last_period)} ({len(periods)} мес.)"
    )).font = bold
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_col)

    ws.cell(row=2, column=1, value="№ мес. (t) →").font = bold
    for p, col in period_col.items():
        t = periods.index(p) + 1
        ws.cell(row=2, column=col, value=t)

    ws.cell(row=3, column=1, value="Ресторан").font = bold
    for p, col in period_col.items():
        c = ws.cell(row=3, column=col, value=_short_month_label(p))
        c.font = bold
        c.alignment = header_fill_align
    ws.cell(row=3, column=total_col, value="Итого").font = bold

    restaurant_row = {}
    row = 4
    for name in names:
        restaurant_row[name] = row
        ws.cell(row=row, column=1, value=name)
        for p, col in period_col.items():
            val = pivot.loc[name, p] if name in pivot.index and p in pivot.columns else None
            if pd.notna(val):
                ws.cell(row=row, column=col, value=float(val))
        first_l, last_l = get_column_letter(2), get_column_letter(total_col - 1)
        ws.cell(row=row, column=total_col, value=f"=SUM({first_l}{row}:{last_l}{row})")
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="ИТОГО").font = bold
    for col in range(2, total_col + 1):
        letter = get_column_letter(col)
        ws.cell(
            row=total_row, column=col,
            value=f"=SUM({letter}{restaurant_row[names[0]]}:{letter}{total_row - 1})",
        ).font = bold

    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = 32
    for col in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    # ================= ЛИСТЫ "План_<Месяц>_<Год>" =================
    for target in target_months:
        _write_plan_sheet(
            wb, target, names, restaurant_row, periods, period_col,
            first_period, deviation_threshold, bold,
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _write_plan_sheet(wb, target, names, restaurant_row, periods, period_col,
                       first_period, deviation_threshold, bold):
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(_sheet_title(target))

    target_t = 1 + (target.year - first_period.year) * 12 + (target.month - first_period.month)
    last_grid_period = periods[-1] if periods else None
    pre_target_periods = [p for p in periods if p < target]
    pre_end_col = period_col[pre_target_periods[-1]] if pre_target_periods else None

    # Сколько годичных пар (Факт мес-1 год-k / Факт целевой мес год-k) реально есть в сетке.
    k_pairs = []
    k = 1
    while True:
        same = shift_months(target, -12 * k)
        prev = shift_months(same, -1)
        if same not in period_col or prev not in period_col:
            break
        k_pairs.append((k, prev, same))
        k += 1

    ws.cell(row=1, column=1, value=f"План по количеству заказов на {_short_month_label(target)} (t={target_t})").font = bold
    ws.cell(row=2, column=1, value=(
        "Коэфф. сезонности = среднее по годам (Факт целевого месяца / Факт предыдущего месяца, год назад); "
        "Прогноз = последний факт перед месяцем × коэфф.; Тренд — линейная регрессия по всей истории (проверка); "
        "ПЛАН = прогноз по сезонности, либо тренд, если сезонных данных нет."
    ))

    col = 1
    header_row = 4
    ws.cell(row=header_row, column=col, value="Ресторан").font = bold
    col += 1

    ratio_cols = []
    for k, prev_p, same_p in k_pairs:
        prev_col, same_col, ratio_col = col, col + 1, col + 2
        ws.cell(row=header_row, column=prev_col, value=f"Факт {_short_month_label(prev_p)}").font = bold
        ws.cell(row=header_row, column=same_col, value=f"Факт {_short_month_label(same_p)}").font = bold
        ws.cell(row=header_row, column=ratio_col, value=f"Коэфф. {same_p.year}").font = bold
        ratio_cols.append((prev_col, same_col, ratio_col, prev_p, same_p))
        col += 3

    coef_col = col
    ws.cell(row=header_row, column=coef_col, value="Коэфф. сезонности").font = bold
    col += 1
    base_col = col
    ws.cell(row=header_row, column=base_col, value="База (посл. факт)").font = bold
    col += 1
    forecast_col = col
    ws.cell(row=header_row, column=forecast_col, value="Прогноз по сезонности").font = bold
    col += 1
    trend_col = col
    ws.cell(row=header_row, column=trend_col, value="Тренд (проверка)").font = bold
    col += 1
    deviation_col = col
    ws.cell(row=header_row, column=deviation_col, value="Отклонение от тренда").font = bold
    col += 1
    plan_col = col
    ws.cell(row=header_row, column=plan_col, value="ПЛАН").font = bold
    col += 1
    review_col = col
    ws.cell(row=header_row, column=review_col, value="Проверить?").font = bold
    last_col = col

    row = header_row + 1
    for name in names:
        src_row = restaurant_row[name]
        ws.cell(row=row, column=1, value=name)

        for prev_col, same_col, ratio_col, prev_p, same_p in ratio_cols:
            prev_letter = get_column_letter(period_col[prev_p])
            same_letter = get_column_letter(period_col[same_p])
            ws.cell(row=row, column=prev_col, value=f"=Данные!{prev_letter}{src_row}")
            ws.cell(row=row, column=same_col, value=f"=Данные!{same_letter}{src_row}")
            ws.cell(
                row=row, column=ratio_col,
                value=f'=IFERROR({get_column_letter(same_col)}{row}/{get_column_letter(prev_col)}{row},"")',
            )

        coef_letter = get_column_letter(coef_col)
        if ratio_cols:
            ratio_letters = ",".join(f"{get_column_letter(rc[2])}{row}" for rc in ratio_cols)
            ws.cell(row=row, column=coef_col, value=f'=IFERROR(AVERAGE({ratio_letters}),"")')
        else:
            ws.cell(row=row, column=coef_col, value='=""')

        if pre_end_col is not None:
            data_first_letter = get_column_letter(2)
            data_last_letter = get_column_letter(pre_end_col)
            row_range = f"Данные!${data_first_letter}{src_row}:${data_last_letter}{src_row}"
            ws.cell(
                row=row, column=base_col,
                value=f'=IFERROR(LOOKUP(2,1/({row_range}<>""),{row_range}),"")',
            )
            t_first, t_last = get_column_letter(2), get_column_letter(pre_end_col)
            t_range = f"Данные!${t_first}$2:${t_last}$2"
            ws.cell(
                row=row, column=trend_col,
                value=(
                    f'=IFERROR(INTERCEPT({row_range},{t_range})'
                    f'+SLOPE({row_range},{t_range})*{target_t},"")'
                ),
            )
        else:
            ws.cell(row=row, column=base_col, value='=""')
            ws.cell(row=row, column=trend_col, value='=""')

        base_letter = get_column_letter(base_col)
        trend_letter = get_column_letter(trend_col)
        ws.cell(
            row=row, column=forecast_col,
            value=(
                f'=IF(OR({coef_letter}{row}="",{base_letter}{row}=""),"",'
                f'{base_letter}{row}*{coef_letter}{row})'
            ),
        )
        forecast_letter = get_column_letter(forecast_col)
        ws.cell(
            row=row, column=deviation_col,
            value=f'=IFERROR({forecast_letter}{row}/{trend_letter}{row}-1,"")',
        )
        deviation_letter = get_column_letter(deviation_col)
        ws.cell(
            row=row, column=plan_col,
            value=f'=IF({forecast_letter}{row}="",{trend_letter}{row},{forecast_letter}{row})',
        ).number_format = "0"
        ws.cell(row=row, column=deviation_col).number_format = "0.0%"
        ws.cell(row=row, column=coef_col).number_format = "0.000"
        for prev_col, same_col, ratio_col, *_ in ratio_cols:
            ws.cell(row=row, column=ratio_col).number_format = "0.000"

        plan_letter = get_column_letter(plan_col)
        ws.cell(
            row=row, column=review_col,
            value=(
                f'=IF({deviation_letter}{row}="","",'
                f'ABS({deviation_letter}{row})>{deviation_threshold})'
            ),
        )
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="ИТОГО").font = bold
    for c in (base_col, forecast_col, trend_col, plan_col):
        letter = get_column_letter(c)
        ws.cell(
            row=total_row, column=c,
            value=f'=SUM({letter}{header_row + 1}:{letter}{total_row - 1})',
        ).font = bold
    fc_letter, tr_letter, dev_letter = get_column_letter(forecast_col), get_column_letter(trend_col), get_column_letter(deviation_col)
    ws.cell(
        row=total_row, column=deviation_col,
        value=f'=IFERROR({fc_letter}{total_row}/{tr_letter}{total_row}-1,"")',
    ).number_format = "0.0%"
    ws.cell(row=total_row, column=plan_col).number_format = "0"

    ws.freeze_panes = get_column_letter(2) + str(header_row + 1)
    ws.column_dimensions["A"].width = 32
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13


if __name__ == "__main__":
    import json
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else "заказов.xlsx"
    target = sys.argv[2] if len(sys.argv) > 2 else "2026-09"

    hist = load_history(path)
    result = build_plan(hist, target)
    print(json.dumps(result, ensure_ascii=False, indent=2))
