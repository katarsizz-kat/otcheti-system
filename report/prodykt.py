import io
import re
import warnings
from typing import BinaryIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ==================== КОНСТАНТЫ ====================
CATEGORY_COLORS = {
    'пиццы': 'FFFFE0', 'закуски': 'F0FFF0', 'напитки': 'F0F8FF',
    'десерты': 'FFE4E1', 'комбо и радости': 'FFFFFF'
}

CATEGORY_MAPPING = {
    'тесто 23 трад': 'пиццы', 'тесто 23 тонкое': 'пиццы',
    'тесто 30 трад': 'пиццы', 'тесто 30 тонкое': 'пиццы',
    'тесто 35 трад': 'пиццы', 'тесто 35 тонкое': 'пиццы',
    'тесто 40 трад': 'пиццы', 'тесто 40 тонкое': 'пиццы',
    'тесто 15 трад': 'пиццы',
    'закуски': 'закуски', 'горячее': 'закуски', 'супы': 'закуски',
    'завтраки': 'завтраки', 'салаты': 'закуски',
    'напитки': 'напитки', 'пиво': 'напитки',
    'десерты': 'десерты', 'радости': 'комбо и радости',
    'дополнительно': 'дополнительно'
}

CATEGORIES_ORDER = ['пиццы', 'закуски', 'напитки', 'десерты', 'комбо и радости']
CITY_COLUMNS = {'СПБ': {'start': 1, 'end': 7}, 'Тюмень': {'start': 9, 'end': 15}}

PIZZA_CATEGORY_MAPPING = {
    7: ["большая бонанза", "итальянская с моцареллой и пепперони", "8 сыров", "любимая папина пицца"],
    6: ["баварская", "супер папа", "мясная", "маленькая италия", "чеддер и бекон", "4 сыра", "четыре сыра", "с чеддером и беконом"],
    5: ["альфредо", "цыпленок рэнч", "папа микс", "цыпленок барбекю", "мясное барбекю", "мексиканская"],
    4: ["цыплёнок флорентина", "гавайская", "двойная пепперони"],
    3: ["пепперони", "ветчина и грибы", "вегетарианская", "маргарита"],
    2: ["капричиоза", "чикен пармеджано", "чизбургер"],
    1: ["сырная пицца", "сырная", "пепперони грин", "цыплёнок грин", "деревенская", "домашняя", "нежная"],
    'Сезонные': ["тоскана", "мишка"]
}
PIZZA_CATEGORY_ORDER = [7, 6, 5, 4, 3, 2, 1, 'Сезонные']

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


# ==================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ====================
def _clean_dish_name(name):
    if not isinstance(name, str): return ""
    name = re.sub(r'\(.*?\)', '', name)
    return ' '.join(name.split()).strip()

def _is_half_pizza(name): return '+' in str(name)

def _reclassify_breakfast(name):
    name_lower = str(name).lower()
    if 'омлет' in name_lower: return 'закуски'
    if 'сырник' in name_lower or 'панкейк' in name_lower: return 'десерты'
    return 'завтраки'

def _map_city(legal_entity):
    le = str(legal_entity)
    if 'СПБ' in le: return 'СПБ'
    elif 'ПД' in le and 'СПБ' not in le: return 'Тюмень'
    return None

def _normalize_text(text):
    if not isinstance(text, str): return ""
    replacements = {'a': 'а', 'e': 'е', 'o': 'о', 'p': 'р', 'c': 'с', 'y': 'у', 'x': 'х',
                   'A': 'А', 'E': 'Е', 'O': 'О', 'P': 'Р', 'C': 'С', 'Y': 'У', 'X': 'Х', 'ё': 'е', 'Ё': 'Е'}
    for lat, cyr in replacements.items(): text = text.replace(lat, cyr)
    text = text.lower().replace('"', '')
    text = re.sub(r'\(.*?\)', '', text)
    text = re.sub(r'^пицца\s*', '', text)
    text = re.sub(r'\bпромо\b', '', text)
    text = re.sub(r'\bподарок\b', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def _find_category(pizza_name):
    normalized_pizza = _normalize_text(pizza_name)
    if normalized_pizza == "с чеддером и беконом": return 6
    for cat_num in PIZZA_CATEGORY_ORDER:
        if cat_num == 'Сезонные': continue
        for name in PIZZA_CATEGORY_MAPPING[cat_num]:
            if normalized_pizza == _normalize_text(name): return cat_num
    for cat_num in PIZZA_CATEGORY_ORDER:
        if cat_num == 'Сезонные': continue
        for name in PIZZA_CATEGORY_MAPPING[cat_num]:
            norm_name = _normalize_text(name)
            if norm_name in normalized_pizza and len(norm_name) > 3: return cat_num
    return 'Сезонные'

def _extract_size(category):
    for size in [15, 23, 30, 35, 40]:
        if str(size) in str(category): return size
    return None

def _classify_for_share(row):
    cat = str(row.get('Категория блюда', '')).strip().lower()
    name = str(row.get('Блюдо', '')).strip().lower()
    if 'тесто' in cat and 'половинки' not in cat: return 'Пиццы'
    if 'половинки' in cat: return 'Пиццы половинки'
    if cat in ['напитки', 'пиво']: return 'Напитки'
    if cat == 'десерты': return 'Десерты'
    if cat == 'завтраки':
        if 'омлет' in name: return 'Закуски'
        if 'сырник' in name or 'панкейк' in name: return 'Десерты'
        return None
    if cat in ['закуски', 'горячее', 'супы', 'салаты']: return 'Закуски'
    if cat == 'дополнительно':
        if 'доставк' in name or 'перчик' in name or 'мастер' in name: return None
        if 'соус' in name or 'мед цветочн' in name or 'сгущен' in name: return 'Соусы'
        return None
    return None


# ==================== ПАРСИНГ (PANDAS ETL) ====================
def _read_main_file_raw(uploaded_file: BinaryIO) -> pd.DataFrame:
    df = pd.read_excel(uploaded_file, header=None)
    header_row = None
    for idx, row in df.iterrows():
        vals = [str(v) for v in row.values if pd.notna(v)]
        if 'Юридическое лицо' in ' '.join(vals) and 'Блюдо' in ' '.join(vals):
            header_row = idx; break
    if header_row is None: raise ValueError("Не найдена строка заголовков в основном файле!")
    
    headers = df.iloc[header_row].values
    df = df[header_row + 1:].copy()
    df.columns = headers
    df = df.reset_index(drop=True)
    df = df[~df['Юридическое лицо'].astype(str).str.contains('OLAP|всего|Итого', na=False)]
    df = df.dropna(how='all')
    df['Юридическое лицо'] = df['Юридическое лицо'].ffill()
    df['Категория блюда'] = df['Категория блюда'].ffill()
    df = df[~df['Блюдо'].astype(str).str.lower().str.contains('персонал', na=False)]
    df = df[df['Блюдо'].notna()]
    df = df[df['Блюдо'].astype(str).str.strip() != '']
    df['Количество блюд'] = pd.to_numeric(df['Количество блюд'], errors='coerce').fillna(0)
    df['Сумма со скидкой, р.'] = pd.to_numeric(df['Сумма со скидкой, р.'], errors='coerce').fillna(0)
    df['Чеков'] = pd.to_numeric(df['Чеков'], errors='coerce').fillna(0)
    return df

def _parse_main_file(uploaded_file: BinaryIO) -> pd.DataFrame:
    df = _read_main_file_raw(uploaded_file)
    df = df[~df['Блюдо'].apply(_is_half_pizza)]
    df['Категория_отчёт'] = df['Категория блюда'].astype(str).str.strip().str.lower().map(CATEGORY_MAPPING)
    df = df[df['Категория_отчёт'].notna()]
    df['Блюдо_очищенное'] = df['Блюдо'].apply(_clean_dish_name)
    
    mask_breakfast = df['Категория_отчёт'] == 'завтраки'
    df.loc[mask_breakfast, 'Категория_отчёт'] = df.loc[mask_breakfast, 'Блюдо_очищенное'].apply(_reclassify_breakfast)
    df.loc[df['Блюдо_очищенное'].str.lower().str.contains('перчик', na=False), 'Категория_отчёт'] = 'закуски'
    
    return df.groupby(['Юридическое лицо', 'Категория_отчёт', 'Блюдо_очищенное']).agg({
        'Количество блюд': 'sum', 'Сумма со скидкой, р.': 'sum', 'Чеков': 'sum'
    }).reset_index()

def _parse_combo_file(uploaded_file: BinaryIO) -> pd.DataFrame:
    df = pd.read_excel(uploaded_file, header=None)
    header_row = None
    for idx, row in df.iterrows():
        vals = [str(v) for v in row.values if pd.notna(v)]
        if 'Юридическое лицо' in ' '.join(vals) and 'Комбо' in ' '.join(vals):
            header_row = idx; break
    if header_row is None: raise ValueError("Не найдена строка заголовков в файле комбо!")
    
    headers = df.iloc[header_row].values
    df = df[header_row + 1:].copy()
    df.columns = headers
    df = df.reset_index(drop=True)
    df = df.dropna(how='all')
    df['Юридическое лицо'] = df['Юридическое лицо'].ffill()
    df = df[~df['Юридическое лицо'].astype(str).str.contains('всего|Итого', na=False)]
    df = df[~df['Название Комбо'].astype(str).str.lower().str.contains('персонал|мастер', na=False)]
    df = df[df['Название Комбо'].notna()]
    df = df[df['Название Комбо'].astype(str).str.strip() != '']
    df['Количество Комбо'] = pd.to_numeric(df['Количество Комбо'], errors='coerce').fillna(0)
    df['Сумма со скидкой, р.'] = pd.to_numeric(df['Сумма со скидкой, р.'], errors='coerce').fillna(0)
    df['Чеков'] = pd.to_numeric(df['Чеков'], errors='coerce').fillna(0)
    df = df.rename(columns={'Название Комбо': 'Блюдо_очищенное', 'Количество Комбо': 'Количество блюд'})
    df['Блюдо_очищенное'] = df['Блюдо_очищенное'].apply(_clean_dish_name)
    df['Категория_отчёт'] = 'комбо и радости'
    df['Город'] = df['Юридическое лицо'].apply(_map_city)
    df = df[df['Город'].notna()]
    
    return df.groupby(['Город', 'Категория_отчёт', 'Блюдо_очищенное']).agg({
        'Количество блюд': 'sum', 'Сумма со скидкой, р.': 'sum', 'Чеков': 'sum'
    }).reset_index()


# ==================== ГЕНЕРАЦИЯ ЛИСТОВ EXCEL (OPENPYXL) ====================
def _create_rating_sheet(wb, df_main, df_combo):
    ws = wb.create_sheet('Рейтинг')
    for idx, cat in enumerate(CATEGORIES_ORDER, 1):
        cell = ws.cell(idx, 1, cat)
        cell.font = Font(bold=True, size=11)
        color = CATEGORY_COLORS.get(cat, 'FFFFFF')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    
    ws.cell(7, 1, 'СПБ').font = Font(bold=True, size=12)
    ws.cell(7, 9, 'Тюмень').font = Font(bold=True, size=12)
    
    headers = ['Наименование', 'Кол-во чеков', 'Рейтинг', 'Сумма со скидкой р.', 'Доля', 'Кол-во блюд', 'Доля']
    for idx, header in enumerate(headers, 1):
        for city_col in [idx, idx + 8]:
            cell = ws.cell(8, city_col, header)
            cell.font = Font(bold=True, size=10)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    df_main_rating = df_main.copy()
    df_main_rating['Город'] = df_main_rating['Юридическое лицо'].apply(_map_city)
    df_main_rating = df_main_rating[df_main_rating['Город'].notna()]
    df_main_rating = df_main_rating[df_main_rating['Категория_отчёт'].isin(CATEGORIES_ORDER + ['завтраки'])]
    
    all_data = pd.concat([df_main_rating, df_combo], ignore_index=True)
    
    city_items = {}
    max_items = 0
    for city in ['СПБ', 'Тюмень']:
        city_data = all_data[all_data['Город'] == city].copy()
        total_checks = city_data['Чеков'].sum()
        city_data['Рейтинг'] = city_data['Чеков'] / total_checks if total_checks > 0 else 0
        city_data = city_data.sort_values('Рейтинг', ascending=False).reset_index(drop=True)
        items = [{'category': r['Категория_отчёт'], 'name': r['Блюдо_очищенное'], 
                  'checks': int(r['Чеков']), 'sum': round(r['Сумма со скидкой, р.'], 2), 
                  'dishes': int(r['Количество блюд'])} for _, r in city_data.iterrows()]
        city_items[city] = items
        max_items = max(max_items, len(items))
    
    data_start_row = 9
    for idx in range(max_items):
        row = data_start_row + idx
        for city in ['СПБ', 'Тюмень']:
            if idx < len(city_items[city]):
                item = city_items[city][idx]
                cols = CITY_COLUMNS[city]
                color = CATEGORY_COLORS.get(item['category'], 'FFFFFF')
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
                ws.cell(row, cols['start']).value = item['name']
                ws.cell(row, cols['start'] + 1).value = item['checks']
                
                total_row = data_start_row + max_items + 3
                cl_checks = get_column_letter(cols['start'] + 1)
                c = ws.cell(row, cols['start'] + 2)
                c.value = f"={cl_checks}{row}/{cl_checks}${total_row}"
                c.number_format = '0.00%'
                
                ws.cell(row, cols['start'] + 3).value = item['sum']
                cl_sum = get_column_letter(cols['start'] + 3)
                c = ws.cell(row, cols['start'] + 4)
                c.value = f"={cl_sum}{row}/{cl_sum}${total_row}"
                c.number_format = '0.00%'
                
                ws.cell(row, cols['start'] + 5).value = item['dishes']
                cl_dishes = get_column_letter(cols['start'] + 5)
                c = ws.cell(row, cols['start'] + 6)
                c.value = f"={cl_dishes}{row}/{cl_dishes}${total_row}"
                c.number_format = '0.00%'
                
                for col in range(cols['start'], cols['end'] + 1):
                    cell = ws.cell(row, col)
                    cell.fill = fill
                    cell.border = THIN_BORDER
    
    total_row = data_start_row + max_items + 3
    for city in ['СПБ', 'Тюмень']:
        cols = CITY_COLUMNS[city]
        ws.cell(total_row, cols['start']).value = 'Итого' if city == 'СПБ' else 'Итого:'
        first_data_row, last_data_row = data_start_row, data_start_row + max_items - 1
        for offset in [1, 3, 5]:
            cl = get_column_letter(cols['start'] + offset)
            ws.cell(total_row, cols['start'] + offset).value = f"=SUM({cl}{first_data_row}:{cl}{last_data_row})"
        for offset in [2, 4, 6]:
            cl = get_column_letter(cols['start'] + offset)
            c = ws.cell(total_row, cols['start'] + offset)
            c.value = f"=SUM({cl}{first_data_row}:{cl}{last_data_row})"
            c.number_format = '0.00%'
        for col in range(cols['start'], cols['end'] + 1):
            cell = ws.cell(total_row, col)
            cell.font = Font(bold=True, size=11)
            cell.border = THIN_BORDER

def _create_pizza_sheet(wb, df_main_raw):
    ws = wb.create_sheet("Пиццы")
    spb_data = df_main_raw[df_main_raw['Юридическое лицо'].astype(str).str.contains('СПБ', na=False)].copy()
    tyumen_data = df_main_raw[
        df_main_raw['Юридическое лицо'].astype(str).str.contains('ООО "ПД"', na=False) & 
        ~df_main_raw['Юридическое лицо'].astype(str).str.contains('СПБ', na=False)
    ].copy()
    
    def aggregate_data(df):
        result = {}
        for _, row in df.iterrows():
            if '+' in str(row['Блюдо']): continue
            size = _extract_size(str(row['Категория блюда']))
            if size is None: continue
            normalized_name = _normalize_text(str(row['Блюдо'])).title()
            cat_num = _find_category(str(row['Блюдо']))
            key = (cat_num, normalized_name)
            if key not in result: result[key] = {15: 0, 23: 0, 30: 0, 35: 0, 40: 0}
            result[key][size] += int(float(row['Количество блюд']))
        return result
    
    def create_table(df_data, title, start_col, start_row):
        result = aggregate_data(df_data)
        sorted_data = []
        for cat_num in PIZZA_CATEGORY_ORDER:
            cat_items = [(k, v) for k, v in result.items() if k[0] == cat_num]
            cat_items.sort(key=lambda x: sum(x[1].values()), reverse=True)
            sorted_data.extend(cat_items)
        
        rows = [{'Категория': k[0], 'Пицца': k[1], **{f'{s}см': v.get(s, 0) for s in [15,23,30,35,40]}} for k, v in sorted_data]
        df = pd.DataFrame(rows)
        if not df.empty: df['Всего'] = df[[f'{s}см' for s in [15,23,30,35,40]]].sum(axis=1)
        
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+8)
        cell = ws.cell(row=start_row, column=start_col, value=title)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        headers = ['Категория', 'Пиццы', '15 см', '23 см', '30 см', '35 см', '40 см', 'Всего', 'Рейтинг продукта']
        for col, header in enumerate(headers, start_col):
            cell = ws.cell(row=start_row+1, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
        
        data_start = start_row + 2
        current_cat, cat_start_row = None, None
        
        for idx, (_, row_data) in enumerate(df.iterrows()):
            if idx >= 36: break
            r = data_start + idx
            cat_num = row_data['Категория']
            
            if current_cat != cat_num:
                if current_cat is not None and cat_start_row < r - 1:
                    ws.merge_cells(start_row=cat_start_row, start_column=start_col, end_row=r-1, end_column=start_col)
                    ws.cell(row=cat_start_row, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
                current_cat, cat_start_row = cat_num, r
            
            ws.cell(row=r, column=start_col, value=cat_num).border = THIN_BORDER
            ws.cell(row=r, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r, column=start_col+1, value=row_data['Пицца']).border = THIN_BORDER
            ws.cell(row=r, column=start_col+1).alignment = Alignment(horizontal='left', vertical='center')
            
            for col_idx, size in enumerate([15, 23, 30, 35, 40], start_col+2):
                ws.cell(row=r, column=col_idx, value=row_data.get(f'{size}см', 0)).border = THIN_BORDER
                ws.cell(row=r, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
            
            c = ws.cell(row=r, column=start_col+7)
            cols_sum = '+'.join([f"{get_column_letter(start_col+i)}{r}" for i in range(2, 7)])
            c.value = f"={cols_sum}"
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal='center', vertical='center')
            
            total_row = start_row + 40
            c = ws.cell(row=r, column=start_col+8)
            c.value = f"={get_column_letter(start_col+7)}{r}/{get_column_letter(start_col+7)}{total_row}"
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.number_format = '0.00%'
        
        if current_cat is not None:
            last_data_row = data_start + min(len(df), 36) - 1
            if cat_start_row < last_data_row:
                ws.merge_cells(start_row=cat_start_row, start_column=start_col, end_row=last_data_row, end_column=start_col)
        
        total_row = start_row + 40
        ws.cell(row=total_row, column=start_col+1, value='Всего пицц').font = Font(bold=True, size=11)
        ws.cell(row=total_row, column=start_col+1).border = THIN_BORDER
        for col in range(start_col+2, start_col+9):
            cl = get_column_letter(col)
            cell = ws.cell(row=total_row, column=col, value=f"=SUM({cl}{data_start}:{cl}{data_start+35})")
            cell.font = Font(bold=True, size=11)
            cell.border = THIN_BORDER
    
    create_table(spb_data, "СПБ", 1, 1)
    create_table(tyumen_data, "Тюмень", 11, 1)

def _create_combo_sheet(wb, df_combo):
    ws = wb.create_sheet("Комбо")
    ws.cell(1, 1, 'СПБ').font = Font(bold=True, size=12)
    ws.cell(1, 6, 'Тюмень').font = Font(bold=True, size=12)
    
    headers = ['Блюдо', 'Количество блюд', 'Сумма со скидкой, р.', '% от числа всех комбо']
    for idx, header in enumerate(headers, 1):
        for col in [idx, idx + 5]:
            cell = ws.cell(2, col, header)
            cell.font = Font(bold=True, size=10)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    city_data_1 = {}
    for city in ['СПБ', 'Тюмень']:
        city_df = df_combo[df_combo['Город'] == city].copy()
        total_qty = city_df['Количество блюд'].sum()
        city_df['%_qty'] = city_df['Количество блюд'] / total_qty if total_qty > 0 else 0
        city_df = city_df.sort_values('%_qty', ascending=False).reset_index(drop=True)
        city_data_1[city] = city_df
    
    max_items_1 = max(len(city_data_1['СПБ']), len(city_data_1['Тюмень']))
    total_row_1 = 3 + max_items_1 + 2
    
    for idx in range(max_items_1):
        row = 3 + idx
        for city, start_col in [('СПБ', 1), ('Тюмень', 6)]:
            if idx < len(city_data_1[city]):
                item = city_data_1[city].iloc[idx]
                ws.cell(row, start_col, item['Блюдо_очищенное']).border = THIN_BORDER
                ws.cell(row, start_col + 1, int(item['Количество блюд'])).border = THIN_BORDER
                ws.cell(row, start_col + 2, round(item['Сумма со скидкой, р.'], 2)).border = THIN_BORDER
                formula = f"={get_column_letter(start_col+1)}{row}/${get_column_letter(start_col+1)}${total_row_1}"
                c = ws.cell(row, start_col + 3, formula)
                c.border = THIN_BORDER
                c.number_format = '0.00%'
    
    for city, start_col in [('СПБ', 1), ('Тюмень', 6)]:
        col_b, col_c, col_d = start_col + 1, start_col + 2, start_col + 3
        total_cell = ws.cell(total_row_1, start_col, 'Итого' if city == 'СПБ' else 'Итого:')
        total_cell.font = Font(bold=True)
        total_cell.border = THIN_BORDER
        ws.cell(total_row_1, col_b, f"=SUM({get_column_letter(col_b)}3:{get_column_letter(col_b)}{total_row_1-3})").font = Font(bold=True)
        ws.cell(total_row_1, col_b).border = THIN_BORDER
        ws.cell(total_row_1, col_c).border = THIN_BORDER
        c = ws.cell(total_row_1, col_d, f"=SUM({get_column_letter(col_d)}3:{get_column_letter(col_d)}{total_row_1-3})")
        c.font = Font(bold=True)
        c.border = THIN_BORDER
        c.number_format = '0.00%'
    
    # Вторая таблица (% от выручки)
    start_row_2 = total_row_1 + 3
    ws.cell(start_row_2, 1, 'СПБ').font = Font(bold=True, size=12)
    ws.cell(start_row_2, 6, 'Тюмень').font = Font(bold=True, size=12)
    headers_2 = ['Блюдо', 'Количество блюд', 'Сумма со скидкой, р.', '% от выручки всех комбо']
    for idx, header in enumerate(headers_2, 1):
        for col in [idx, idx + 5]:
            cell = ws.cell(start_row_2 + 1, col, header)
            cell.font = Font(bold=True, size=10)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    city_data_2 = {}
    for city in ['СПБ', 'Тюмень']:
        city_df = df_combo[df_combo['Город'] == city].copy()
        total_sum = city_df['Сумма со скидкой, р.'].sum()
        city_df['%_sum'] = city_df['Сумма со скидкой, р.'] / total_sum if total_sum > 0 else 0
        city_df = city_df.sort_values('%_sum', ascending=False).reset_index(drop=True)
        city_data_2[city] = city_df
    
    max_items_2 = max(len(city_data_2['СПБ']), len(city_data_2['Тюмень']))
    data_start_2 = start_row_2 + 2
    total_row_2 = data_start_2 + max_items_2 + 2
    
    for idx in range(max_items_2):
        row = data_start_2 + idx
        for city, start_col in [('СПБ', 1), ('Тюмень', 6)]:
            if idx < len(city_data_2[city]):
                item = city_data_2[city].iloc[idx]
                ws.cell(row, start_col, item['Блюдо_очищенное']).border = THIN_BORDER
                ws.cell(row, start_col + 1, int(item['Количество блюд'])).border = THIN_BORDER
                ws.cell(row, start_col + 2, round(item['Сумма со скидкой, р.'], 2)).border = THIN_BORDER
                formula = f"={get_column_letter(start_col+2)}{row}/${get_column_letter(start_col+2)}${total_row_2}"
                c = ws.cell(row, start_col + 3, formula)
                c.border = THIN_BORDER
                c.number_format = '0.00%'
    
    for city, start_col in [('СПБ', 1), ('Тюмень', 6)]:
        col_b, col_c, col_d = start_col + 1, start_col + 2, start_col + 3
        total_cell = ws.cell(total_row_2, start_col, 'Итого' if city == 'СПБ' else 'Итого:')
        total_cell.font = Font(bold=True)
        total_cell.border = THIN_BORDER
        ws.cell(total_row_2, col_b).border = THIN_BORDER
        ws.cell(total_row_2, col_c, f"=SUM({get_column_letter(col_c)}{data_start_2}:{get_column_letter(col_c)}{total_row_2-3})").font = Font(bold=True)
        ws.cell(total_row_2, col_c).border = THIN_BORDER
        c = ws.cell(total_row_2, col_d, f"=SUM({get_column_letter(col_d)}{data_start_2}:{get_column_letter(col_d)}{total_row_2-3})")
        c.font = Font(bold=True)
        c.border = THIN_BORDER
        c.number_format = '0.00%'

def _create_category_share_sheet(wb, df_main_raw, df_combo):
    ws = wb.create_sheet("Доля категорий")
    CATEGORIES_SHARE = ['Десерты', 'Закуски', 'Напитки', 'Пиццы', 'Пиццы половинки', 'Сеты', 'Соусы']
    
    df_main_share = df_main_raw.copy()
    df_main_share['Город'] = df_main_share['Юридическое лицо'].apply(_map_city)
    df_main_share = df_main_share[df_main_share['Город'].notna()]
    df_main_share['Категория_доля'] = df_main_share.apply(_classify_for_share, axis=1)
    df_main_share = df_main_share[df_main_share['Категория_доля'].notna()]
    
    df_combo_share = df_combo.copy()
    df_combo_share['Категория_доля'] = 'Сеты'
    
    df_all_share = pd.concat([df_main_share, df_combo_share], ignore_index=True)
    grouped = df_all_share.groupby(['Город', 'Категория_доля']).agg({'Количество блюд': 'sum', 'Сумма со скидкой, р.': 'sum'}).reset_index()
    
    city_totals = {}
    for city in ['СПБ', 'Тюмень']:
        city_data = grouped[grouped['Город'] == city]
        city_totals[city] = {'qty': city_data['Количество блюд'].sum(), 'sum': city_data['Сумма со скидкой, р.'].sum()}
    
    city_category_data = {}
    for city in ['СПБ', 'Тюмень']:
        city_data = grouped[grouped['Город'] == city]
        total_qty, total_sum = city_totals[city]['qty'], city_totals[city]['sum']
        cat_data = {}
        for cat in CATEGORIES_SHARE:
            cat_row = city_data[city_data['Категория_доля'] == cat]
            qty = int(cat_row['Количество блюд'].sum()) if len(cat_row) > 0 else 0
            sum_val = round(cat_row['Сумма со скидкой, р.'].sum(), 2) if len(cat_row) > 0 else 0
            pct_qty = qty / total_qty if total_qty > 0 else 0
            pct_sum = sum_val / total_sum if total_sum > 0 else 0
            cat_data[cat] = {'qty': qty, 'sum': sum_val, 'pct_qty': pct_qty, 'pct_sum': pct_sum}
        city_category_data[city] = cat_data
    
    sorted_cats_qty = sorted(CATEGORIES_SHARE, key=lambda cat: city_category_data['СПБ'][cat]['pct_qty'], reverse=True)
    
    ws.cell(1, 1, 'СПБ').font = Font(bold=True, size=12)
    ws.cell(1, 6, 'Тюмень').font = Font(bold=True, size=12)
    headers_1 = ['Категория', 'кол-во', 'выручка', '%']
    for idx, header in enumerate(headers_1, 1):
        for col in [idx, idx + 5]:
            cell = ws.cell(2, col, header)
            cell.font = Font(bold=True, size=10)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for idx, cat in enumerate(sorted_cats_qty):
        row = 3 + idx
        for city, start_col in [('СПБ', 1), ('Тюмень', 6)]:
            data = city_category_data[city][cat]
            ws.cell(row, start_col, cat).border = THIN_BORDER
            ws.cell(row, start_col + 1, data['qty']).border = THIN_BORDER
            ws.cell(row, start_col + 2, data['sum']).border = THIN_BORDER
            total_row_1 = 3 + len(CATEGORIES_SHARE) + 2
            formula = f"={get_column_letter(start_col+1)}{row}/${get_column_letter(start_col+1)}${total_row_1}"
            c = ws.cell(row, start_col + 3, formula)
            c.border = THIN_BORDER
            c.number_format = '0.00%'
    
    total_row_1 = 3 + len(CATEGORIES_SHARE) + 2
    for city, start_col in [('СПБ', 1), ('Тюмень', 6)]:
        col_b, col_c, col_d = start_col + 1, start_col + 2, start_col + 3
        ws.cell(total_row_1, start_col, 'Итого:').font = Font(bold=True)
        ws.cell(total_row_1, start_col).border = THIN_BORDER
        ws.cell(total_row_1, col_b, f"=SUM({get_column_letter(col_b)}3:{get_column_letter(col_b)}{total_row_1-1})").font = Font(bold=True)
        ws.cell(total_row_1, col_b).border = THIN_BORDER
        ws.cell(total_row_1, col_c).border = THIN_BORDER
        c = ws.cell(total_row_1, col_d, f"=SUM({get_column_letter(col_d)}3:{get_column_letter(col_d)}{total_row_1-1})")
        c.font = Font(bold=True)
        c.border = THIN_BORDER
        c.number_format = '0.00%'
    
    # Вторая таблица (по выручке)
    start_row_2 = total_row_1 + 3
    ws.cell(start_row_2, 1, 'СПБ').font = Font(bold=True, size=12)
    ws.cell(start_row_2, 6, 'Тюмень').font = Font(bold=True, size=12)
    for idx, header in enumerate(headers_1, 1):
        for col in [idx, idx + 5]:
            cell = ws.cell(start_row_2 + 1, col, header)
            cell.font = Font(bold=True, size=10)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    sorted_cats_sum = sorted(CATEGORIES_SHARE, key=lambda cat: city_category_data['СПБ'][cat]['pct_sum'], reverse=True)
    data_start_2 = start_row_2 + 2
    
    for idx, cat in enumerate(sorted_cats_sum):
        row = data_start_2 + idx
        for city, start_col in [('СПБ', 1), ('Тюмень', 6)]:
            data = city_category_data[city][cat]
            ws.cell(row, start_col, cat).border = THIN_BORDER
            ws.cell(row, start_col + 1, data['qty']).border = THIN_BORDER
            ws.cell(row, start_col + 2, data['sum']).border = THIN_BORDER
            total_row_2 = data_start_2 + len(CATEGORIES_SHARE) + 2
            formula = f"={get_column_letter(start_col+2)}{row}/${get_column_letter(start_col+2)}${total_row_2}"
            c = ws.cell(row, start_col + 3, formula)
            c.border = THIN_BORDER
            c.number_format = '0.00%'
    
    total_row_2 = data_start_2 + len(CATEGORIES_SHARE) + 2
    for city, start_col in [('СПБ', 1), ('Тюмень', 6)]:
        col_b, col_c, col_d = start_col + 1, start_col + 2, start_col + 3
        ws.cell(total_row_2, start_col, 'Итого:').font = Font(bold=True)
        ws.cell(total_row_2, start_col).border = THIN_BORDER
        ws.cell(total_row_2, col_b).border = THIN_BORDER
        ws.cell(total_row_2, col_c, f"=SUM({get_column_letter(col_c)}{data_start_2}:{get_column_letter(col_c)}{total_row_2-1})").font = Font(bold=True)
        ws.cell(total_row_2, col_c).border = THIN_BORDER
        c = ws.cell(total_row_2, col_d, f"=SUM({get_column_letter(col_d)}{data_start_2}:{get_column_letter(col_d)}{total_row_2-1})")
        c.font = Font(bold=True)
        c.border = THIN_BORDER
        c.number_format = '0.00%'

def _create_category_dynamics_sheet(wb, df_main_raw, df_combo):
    ws = wb.create_sheet("Доля категорий динамика")
    CATEGORIES_DYN = ['Десерты', 'Закуски', 'Напитки', 'Пиццы', 'Пиццы половинки', 'Сеты', 'Соусы']
    
    df_main_dyn = df_main_raw.copy()
    df_main_dyn['Город'] = df_main_dyn['Юридическое лицо'].apply(_map_city)
    df_main_dyn = df_main_dyn[df_main_dyn['Город'].notna()]
    df_main_dyn['Категория_доля'] = df_main_dyn.apply(_classify_for_share, axis=1)
    df_main_dyn = df_main_dyn[df_main_dyn['Категория_доля'].notna()]
    
    df_combo_dyn = df_combo.copy()
    df_combo_dyn['Категория_доля'] = 'Сеты'
    
    df_all_dyn = pd.concat([df_main_dyn, df_combo_dyn], ignore_index=True)
    grouped = df_all_dyn.groupby(['Город', 'Категория_доля']).agg({'Количество блюд': 'sum', 'Сумма со скидкой, р.': 'sum'}).reset_index()
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    cell = ws.cell(1, 1, "Июль 2026г.")
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    ws.cell(3, 1, 'СПБ').font = Font(bold=True, size=12)
    ws.cell(3, 1, 'СПБ').fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    headers = ['Категория', 'Доля по кол-ву %', 'Доля по выручке%']
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(4, idx, header)
        cell.font = Font(bold=True, size=10)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    spb_data = grouped[grouped['Город'] == 'СПБ']
    total_qty_spb = spb_data['Количество блюд'].sum()
    total_sum_spb = spb_data['Сумма со скидкой, р.'].sum()
    for idx, cat in enumerate(CATEGORIES_DYN):
        row = 5 + idx
        cat_row = spb_data[spb_data['Категория_доля'] == cat]
        qty = int(cat_row['Количество блюд'].sum()) if len(cat_row) > 0 else 0
        sum_val = round(cat_row['Сумма со скидкой, р.'].sum(), 2) if len(cat_row) > 0 else 0
        pct_qty = qty / total_qty_spb if total_qty_spb > 0 else 0
        pct_sum = sum_val / total_sum_spb if total_sum_spb > 0 else 0
        ws.cell(row, 1, cat).border = THIN_BORDER
        c1 = ws.cell(row, 2, pct_qty)
        c1.border = THIN_BORDER
        c1.number_format = '0.00%'
        c2 = ws.cell(row, 3, pct_sum)
        c2.border = THIN_BORDER
        c2.number_format = '0.00%'
    
    ws.cell(14, 1, 'Тюмень').font = Font(bold=True, size=12)
    ws.cell(14, 1, 'Тюмень').fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(15, idx, header)
        cell.font = Font(bold=True, size=10)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    tmn_data = grouped[grouped['Город'] == 'Тюмень']
    total_qty_tmn = tmn_data['Количество блюд'].sum()
    total_sum_tmn = tmn_data['Сумма со скидкой, р.'].sum()
    for idx, cat in enumerate(CATEGORIES_DYN):
        row = 16 + idx
        cat_row = tmn_data[tmn_data['Категория_доля'] == cat]
        qty = int(cat_row['Количество блюд'].sum()) if len(cat_row) > 0 else 0
        sum_val = round(cat_row['Сумма со скидкой, р.'].sum(), 2) if len(cat_row) > 0 else 0
        pct_qty = qty / total_qty_tmn if total_qty_tmn > 0 else 0
        pct_sum = sum_val / total_sum_tmn if total_sum_tmn > 0 else 0
        ws.cell(row, 1, cat).border = THIN_BORDER
        c1 = ws.cell(row, 2, pct_qty)
        c1.border = THIN_BORDER
        c1.number_format = '0.00%'
        c2 = ws.cell(row, 3, pct_sum)
        c2.border = THIN_BORDER
        c2.number_format = '0.00%'


# ==================== ГЛАВНАЯ ТОЧКА ВХОДА ЛОГИКИ ====================
def generate_product_report(file_main: BinaryIO, file_combo: BinaryIO) -> io.BytesIO:
    """
    Генерирует итоговый Excel-отчет по продукту.
    Принимает бинарные потоки файлов, возвращает BytesIO с готовым xlsx.
    НЕ содержит streamlit.
    """
    df_main_raw = _read_main_file_raw(file_main)
    file_main.seek(0)
    df_main = _parse_main_file(file_main)
    df_combo = _parse_combo_file(file_combo)
    
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    _create_rating_sheet(wb, df_main, df_combo)
    _create_pizza_sheet(wb, df_main_raw)
    _create_combo_sheet(wb, df_combo)
    _create_category_share_sheet(wb, df_main_raw, df_combo)
    _create_category_dynamics_sheet(wb, df_main_raw, df_combo)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output