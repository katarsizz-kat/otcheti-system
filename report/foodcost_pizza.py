from io import BytesIO
from typing import Any, Dict, List, Optional, Set, Tuple

import re

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from report.foodcost_common import safe_file_name


RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FFFF0000",
    end_color="FFFF0000",
)

GREY_FILL = PatternFill(
    fill_type="solid",
    start_color="FFD9D9D9",
    end_color="FFD9D9D9",
)

NO_FILL = PatternFill(fill_type=None)


COST_NAME_COL = 2
VALID_SIZES = {23, 30, 35, 40}

SIZE_TO_COLUMNS = {
    23: {"menu": 3, "target": 4},
    30: {"menu": 6, "target": 7},
    35: {"menu": 9, "target": 10},
    40: {"menu": 12, "target": 13},
}


TEMPLATE_PIZZA_NAMES_RAW = [
    "Большая Бонанза",
    "8 сыров",
    "Итальянская с моцареллой и пепперони",
    "Любимая папина пицца",
    "4 сыра",
    "Маленькая Италия",
    "Мясная",
    "Супер Папа",
    "Баварская",
    "Альфредо",
    "Цыпленок Рэнч",
    "Папа Микс",
    "Мясное барбекю",
    "Цыпленок Барбекю",
    "Мексиканская",
    "Двойная пепперони",
    "Гавайская",
    "Цыплёнок Флорентина",
    "Маргарита",
    "Вегетарианская",
    "Ветчина и Грибы",
    "Пепперони",
    "Чизбургер",
    "Чикен Пармеджано",
    "Капричиоза",
    "Домашняя",
    "Деревенская",
    "Нежная",
    "Цыплёнок Грин",
    "Пепперони Грин",
    "Сырная Пицца",
]


PIZZA_COST_ALIAS_RAW = {
    "Пицца 8 сыров NEW": "8 сыров",
    "Четыре Сыра": "4 сыра",
    "Сырная": "Сырная Пицца",
}


PIZZA_SALES_ALIAS_RAW = {
    "Пицца 8 сыров": "8 сыров",
    "Четыре Сыра": "4 сыра",
    "Сырная": "Сырная Пицца",
}


def _normalize(name: Any) -> str:
    if name is None:
        return ""

    s = str(name).lower().strip()
    s = s.replace("\u00a0", " ")
    s = s.replace("ё", "е")

    if re.search("[а-я]", s):
        s = s.replace("o", "о")

    s = re.sub(r"['\"«»`]", "", s)
    s = re.sub(r"[^a-zа-я0-9]+", " ", s)

    tokens = s.split()
    tokens = [
        token
        for token in tokens
        if token not in {"гр", "г", "шт", "л", "мл"}
    ]

    return " ".join(tokens).strip()


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if not s:
        return None

    s = s.replace("\u00a0", " ").replace("₽", "").strip()
    s = re.sub(r"\s+", "", s)

    if not s:
        return None

    # Примеры:
    # 36,599 -> 36599
    # 1,079 -> 1079
    # 69.88 -> 69.88
    # 69,88 -> 69.88
    # 1 234 -> 1234
    if "," in s and "." in s:
        s = s.replace(",", "")
    elif "," in s:
        if re.fullmatch(r"-?\d{1,3}(,\d{3})+", s):
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")

    try:
        return float(s)
    except Exception:
        return None


def _parse_size(value: Any) -> Optional[int]:
    numeric = _to_float(value)
    if numeric is not None:
        size = int(round(numeric))
        return size if size in VALID_SIZES else None

    s = str(value).lower()
    match = re.search(r"(23|30|35|40)", s)
    if not match:
        return None

    return int(match.group(1))


def _round_cost(value: float) -> int | float:
    if value < 50:
        return round(value, 1)

    return int(round(value))


def _is_formula(cell: Any) -> bool:
    if cell.data_type == "f":
        return True

    value = cell.value
    return isinstance(value, str) and value.strip().startswith("=")


def _is_positive_menu(ws: Any, row: int, column: int) -> bool:
    value = _to_float(ws.cell(row=row, column=column).value)
    return value is not None and value > 0


def _get_sheet(wb: Any, names: Set[str], fallback_index: int) -> Any:
    for ws in wb.worksheets:
        if _normalize(ws.title) in names:
            return ws

    if len(wb.worksheets) > fallback_index:
        return wb.worksheets[fallback_index]

    return wb.active


def _build_alias_map(raw: Dict[str, str]) -> Dict[str, str]:
    return {_normalize(key): _normalize(value) for key, value in raw.items()}


TEMPLATE_PIZZA_KEYS = {_normalize(name) for name in TEMPLATE_PIZZA_NAMES_RAW}
PIZZA_COST_ALIASES = _build_alias_map(PIZZA_COST_ALIAS_RAW)
PIZZA_SALES_ALIASES = _build_alias_map(PIZZA_SALES_ALIAS_RAW)


def _map_cost_base(base_key: str) -> Optional[str]:
    mapped = PIZZA_COST_ALIASES.get(base_key)
    if mapped:
        return mapped

    if base_key in TEMPLATE_PIZZA_KEYS:
        return base_key

    return None


def _map_sales_name(name_key: str) -> Optional[str]:
    mapped = PIZZA_SALES_ALIASES.get(name_key)
    if mapped:
        return mapped

    if name_key in TEMPLATE_PIZZA_KEYS:
        return name_key

    return None


def _parse_cost_pizza_name(raw_name: Any) -> Optional[Tuple[str, int]]:
    if raw_name is None:
        return None

    name = str(raw_name).strip()
    if not name or "(" not in name:
        return None

    s = name.lower()

    if "+" in s:
        return None

    bad_words = [
        "тонкое",
        "подарок",
        "не использовать",
        "кусок",
        "половинк",
        "корочка",
        "15 см",
        "15)",
    ]

    if any(bad in s for bad in bad_words):
        return None

    if re.search(r"(^|[^а-яa-z0-9])(кб|сб)([^а-яa-z0-9]|$)", s):
        return None

    if "трад" not in s:
        return None

    size_match = re.search(r"(23|30|35|40)", s)
    if not size_match:
        return None

    size = int(size_match.group(1))
    if size not in VALID_SIZES:
        return None

    base_raw = name.split("(")[0].strip()
    base_key = _normalize(base_raw)

    if not base_key:
        return None

    return base_key, size


def _read_cost_items(
    cost_file: Any,
    city_normalized: str,
) -> Tuple[Dict[Tuple[str, int], Dict[str, Any]], Dict[Tuple[str, int], Dict[str, Any]]]:
    df = pd.read_excel(cost_file, header=None)

    header_row = None
    spb_sum_col = None
    tyumen_sum_col = None

    for row_index, row in df.iterrows():
        values = [_normalize(cell) for cell in row.values]
        sum_cols = [
            col_index
            for col_index, value in enumerate(values)
            if value == "сумма"
        ]

        if len(sum_cols) >= 2:
            header_row = row_index
            spb_sum_col = sum_cols[0]
            tyumen_sum_col = sum_cols[1]
            break

    if header_row is None or spb_sum_col is None or tyumen_sum_col is None:
        raise ValueError(
            "Не удалось распознать структуру файла себестоимости пиццы."
        )

    city_col = (
        tyumen_sum_col
        if city_normalized == "тюмень"
        else spb_sum_col
    )

    mapped_items: Dict[Tuple[str, int], Dict[str, Any]] = {}
    all_items: Dict[Tuple[str, int], Dict[str, Any]] = {}

    for row in df.iloc[header_row + 1 :].itertuples(index=False, name=None):
        if len(row) <= max(COST_NAME_COL, city_col):
            continue

        raw_name = row[COST_NAME_COL]
        if raw_name is None or pd.isna(raw_name):
            continue

        parsed = _parse_cost_pizza_name(raw_name)
        if parsed is None:
            continue

        base_key, size = parsed

        value = _to_float(row[city_col])
        if value is None:
            continue

        template_key = _map_cost_base(base_key)

        item_key = (base_key, size)
        if item_key not in all_items:
            all_items[item_key] = {
                "original": str(raw_name).strip(),
                "size": size,
                "value": value,
                "template_key": template_key,
            }

        if template_key:
            mapped_key = (template_key, size)
            if mapped_key not in mapped_items:
                mapped_items[mapped_key] = {
                    "original": str(raw_name).strip(),
                    "value": value,
                }

    return mapped_items, all_items


def _read_sales_items(
    sales_file: Any,
) -> Tuple[Dict[Tuple[str, int], Dict[str, Any]], Dict[Tuple[str, int], Dict[str, Any]]]:
    df = pd.read_excel(sales_file)

    if df.empty:
        raise ValueError("Файл продаж пиццы пустой.")

    name_col = None
    size_col = None
    qty_col = None

    for col in df.columns:
        normalized = _normalize(col).replace(" ", "")

        if normalized == "dishpizzatype":
            name_col = col

        if normalized == "dishpizzasize":
            size_col = col

        if normalized == "quantityofdishes":
            qty_col = col

    if name_col is None:
        name_col = df.columns[0]

    if size_col is None:
        size_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]

    if qty_col is None:
        qty_col = df.columns[2] if len(df.columns) > 2 else df.columns[0]

    mapped_items: Dict[Tuple[str, int], Dict[str, Any]] = {}
    all_items: Dict[Tuple[str, int], Dict[str, Any]] = {}

    for raw_name, raw_size, raw_qty in zip(df[name_col], df[size_col], df[qty_col]):
        if raw_name is None or pd.isna(raw_name):
            continue

        name = str(raw_name).strip()
        if not name or name == "-":
            continue

        if "+" in name:
            continue

        lowered = name.lower()
        if "кусок" in lowered or "половинк" in lowered:
            continue

        size = _parse_size(raw_size)
        if size is None:
            continue

        qty = _to_float(raw_qty)
        if qty is None:
            continue

        name_key = _normalize(name)
        template_key = _map_sales_name(name_key)

        item_key = (name_key, size)

        if item_key in all_items:
            all_items[item_key]["qty"] += qty
        else:
            all_items[item_key] = {
                "original": name,
                "size": size,
                "qty": qty,
                "template_key": template_key,
            }

        if template_key:
            mapped_key = (template_key, size)

            if mapped_key in mapped_items:
                mapped_items[mapped_key]["qty"] += qty
            else:
                mapped_items[mapped_key] = {
                    "original": name,
                    "qty": qty,
                }

    return mapped_items, all_items


def _find_cc_item_rows(ws: Any) -> List[Tuple[int, str]]:
    header_row = None

    for row_index in range(1, ws.max_row + 1):
        name_header = _normalize(ws.cell(row=row_index, column=2).value)
        menu_header = _normalize(ws.cell(row=row_index, column=3).value)
        cc_header = _normalize(ws.cell(row=row_index, column=4).value)

        if (
            name_header == "пиццы"
            and menu_header == "в меню"
            and cc_header == "сс"
        ):
            header_row = row_index
            break

    if header_row is None:
        raise ValueError(
            "Не удалось найти блок СС в шаблоне пиццы."
        )

    rows: List[Tuple[int, str]] = []
    started = False

    for row_index in range(header_row + 1, ws.max_row + 1):
        raw_name = ws.cell(row=row_index, column=2).value

        if raw_name is None or not str(raw_name).strip():
            if started:
                break
            continue

        key = _normalize(raw_name)
        if not key or key == "пиццы":
            break

        started = True
        rows.append((row_index, key))

    return rows


def _find_sales_item_rows(ws: Any) -> List[Tuple[int, str]]:
    header_row = None

    for row_index in range(1, ws.max_row + 1):
        menu_header = _normalize(ws.cell(row=row_index, column=3).value)
        qty_header = _normalize(ws.cell(row=row_index, column=4).value)
        total_header = _normalize(ws.cell(row=row_index, column=5).value)

        if (
            menu_header == "в меню"
            and qty_header == "кол во"
            and total_header == "итого"
        ):
            header_row = row_index
            break

    if header_row is None:
        raise ValueError(
            "Не удалось найти блок продаж в шаблоне пиццы."
        )

    rows: List[Tuple[int, str]] = []
    started = False

    for row_index in range(header_row + 1, ws.max_row + 1):
        raw_name = ws.cell(row=row_index, column=2).value

        if raw_name is None or not str(raw_name).strip():
            if started:
                break
            continue

        key = _normalize(raw_name)
        if not key or key == "пиццы":
            break

        started = True
        rows.append((row_index, key))

    return rows


def _write_extra_positions_sheet(
    wb: Any,
    cost_all_items: Dict[Tuple[str, int], Dict[str, Any]],
    sales_all_items: Dict[Tuple[str, int], Dict[str, Any]],
    matched_cost: Set[Tuple[str, int]],
    matched_sales: Set[Tuple[str, int]],
) -> None:
    sheet_name = "Лишние позиции"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    cost_extras = []
    for item in cost_all_items.values():
        template_key = item["template_key"]
        size = item["size"]

        if template_key is None or (template_key, size) not in matched_cost:
            cost_extras.append(
                (
                    item["original"],
                    size,
                    _round_cost(item["value"]),
                )
            )

    cost_extras.sort(key=lambda item: (str(item[0]).lower(), item[1]))

    sales_extras = []
    for item in sales_all_items.values():
        template_key = item["template_key"]
        size = item["size"]

        if template_key is None or (template_key, size) not in matched_sales:
            sales_extras.append(
                (
                    item["original"],
                    size,
                    int(round(item["qty"])),
                )
            )

    sales_extras.sort(key=lambda item: (str(item[0]).lower(), item[1]))

    def append_section(
        title: str,
        headers: List[str],
        rows: List[Tuple[Any, ...]],
    ) -> None:
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).fill = GREY_FILL

        ws.append(headers)
        for col_index in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col_index).fill = GREY_FILL

        if rows:
            for row in rows:
                ws.append(list(row))
                for col_index in range(1, len(row) + 1):
                    ws.cell(row=ws.max_row, column=col_index).fill = GREY_FILL
        else:
            ws.append(["Нет"])
            for col_index in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col_index).fill = GREY_FILL

        ws.append([])

    append_section(
        title="Позиции из себестоимости, которых нет в шаблоне",
        headers=["Название", "Размер", "Себестоимость"],
        rows=cost_extras,
    )

    append_section(
        title="Позиции из продаж, которых нет в шаблоне",
        headers=["Название", "Размер", "Кол-во"],
        rows=sales_extras,
    )

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20


def build_pizza_report(
    *,
    city: str,
    template_file: Any,
    cost_file: Any,
    sales_file: Any,
) -> Tuple[bytes, str]:
    if template_file is None:
        raise ValueError("Нужно загрузить шаблон пиццы.")

    if cost_file is None:
        raise ValueError("Нужно загрузить файл себестоимости пиццы.")

    if sales_file is None:
        raise ValueError("Нужно загрузить файл продаж пиццы.")

    city_normalized = _normalize(city)
    if city_normalized not in {"спб", "тюмень"}:
        city_normalized = "спб"

    cost_items, cost_all_items = _read_cost_items(
        cost_file=cost_file,
        city_normalized=city_normalized,
    )

    sales_items, sales_all_items = _read_sales_items(sales_file)

    wb = load_workbook(template_file, data_only=False)

    cc_ws = _get_sheet(wb, {"сс", "cc"}, 0)
    sales_ws = _get_sheet(wb, {"продажи", "продажа", "sales"}, 1)

    cc_rows = _find_cc_item_rows(cc_ws)
    sales_rows = _find_sales_item_rows(sales_ws)

    if not cc_rows:
        raise ValueError("Не удалось найти строки пицц на листе СС.")

    if not sales_rows:
        raise ValueError("Не удалось найти строки пицц на листе продаж.")

    formula_cells: List[str] = []
    cc_actions: List[Tuple[int, int, str, int]] = []
    sales_actions: List[Tuple[int, int, str, int]] = []

    for row_index, template_key in cc_rows:
        for size, cols in SIZE_TO_COLUMNS.items():
            menu_col = cols["menu"]
            target_col = cols["target"]

            if not _is_positive_menu(cc_ws, row_index, menu_col):
                continue

            cell = cc_ws.cell(row=row_index, column=target_col)
            if _is_formula(cell):
                formula_cells.append(f"{cc_ws.title}!{cell.coordinate}")

            cc_actions.append((row_index, target_col, template_key, size))

    for row_index, template_key in sales_rows:
        for size, cols in SIZE_TO_COLUMNS.items():
            menu_col = cols["menu"]
            target_col = cols["target"]

            if not _is_positive_menu(sales_ws, row_index, menu_col):
                continue

            cell = sales_ws.cell(row=row_index, column=target_col)
            if _is_formula(cell):
                formula_cells.append(f"{sales_ws.title}!{cell.coordinate}")

            sales_actions.append((row_index, target_col, template_key, size))

    if formula_cells:
        raise RuntimeError(
            "Обнаружены формулы в целевых ячейках сс/кол-во: "
            + ", ".join(formula_cells)
        )

    matched_cost: Set[Tuple[str, int]] = set()
    matched_sales: Set[Tuple[str, int]] = set()

    for row_index, target_col, template_key, size in cc_actions:
        cell = cc_ws.cell(row=row_index, column=target_col)
        item = cost_items.get((template_key, size))

        if item:
            cell.value = _round_cost(item["value"])
            cell.fill = NO_FILL
            matched_cost.add((template_key, size))
        else:
            cell.value = None
            cell.fill = RED_FILL

    for row_index, target_col, template_key, size in sales_actions:
        cell = sales_ws.cell(row=row_index, column=target_col)
        item = sales_items.get((template_key, size))

        if item:
            cell.value = int(round(item["qty"]))
            cell.fill = NO_FILL
            matched_sales.add((template_key, size))
        else:
            cell.value = None
            cell.fill = RED_FILL

    _write_extra_positions_sheet(
        wb=wb,
        cost_all_items=cost_all_items,
        sales_all_items=sales_all_items,
        matched_cost=matched_cost,
        matched_sales=matched_sales,
    )

    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    file_name = safe_file_name(city, "pizza")

    return buffer.getvalue(), file_name