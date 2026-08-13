from io import BytesIO

from openpyxl import Workbook


def make_placeholder_report(marker: str) -> bytes:
    """
    Временная заглушка для проверки UI.

    Далее этот модуль будет заменён реальной генерацией Excel:
    - чтение шаблона через openpyxl с data_only=False;
    - заполнение только целевых ячеек;
    - сохранение формул и форматирования;
    - проверка, что целевая ячейка не содержит формулу;
    - красная подсветка пропусков;
    - отдельный лист для лишних позиций.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    ws["A1"] = marker
    ws["A2"] = (
        "Каркас страницы. Бизнес-логика заполнения Excel-шаблона "
        "будет подключена следующим шагом."
    )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()


def safe_file_name(city: str, report_type: str) -> str:
    city_normalized = city.strip().lower()

    if city_normalized == "спб":
        city_slug = "spb"
    elif city_normalized == "тюмень":
        city_slug = "tyumen"
    else:
        city_slug = city_normalized

    return f"foodcost_{report_type}_{city_slug}.xlsx"