from io import BytesIO
from typing import Any, Dict, List, Set, Tuple

import re

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from report.foodcost_common import safe_file_name


RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FFFF0000",
    end_color="FFFF0000",
)

GREY_FILL = PatternFill(
    fill_type="solid",
    start_color="FFD9D9D9",
    end_color="FFD9D9D9",
)

NO_FILL = PatternFill(fill_type=None)


COST_NAME_COL = 2
CURRENT_CC_COL = 2
CURRENT_MENU_COL = 3

# ВАЖНО: в шаблоне закусок блок "Продажи" начинается с 13-й колонки.
# 13 — название позиции, 14 — количество.
SALES_NAME_COL = 13
SALES_QTY_COL = 14


SKIP_ROW_KEYS = {
    "спб",
    "тюмень",
    "сейчас",
    "изменения",
    "продажи",
    "сс",
    "меню",
    "кол во",
    "доля",
    "на сколько",
    "от сс",
    "итого",
}

SKIP_MENU_KEYS = {
    "меню",
    "в меню",
    "сс",
    "сейчас",
    "изменения",
    "продажи",
}


COST_ALIAS_RAW: Dict[str, List[str]] = {
    "Перчики Пепперончини": ["Перчик Пепперончини"],
    "Картофельные Дольки 170 гр": ["Картофельные дольки"],
    "Картофель из печи 150 гр": ["Картофель из печи 150 гр."],
    "Картофель из печи 200 гр": ["Картофель из печи 200 гр."],
    "Куриные наггетсы 9 шт": ["Куриные наггетсы 9 шт"],
    "Куриные наггетсы 16 шт": ["Куриные наггетсы 16 шт"],
    "Сэндвич Цыпленок Барбекю": ["Римский сэндвич Цыпленок Барбекю"],
    "Сэндвич Ветчина и грибы": ["Римский сэндвич Ветчина и грибы"],
    "Сэндвич Пепперони": ["Римский сэндвич Пепперони"],
    "Джонер Чикен 1000": ["Джонер Чикен 1000"],
    "Пепперони Джонер": ["Пепперони Джонер"],
    "Джонер с ветчиной": ["Джонер с ветчиной"],
    "Чикен Джонер": ["Чикен Джонер"],
    "Веджи Джонер": ["Веджи Джонер"],
    "Рогалики с сыром": ["Рогалики с сыром"],
    "Рогалики с колбасками": ["Рогалики с колбасками"],
    "Рогалики с ветчиной": ["Рогалики с ветчиной"],
    "Сырные Палочки": ["Сырные палочки"],
    "Чеддер палочки": ["Чеддер палочки"],
    "Сырные Палочки с ветчиной и грибами": [
        "Сырные палочки с ветчиной и грибами"
    ],
    "Пападиас Ветчина и грибы": ["Пападиас Ветчина и грибы"],
    "Пападиас Чикен Барбекю": ["Пападиас Чикен Барбекю"],
    "Кальзоне Овощной": ["Кальзоне Овощной"],
    "Кальзоне Мясной": ["Кальзоне Мясной"],
    "Куриные крылышки Барбекю 210": ["Крылышки Барбекю (210 гр)"],
    "Куриные крылышки Барбекю 425": ["Крылышки Барбекю (425 гр)"],
    "Куриные крылышки Барбекю 640": ["Крылышки Барбекю (640 гр)"],
    "Пайнэпл ролл": ["Пайнэпл ролл (12 шт)"],
    "Пепперони ролл": ["Пепперони ролл (12 шт)"],
    "Картофель с ветчиной и грибами": ["Картофель с ветчиной и грибами"],
    "Картофель по-баварски": ["Картофель по-баварски"],
    "Хлебные Ломтики": [
        "Хлебные Ломтики",
        "Хлебные Ломтики (2 шт)",
    ],
    "Крем-суп с белыми грибами и шампиньонами": [
        "Крем-суп с белыми грибами и шампиньонами"
    ],
    "Борщ с говядиной": ["Борщ с говядиной"],
    "Солянка": ["Солянка"],
    "Паста Карбонара": ["Паста Карбонара"],
    "Паста Цыпленок Рэнч": ["Паста Цыпленок Рэнч"],
    "Паста Барбекю с колбасками": ["Паста Барбекю с колбасками"],
    "Паста Цыпленок Флорентина": ["Паста Цыпленок Флорентина"],
    "Омлет с курицей и томатами": ["Омлет с курицей и томатами"],
    "Омлет с беконом": ["Омлет с беконом"],
    "Омлет с томатами": ["Омлет с томатами"],
    "Панкейки": ["Панкейки"],
    "Сырники с соусом": ["Сырники с соусом (3 шт)"],
    "Скроллы с яблоком и брусникой": ["Скроллы с яблоком и брусникой"],
    "Чизкейк Нью-Йорк": ["Чизкейк Нью-Йорк"],
    "Чизкейк Шоколадный": ["Чизкейк Шоколадный"],
    "Тирамису": ["Тирамису"],
    "Донат": ["Донат"],
    "Мороженое Магнат DOUBLE Шоколад": [
        "Мороженое Магнат DOUBLE Шоколад"
    ],
    "Мороженое Магнат DOUBLE Соленая карамель Пинта": [
        "Мороженое Магнат DOUBLE Соленая карамель Пинта"
    ],
}


SALES_ALIAS_RAW: Dict[str, List[str]] = {
    "Перчики Пепперончини": [],
    "Картофельные Дольки 170 гр": ["Картофельные дольки"],
    "Картофель из печи 150 гр": ["Картофель из печи 150 гр."],
    "Картофель из печи 200 гр": ["Картофель из печи 200 гр."],
    "Куриные наггетсы 9 шт": ["Куриные наггетсы 9 шт"],
    "Куриные наггетсы 16 шт": ["Куриные наггетсы 16 шт"],
    "Сэндвич Цыпленок Барбекю": ["Римский сэндвич Цыпленок Барбекю"],
    "Сэндвич Ветчина и грибы": ["Римский сэндвич Ветчина и грибы"],
    "Сэндвич Пепперони": ["Римский сэндвич Пепперони"],
    "Джонер Чикен 1000": ["Джонер Чикен 1000"],
    "Пепперони Джонер": ["Пепперони Джонер"],
    "Джонер с ветчиной": ["Джонер с ветчиной"],
    "Чикен Джонер": ["Чикен Джонер"],
    "Веджи Джонер": ["Веджи Джонер"],
    "Рогалики с сыром": ["Рогалики с сыром"],
    "Рогалики с колбасками": ["Рогалики с колбасками"],
    "Рогалики с ветчиной": ["Рогалики с ветчиной"],
    "Сырные Палочки": ["Сырные палочки"],
    "Чеддер палочки": ["Чеддер палочки"],
    "Сырные Палочки с ветчиной и грибами": [
        "Сырные палочки с ветчиной и грибами"
    ],
    "Пападиас Ветчина и грибы": ["Пападиас Ветчина и грибы"],
    "Пападиас Чикен Барбекю": ["Пападиас Чикен Барбекю"],
    "Кальзоне Овощной": ["Кальзоне Овощной"],
    "Кальзоне Мясной": ["Кальзоне Мясной"],
    "Куриные крылышки Барбекю 210": ["Крылышки Барбекю (210 гр)"],
    "Куриные крылышки Барбекю 425": ["Крылышки Барбекю (425 гр)"],
    "Куриные крылышки Барбекю 640": ["Крылышки Барбекю (640 гр)"],
    "Пайнэпл ролл": ["Пайнэпл ролл (12 шт)"],
    "Пепперони ролл": ["Пепперони ролл (12 шт)"],
    "Картофель с ветчиной и грибами": ["Картофель с ветчиной и грибами"],
    "Картофель по-баварски": ["Картофель по-баварски"],
    "Хлебные Ломтики": ["Хлебные Ломтики (2 шт)"],
    "Крем-суп с белыми грибами и шампиньонами": [
        "Крем-суп с белыми грибами и шампиньонами"
    ],
    "Борщ с говядиной": ["Борщ с говядиной"],
    "Солянка": ["Солянка"],
    "Паста Карбонара": ["Паста Карбонара"],
    "Паста Цыпленок Рэнч": ["Паста Цыпленок Рэнч"],
    "Паста Барбекю с колбасками": ["Паста Барбекю с колбасками"],
    "Паста Цыпленок Флорентина": ["Паста Цыпленок Флорентина"],
    "Омлет с курицей и томатами": ["Омлет с курицей и томатами"],
    "Омлет с беконом": ["Омлет с беконом"],
    "Омлет с томатами": ["Омлет с томатами"],
    "Панкейки": [
        "Панкейки с ореховой начинкой (3 шт)",
        "Панкейки с ореховой пастой (3 шт)",
    ],
    "Сырники с соусом": [
        "Сырники с соусом (3 шт)",
        "Сырники с соусом на выбор (3 шт)",
    ],
    "Скроллы с яблоком и брусникой": ["Скроллы с яблоком и брусникой"],
    "Чизкейк Нью-Йорк": ["Чизкейк Нью-Йорк"],
    "Чизкейк Шоколадный": ["Чизкейк Шоколадный"],
    "Тирамису": ["Тирамису"],
    "Донат": ["Донат с клубничной начинкой"],
    "Мороженое Магнат DOUBLE Шоколад": ["Магнат DOUBLE Шоколад"],
    "Мороженое Магнат DOUBLE Соленая карамель Пинта": [
        "Магнат DOUBLE Соленая карамель Пинта"
    ],
}


def _normalize(name: Any) -> str:
    if name is None:
        return ""

    s = str(name).lower().strip()
    s = s.replace("\u00a0", " ")
    s = s.replace("ё", "е")

    if re.search("[а-я]", s):
        s = s.replace("o", "о")

    s = re.sub(r"['\"«»`]", "", s)
    s = re.sub(r"[^a-zа-я0-9]+", " ", s)

    tokens = s.split()
    tokens = [
        token
        for token in tokens
        if token not in {"гр", "г", "шт", "л", "мл"}
    ]

    return " ".join(tokens).strip()


def _to_float(value: Any) -> float | None:
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if not s:
        return None

    s = s.replace("\u00a0", " ").replace("₽", "").strip()
    s = re.sub(r"\s+", "", s)

    if not s:
        return None

    # Примеры:
    # 36,599 -> 36599
    # 1,079 -> 1079
    # 69.88 -> 69.88
    # 69,88 -> 69.88
    # 1 234 -> 1234
    if "," in s and "." in s:
        s = s.replace(",", "")
    elif "," in s:
        if re.fullmatch(r"-?\d{1,3}(,\d{3})+", s):
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")

    try:
        return float(s)
    except Exception:
        return None


def _build_aliases(raw: Dict[str, List[str]]) -> Dict[str, List[str]]:
    result: Dict[str, List[str]] = {}

    for key, values in raw.items():
        result[_normalize(key)] = [_normalize(value) for value in values]

    return result


COST_ALIASES = _build_aliases(COST_ALIAS_RAW)
SALES_ALIASES = _build_aliases(SALES_ALIAS_RAW)


def _is_formula(cell: Any) -> bool:
    if cell.data_type == "f":
        return True

    value = cell.value
    return isinstance(value, str) and value.strip().startswith("=")


def _round_cost(value: float) -> int | float:
    if value < 50:
        return round(value, 1)

    return int(round(value))


def _read_cost_items(
    cost_file: Any,
    city_normalized: str,
) -> Dict[str, Dict[str, Any]]:
    df = pd.read_excel(cost_file, header=None)

    header_row = None
    spb_sum_col = None
    tyumen_sum_col = None

    for row_index, row in df.iterrows():
        values = [_normalize(cell) for cell in row.values]
        sum_cols = [
            col_index
            for col_index, value in enumerate(values)
            if value == "сумма"
        ]

        if len(sum_cols) >= 2:
            header_row = row_index
            spb_sum_col = sum_cols[0]
            tyumen_sum_col = sum_cols[1]
            break

    if header_row is None or spb_sum_col is None or tyumen_sum_col is None:
        raise ValueError(
            "Не удалось распознать структуру файла себестоимости закусок."
        )

    city_col = (
        tyumen_sum_col
        if city_normalized == "тюмень"
        else spb_sum_col
    )

    items: Dict[str, Dict[str, Any]] = {}

    for row in df.iloc[header_row + 1 :].itertuples(index=False, name=None):
        if len(row) <= max(COST_NAME_COL, city_col):
            continue

        raw_name = row[COST_NAME_COL]
        if raw_name is None or pd.isna(raw_name):
            continue

        name = str(raw_name).strip()
        if not name:
            continue

        key = _normalize(name)
        if not key:
            continue

        value = _to_float(row[city_col])
        if value is None:
            continue

        if key not in items:
            items[key] = {
                "value": value,
                "original": name,
            }

    return items


def _read_sales_items(sales_file: Any) -> Dict[str, Dict[str, Any]]:
    df = pd.read_excel(sales_file)

    if df.empty:
        raise ValueError("Файл продаж закусок пустой.")

    name_col = None
    qty_col = None

    for col in df.columns:
        normalized = _normalize(col).replace(" ", "")

        if normalized == "dishname":
            name_col = col

        if normalized == "quantityofdishes":
            qty_col = col

    if name_col is None:
        name_col = df.columns[0]

    if qty_col is None:
        qty_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]

    items: Dict[str, Dict[str, Any]] = {}

    for raw_name, raw_qty in zip(df[name_col], df[qty_col]):
        if raw_name is None or pd.isna(raw_name):
            continue

        name = str(raw_name).strip()
        if not name or name == "-":
            continue

        key = _normalize(name)
        if not key:
            continue

        qty = _to_float(raw_qty)
        if qty is None:
            continue

        if key in items:
            items[key]["qty"] += qty
        else:
            items[key] = {
                "qty": qty,
                "original": name,
            }

    return items


def _get_cost_value(
    template_key: str,
    cost_items: Dict[str, Dict[str, Any]],
    matched_cost_keys: Set[str],
) -> Tuple[bool, float | None]:
    if template_key in COST_ALIASES:
        candidates = COST_ALIASES[template_key]
    else:
        candidates = [template_key]

    found = False
    result_value = None

    for candidate in candidates:
        if candidate in cost_items:
            matched_cost_keys.add(candidate)

            if not found:
                found = True
                result_value = cost_items[candidate]["value"]

    return found, result_value


def _get_sales_value(
    template_key: str,
    sales_items: Dict[str, Dict[str, Any]],
    matched_sales_keys: Set[str],
) -> Tuple[bool, float]:
    if template_key in SALES_ALIASES:
        candidates = SALES_ALIASES[template_key]
    else:
        candidates = [template_key]

    found = False
    total = 0.0

    for candidate in candidates:
        if candidate in sales_items:
            matched_sales_keys.add(candidate)
            total += sales_items[candidate]["qty"]
            found = True

    return found, total


def _collect_template_rows(ws: Any) -> Tuple[List[Tuple[int, str, str]], List[str]]:
    item_rows: List[Tuple[int, str, str]] = []
    formula_cells: List[str] = []

    for row_index in range(1, ws.max_row + 1):
        raw_name = ws.cell(row=row_index, column=1).value
        if raw_name is None:
            continue

        name = str(raw_name).strip()
        if not name:
            continue

        template_key = _normalize(name)
        if not template_key or template_key in SKIP_ROW_KEYS:
            continue

        menu_value = ws.cell(row=row_index, column=CURRENT_MENU_COL).value
        if menu_value is None:
            continue

        menu_key = _normalize(menu_value)
        if menu_key in SKIP_MENU_KEYS:
            continue

        cc_cell = ws.cell(row=row_index, column=CURRENT_CC_COL)
        qty_cell = ws.cell(row=row_index, column=SALES_QTY_COL)

        if _is_formula(cc_cell):
            formula_cells.append(f"{ws.title}!{cc_cell.coordinate}")

        if _is_formula(qty_cell):
            formula_cells.append(f"{ws.title}!{qty_cell.coordinate}")

        raw_sales_name = ws.cell(row=row_index, column=SALES_NAME_COL).value
        sales_name = (
            str(raw_sales_name).strip()
            if raw_sales_name is not None
            else name
        )

        item_rows.append((row_index, template_key, sales_name))

    return item_rows, formula_cells


def _write_extra_positions_sheet(
    wb: Any,
    cost_items: Dict[str, Dict[str, Any]],
    sales_items: Dict[str, Dict[str, Any]],
    matched_cost_keys: Set[str],
    matched_sales_keys: Set[str],
) -> None:
    sheet_name = "Лишние позиции"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    cost_extras = [
        (data["original"], data["value"])
        for key, data in cost_items.items()
        if key not in matched_cost_keys
    ]
    cost_extras.sort(key=lambda item: str(item[0]).lower())

    sales_extras = [
        (data["original"], data["qty"])
        for key, data in sales_items.items()
        if key not in matched_sales_keys
    ]
    sales_extras.sort(key=lambda item: str(item[0]).lower())

    ws.append(["Позиции из себестоимости, которых нет в шаблоне"])
    ws.cell(row=ws.max_row, column=1).fill = GREY_FILL

    ws.append(["Название", "Себестоимость"])
    for col_index in range(1, 3):
        ws.cell(row=ws.max_row, column=col_index).fill = GREY_FILL

    if cost_extras:
        for original_name, value in cost_extras:
            ws.append([original_name, _round_cost(value)])
            for col_index in range(1, 3):
                ws.cell(row=ws.max_row, column=col_index).fill = GREY_FILL
    else:
        ws.append(["Нет"])
        for col_index in range(1, 3):
            ws.cell(row=ws.max_row, column=col_index).fill = GREY_FILL

    ws.append([])

    ws.append(["Позиции из продаж, которых нет в шаблоне"])
    ws.cell(row=ws.max_row, column=1).fill = GREY_FILL

    ws.append(["Название", "Кол-во"])
    for col_index in range(1, 3):
        ws.cell(row=ws.max_row, column=col_index).fill = GREY_FILL

    if sales_extras:
        for original_name, qty in sales_extras:
            ws.append([original_name, int(round(qty))])
            for col_index in range(1, 3):
                ws.cell(row=ws.max_row, column=col_index).fill = GREY_FILL
    else:
        ws.append(["Нет"])
        for col_index in range(1, 3):
            ws.cell(row=ws.max_row, column=col_index).fill = GREY_FILL

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 20


def build_snacks_report(
    *,
    city: str,
    template_file: Any,
    cost_file: Any,
    sales_file: Any,
) -> Tuple[bytes, str]:
    if template_file is None:
        raise ValueError("Нужно загрузить шаблон закусок.")

    if cost_file is None:
        raise ValueError("Нужно загрузить файл себестоимости закусок.")

    if sales_file is None:
        raise ValueError("Нужно загрузить файл продаж закусок.")

    city_normalized = _normalize(city)
    if city_normalized not in {"спб", "тюмень"}:
        city_normalized = "спб"

    cost_items = _read_cost_items(cost_file, city_normalized)
    sales_items = _read_sales_items(sales_file)

    wb = load_workbook(template_file, data_only=False)
    ws = wb.active

    item_rows, formula_cells = _collect_template_rows(ws)

    if formula_cells:
        raise RuntimeError(
            "Обнаружены формулы в целевых ячейках сс/кол-во: "
            + ", ".join(formula_cells)
        )

    if not item_rows:
        raise ValueError(
            "Не удалось найти строки закусок в шаблоне. "
            "Проверь структуру шаблона."
        )

    matched_cost_keys: Set[str] = set()
    matched_sales_keys: Set[str] = set()

    for row_index, template_key, _sales_name in item_rows:
        found_cost, cost_value = _get_cost_value(
            template_key=template_key,
            cost_items=cost_items,
            matched_cost_keys=matched_cost_keys,
        )

        cc_cell = ws.cell(row=row_index, column=CURRENT_CC_COL)

        if found_cost and cost_value is not None:
            cc_cell.value = _round_cost(cost_value)
            cc_cell.fill = NO_FILL
        else:
            cc_cell.value = None
            cc_cell.fill = RED_FILL

        found_sales, sales_value = _get_sales_value(
            template_key=template_key,
            sales_items=sales_items,
            matched_sales_keys=matched_sales_keys,
        )

        qty_cell = ws.cell(row=row_index, column=SALES_QTY_COL)

        if found_sales:
            qty_cell.value = int(round(sales_value))
            qty_cell.fill = NO_FILL
        else:
            qty_cell.value = None
            qty_cell.fill = RED_FILL

    _write_extra_positions_sheet(
        wb=wb,
        cost_items=cost_items,
        sales_items=sales_items,
        matched_cost_keys=matched_cost_keys,
        matched_sales_keys=matched_sales_keys,
    )

    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    file_name = safe_file_name(city, "snacks")

    return buffer.getvalue(), file_name