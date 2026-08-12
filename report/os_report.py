"""Логика отчёта ОС: анализ данных и генерация Excel через openpyxl.
БЕЗ streamlit. Один файл = одна задача.
"""
import io
import re
from datetime import timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


# ============================================================
# КОНСТАНТЫ
# ============================================================
MSK_OFFSET = timedelta(hours=3)

TYUMEN_ORDER = [
    'Тюмень №2 – Орджоникидзе',
    'Тюмень №3 – Мельникайте',
]

EXCLUDED_OPERATORS = ['Marketing SPB']

CATEGORY_MAPPING = {
    'Отравление': 'Критические инциденты',
    'Инородный предмет в блюде': 'Критические инциденты',
    'Опоздания': 'Опоздания',
    'Не доставили заказ': 'Перепутанная/недовезённая позиция',
    'Холодная пицца': 'Жалоба на блюдо',
    'Разлит / Повреждение упаковки': 'Жалоба на блюдо',
    'Забыли позицию': 'Перепутанная/недовезённая позиция',
    'Перепутали позицию': 'Перепутанная/недовезённая позиция',
    'Невкусно': 'Жалоба на блюдо',
    'Некачественно': 'Жалоба на блюдо',
    'Приложение / Сайт / IT-сбои': 'Сайт и IT-сбои',
    'Проблема с бонусами или промокодом': 'Программа лояльности',
    'Скидка на День Рождения': 'Программа лояльности',
    'Возврат ДС': 'Возврат ДС',
    'Слишком дорого': 'Жалоба на сервис',
    'Жалоба на сервис': 'Жалоба на сервис',
    'Не дозвониться': 'Жалоба на сервис',
    'Управление аккаунтом': 'Аккаунт',
}

AGG_ORDER = [
    'Опоздания', 'Перепутанная/недовезённая позиция', 'Жалоба на блюдо',
    'Сайт и IT-сбои', 'Программа лояльности', 'Возврат ДС',
    'Жалоба на сервис', 'Аккаунт', 'Критические инциденты',
]

INTERVAL_ORDER = [
    '00:00-01:00', '01:00-02:00', '02:00-08:00',
    '08:00-09:00', '09:00-10:00', '10:00-11:00', '11:00-12:00',
    '12:00-13:00', '13:00-14:00', '14:00-15:00', '15:00-16:00',
    '16:00-17:00', '17:00-18:00', '18:00-19:00', '19:00-20:00',
    '20:00-21:00', '21:00-22:00', '22:00-23:00', '23:00-00:00',
]

WEEKDAY_ORDER = [
    'Понедельник', 'Вторник', 'Среда', 'Четверг',
    'Пятница', 'Суббота', 'Воскресенье',
]


# ============================================================
# КАТЕГОРИЗАЦИЯ
# ============================================================
def categorize_complaint_detailed(text, reason):
    """Детальная категоризация по ключевым словам."""
    t = (str(text) + " " + str(reason)).lower()
    rules = [
        (['отравл', 'тошнит', 'тошнило', 'диарея', 'рвота', 'плохо стало', 'заболел', 'заболели'], 'Отравление'),
        (['волос', 'проволока', 'инородн', 'предмет', 'стекло', 'металл', 'мусор', 'плесень', 'жук', 'таракан', 'гвоздь', 'щепка'], 'Инородный предмет в блюде'),
        (['опозд', 'задержк', 'долго', 'где курьер', 'не везут', 'нарушение сроков', 'отодвигается', 'сдвинули время', 'принудительн', 'изменили время', 'праздник не вышло', 'ждали на праздник', 'срыв', 'не привезли вовремя', 'не в заявленное', 'час задержк', 'два часа', 'больше часа', 'более чем на час', 'на 40 минут', 'на 50 минут', 'на 30 минут', 'на полтора', 'статус не меняется', 'курьер не едет', 'не выехал'], 'Опоздания'),
        (['не привезли заказ', 'не доставили заказ', 'нет моего заказа', 'заказ не привезли', 'не получил заказ', 'заказ так и не', 'отменили заказ', 'заказ не попал'], 'Не доставили заказ'),
        (['холодн', 'остывш', 'ледян', 'не горяч', 'еле тепл', 'тепленьк'], 'Холодная пицца'),
        (['разлит', 'вылит', 'растаяло', 'растаявш', 'пострадал десерт', 'повреждение упаковк', 'мятая', 'смята', 'прилипла', 'перевёрнут', 'перевернут', 'соус в пакет', 'соус в картошк', 'раздавлен', 'протекает', 'коробка вся'], 'Разлит / Повреждение упаковки'),
        (['забыли', 'не положили', 'не доложили', 'отсутствует', 'нет в заказе', 'не хватает', 'не привезли соус', 'не привезли перец', 'не привезли фри', 'не привезли крылья', 'не привезли подарочн', 'недовоз', 'не положили один', 'не положили два', 'не было соуса', 'нет соуса', 'без соуса', 'без перца', 'без бортиков', 'без курицы', 'без картошки', 'некомплектн', 'не полный заказ', 'не разрезана', 'не порезана'], 'Забыли позицию'),
        (['перепутали', 'не тот', 'не ту', 'другой заказ', 'вместо', 'привезли одинаковые', 'не то тесто', 'не тот борт', 'сырный борт а не колбасный', 'колбасный а не сырный', 'другую пиццу', 'не та пицца', 'не те', 'ошибк в заказ', 'привезли не то', 'пришла не та', '20 см вместо', '30 см вместо', 'не тот размер'], 'Перепутали позицию'),
        (['невкусно', 'ужасн вкус', 'резин', 'сухая', 'без начинки', 'мало начинки', 'горелые', 'горелая', 'пережарен', 'пересушен', 'никакая на вкус', 'отвратительн вкус'], 'Невкусно'),
        (['некачествен', 'сырое', 'сырая', 'непропечен', 'недожарен', 'контроль качества', 'не соответствует', 'ужасно приготовлен', 'начинка размазан', 'тесто не пропеклос', 'слипшаяся', 'качество доставленн'], 'Некачественно'),
        (['не работает', 'ошибка', 'глючит', 'глючная', 'не оформляется', 'зависает', 'корзина пуст', 'внутренняя ошибка', 'не тот адрес', 'не отображается', 'пропали заказ', 'не могу оформить', 'сложность в оформлен', 'не прошла оплат', 'оплату не засчитал', 'не могу применить', 'сброс', 'не загружается', 'не предлагает оплатить', 'приложение у меня', 'указало этот адрес', 'некорректная работа', 'не могу сделать заказ', 'не могу зарегистрир', 'не приходят сообщен', 'не отображается истор', 'статус заказа в приложени'], 'Приложение / Сайт / IT-сбои'),
        (['промокод', 'бонус', 'папа бонус', 'балл', 'не начислили', 'списали', 'не применился', 'не сработал', 'начислен', 'баллы за', 'не приходят бонусн', 'автоматом спиш', 'не дали списать', 'не дает списать'], 'Проблема с бонусами или промокодом'),
        (['день рождения', 'день рождени', 'др', 'скидка на др', 'промокод на день', 'подарок на др', 'на днях был др', 'дню рождении', 'днем рождения', 'не пришёл промокод', 'не присылают'], 'Скидка на День Рождения'),
        (['возврат', 'верните деньги', 'верните средств', 'вернуть деньги', 'вернуть средств', 'двойное списан', 'два раза оплатил', 'списались за оба', 'компенсац', 'возмещен', 'когда вернутся', 'возврат денежн', 'частичную компенс'], 'Возврат ДС'),
        (['слишком дорог', 'очень дорог', 'завышен', 'содрать', 'платная доставка', 'стоимость доставк', 'цена завыш', 'неприемлем', 'заплатила очень много'], 'Слишком дорого'),
        (['груб', 'хам', 'невежл', 'хамили', 'наказать', 'конфликт', 'перекладывание вины', 'не умеете работать', 'отвратительн сервис', 'курьер грубил', 'смеялся', 'не отдал заказ', 'без термосумки', 'без терминала', 'не русского', 'отказалась принять', 'не приняли заказ', 'отказ принять'], 'Жалоба на сервис'),
        (['не берут трубку', 'не отвечает', 'сбрасывает', 'недозвон', 'телефон недоступен', 'не дозвониться', 'номер занят', 'не подходит', 'никто не взял', 'не отвечает на звонки', 'оператор не отвечает', 'поддержка не отвечает', 'горячую линию', 'никто не подходит', 'сброс звонка', 'идет сброс'], 'Не дозвониться'),
        (['удалите профиль', 'удаление аккаунта', 'восстановить аккаунт', 'объединить', 'уведомления', 'спам', 'рассылк', 'отключите', 'изменился номер', 'объединение аккаунт', 'не могу зарегистрир', 'случайно удалила', 'удалил аккаунт'], 'Управление аккаунтом'),
    ]
    for keywords, cat in rules:
        if any(w in t for w in keywords):
            return cat
    return 'Другое / Запрос'


def get_aggregated_category(detailed_cat):
    """Возвращает укрупнённую категорию."""
    return CATEGORY_MAPPING.get(detailed_cat, 'Другое')


# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ
# ============================================================
def extract_restaurant_number(restaurant_name):
    if pd.isna(restaurant_name):
        return 9999
    match = re.search(r'№\s*(\d+)', str(restaurant_name))
    return int(match.group(1)) if match else 9999


def get_restaurant_sort_key(restaurant_name, city_name=''):
    city = str(city_name)
    name = str(restaurant_name)
    if 'Санкт-Петербург' in city or 'Санкт-Петербург' in name:
        return (0, extract_restaurant_number(name), name)
    if 'Тюмень' in city or 'Тюмень' in name:
        try:
            return (1, TYUMEN_ORDER.index(name), name)
        except ValueError:
            return (1, 9999, name)
    return (2, extract_restaurant_number(name), name)


def analyze_bot_fails(df):
    """Анализ сбоев чат-бота."""
    bot_fails = []
    fail_kw = ['запуталась', 'не поняла', 'упс, кажется', 'пошло не так', 'не совсем поняла']
    demand_kw = ['оператор', 'позови', 'соедини', 'живого', 'человек']
    for _, row in df.iterrows():
        text = str(row.get('Первичное сообщение', '')).lower()
        is_fail = any(k in text for k in fail_kw)
        is_demand = any(k in text for k in demand_kw)
        if is_fail or is_demand:
            msgs = re.findall(
                r'Клиент\s*([^)]+):\s*(.*?)(?:\n|$)',
                str(row.get('Первичное сообщение', '')),
                re.DOTALL,
            )
            summary = msgs[0][1].strip()[:120] if msgs else "Нет текста"
            bot_fails.append({
                '№ Обращения': row.get('Номер обращения', ''),
                'Дата': row.get('Дата отзыва', ''),
                'Ресторан': row.get('Ресторан', ''),
                'Суть обращения': summary,
                'Тип ошибки бота': 'Требовал оператора' if is_demand else 'Бот не понял',
            })
    return pd.DataFrame(bot_fails)


def get_hour_interval(hour):
    if hour == 0:
        return '00:00-01:00'
    if hour == 1:
        return '01:00-02:00'
    if 2 <= hour < 8:
        return '02:00-08:00'
    next_hour = (hour + 1) % 24
    return f"{hour:02d}:00-{next_hour:02d}:00"


def get_interval_hours(label):
    if label == '02:00-08:00':
        return 6
    return 1


def get_weekday_name(d):
    return WEEKDAY_ORDER[d.weekday()]


# ============================================================
# ЗАГРУЗКА И ФИЛЬТРАЦИЯ
# ============================================================
def load_data(uploaded_file):
    """Читает Excel/CSV, парсит даты."""
    if uploaded_file.name.endswith('.csv'):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)
    df['Дата отзыва'] = pd.to_datetime(df['Дата отзыва'], errors='coerce')
    return df


def apply_filters(df, cities, date_type, date_range=None, selected_dates=None):
    """Фильтры по городам и датам."""
    df_filtered = df[df['Город'].isin(cities)] if cities else df.copy()
    valid_dates = df_filtered['Дата отзыва'].dropna()
    if valid_dates.empty:
        return df_filtered

    if date_type == "Интервал дат" and date_range and len(date_range) == 2:
        d1, d2 = date_range
        df_filtered = df_filtered[
            (df_filtered['Дата отзыва'].dt.date >= d1) &
            (df_filtered['Дата отзыва'].dt.date <= d2)
        ]
    elif date_type == "Отдельные даты" and selected_dates:
        df_filtered = df_filtered[
            df_filtered['Дата отзыва'].dt.date.isin(selected_dates)
        ]
    return df_filtered


# ============================================================
# РАСЧЁТЫ
# ============================================================
def compute_operator_stats(df_op):
    rows = []
    for op in df_op['Исполнитель'].unique():
        sub = df_op[df_op['Исполнитель'] == op]
        days = sub['Дата_рабочая'].nunique()
        total = len(sub)
        rows.append({
            'Оператор': op,
            'Всего обращений': total,
            'Рабочих дней': days,
            'Среднее в день': round(total / days, 2) if days else 0,
        })
    return pd.DataFrame(rows).sort_values('Всего обращений', ascending=False)


def compute_joint_work(df_op):
    df_joint = df_op[~df_op['Исполнитель'].isin(EXCLUDED_OPERATORS)].copy()
    day_ops = df_joint.groupby('Дата_рабочая')['Исполнитель'].apply(
        lambda x: sorted(set(x))
    ).reset_index()
    day_ops['Кол-во'] = day_ops['Исполнитель'].apply(len)
    multi = day_ops[day_ops['Кол-во'] > 1].copy()
    if multi.empty:
        return multi
    multi['Операторы'] = multi['Исполнитель'].apply(', '.join)
    return multi[['Дата_рабочая', 'Кол-во', 'Операторы']].sort_values(
        'Дата_рабочая', ascending=False,
    )


def compute_hourly_stats(df_op):
    total_days = df_op['Дата_рабочая'].nunique()
    df = df_op.copy()
    df['Интервал'] = df['Час_МСК'].apply(get_hour_interval)
    h_stats = df.groupby('Интервал').size().reset_index(name='Обращений')
    h_stats['Обращений/час'] = h_stats.apply(
        lambda r: round(r['Обращений'] / get_interval_hours(r['Интервал']), 2), axis=1,
    )
    h_stats['Обращений/час/день'] = h_stats.apply(
        lambda r: round(
            r['Обращений'] / get_interval_hours(r['Интервал']) / total_days, 2,
        ) if total_days > 0 else 0, axis=1,
    )
    h_stats['_s'] = h_stats['Интервал'].apply(
        lambda x: INTERVAL_ORDER.index(x) if x in INTERVAL_ORDER else 99,
    )
    h_stats = h_stats.sort_values('_s').drop('_s', axis=1)
    all_intervals = pd.DataFrame({'Интервал': INTERVAL_ORDER})
    h_stats = all_intervals.merge(h_stats, on='Интервал', how='left').fillna(0)
    h_stats['Обращений'] = h_stats['Обращений'].astype(int)
    h_stats['Обращений/час'] = h_stats['Обращений/час'].astype(float)
    h_stats['Обращений/час/день'] = h_stats['Обращений/час/день'].astype(float)
    return h_stats, total_days


def compute_day_stats(df_op):
    day_stats = df_op.groupby('Дата_рабочая').size().reset_index(name='Обращений')
    day_stats = day_stats.sort_values('Обращений', ascending=False)
    day_stats['День недели'] = day_stats['Дата_рабочая'].apply(get_weekday_name)

    wd_stats = df_op.groupby(
        df_op['Дата_рабочая'].apply(get_weekday_name),
    ).size().reset_index(name='Обращений')
    wd_stats.columns = ['День недели', 'Обращений']
    wd_stats['_s'] = wd_stats['День недели'].apply(
        lambda x: WEEKDAY_ORDER.index(x) if x in WEEKDAY_ORDER else 99,
    )
    return day_stats, wd_stats.sort_values('_s').drop('_s', axis=1)


def compute_complaint_pivots(df_f):
    """Возвращает (df_с_категориями, p1_ui, p2_ui, city_map)."""
    df = df_f.copy()
    df['Категория'] = df.apply(
        lambda r: categorize_complaint_detailed(
            r.get('Первичное сообщение', ''), r.get('Причина обращения', ''),
        ), axis=1,
    )
    df['Категория_укрупн'] = df['Категория'].apply(get_aggregated_category)

    city_map = df.drop_duplicates('Ресторан').set_index('Ресторан')['Город'].to_dict()

    pivot1 = pd.pivot_table(
        df, index='Ресторан', columns='Категория', aggfunc='size', fill_value=0,
    )
    pivot1['ИТОГО'] = pivot1.sum(axis=1)
    p1_reset = pivot1.reset_index()
    p1_reset['Город'] = p1_reset['Ресторан'].map(city_map).fillna('')
    p1_reset['_sk'] = p1_reset.apply(
        lambda r: get_restaurant_sort_key(r['Ресторан'], r['Город']), axis=1,
    )
    p1_sorted = p1_reset.sort_values('_sk')

    pivot2 = pd.pivot_table(
        df, index='Ресторан', columns='Категория_укрупн', aggfunc='size', fill_value=0,
    )
    pivot2['ИТОГО'] = pivot2.sum(axis=1)
    p2_reset = pivot2.reset_index()
    p2_reset['Город'] = p2_reset['Ресторан'].map(city_map).fillna('')
    p2_reset['_sk'] = p2_reset.apply(
        lambda r: get_restaurant_sort_key(r['Ресторан'], r['Город']), axis=1,
    )
    p2_sorted = p2_reset.sort_values('_sk')

    p1_for_ui = p1_sorted.drop(columns=['Город', '_sk'], errors='ignore')
    p2_for_ui = p2_sorted.drop(columns=['Город', '_sk'], errors='ignore')

    return df, p1_for_ui, p2_for_ui, city_map


# ============================================================
# ГЕНЕРАЦИЯ EXCEL (openpyxl, формулы =SUM)
# ============================================================
def _apply_header_style(ws, row, headers):
    font = Font(bold=True, color='FFFFFF', size=11)
    fill = PatternFill('solid', fgColor='2C3E50')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.border = border
        cell.alignment = align


def _write_df_to_sheet(ws, start_row, df, headers=None):
    if headers is None:
        headers = list(df.columns)
    _apply_header_style(ws, start_row, headers)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=False), start=start_row + 1,
    ):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin
    return start_row + len(df) + 1


def _style_total_row(ws, row_idx, col_count):
    total_font = Font(bold=True, size=11, color='2C3E50')
    total_fill = PatternFill('solid', fgColor='E8F8F5')
    total_border = Border(
        left=Side(style='double'), right=Side(style='double'),
        top=Side(style='double'), bottom=Side(style='double'),
    )
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border


def _color_city_rows(ws, start_row, end_row, city_map, restaurant_col_idx=1):
    """Раскрашивает строки: СПб — голубой, Тюмень — персиковый."""
    spb_fill = PatternFill('solid', fgColor='D6EAF8')
    tyumen_fill = PatternFill('solid', fgColor='FAE5D3')
    for row_idx in range(start_row, end_row + 1):
        restaurant_name = ws.cell(row=row_idx, column=restaurant_col_idx).value
        city = city_map.get(restaurant_name, '')
        fill = None
        if 'Санкт-Петербург' in str(city):
            fill = spb_fill
        elif 'Тюмень' in str(city):
            fill = tyumen_fill
        if fill:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill


def _replace_itogo_with_formula(ws, headers, last_row):
    """Заменяет значения ИТОГО формулами =SUM."""
    if 'ИТОГО' not in headers:
        return
    itog_col = headers.index('ИТОГО') + 1
    itog_letter = get_column_letter(itog_col)
    first_letter = get_column_letter(2)
    last_letter = get_column_letter(itog_col - 1)
    bold_font = Font(bold=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for r in range(2, last_row):
        cell = ws[f"{itog_letter}{r}"]
        cell.value = f"=SUM({first_letter}{r}:{last_letter}{r})"
        cell.font = bold_font
        cell.border = thin


def _add_total_row(ws, headers, total_row):
    """Добавляет итоговую строку с формулами =SUM."""
    ws.cell(row=total_row, column=1, value="ИТОГО")
    for col_idx in range(2, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws[f"{col_letter}{total_row}"] = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
    _style_total_row(ws, total_row, len(headers))


def generate_excel_report(df_op_stats, h_stats, day_stats, wd_stats,
                          p1, p2, bot_df, city_map):
    """6 листов с формулами =SUM и графиками."""
    wb = Workbook()

    # ── Лист 1: Операторы ──
    if df_op_stats is not None and not df_op_stats.empty:
        ws = wb.active
        ws.title = "Операторы"
        _write_df_to_sheet(ws, 1, df_op_stats, list(df_op_stats.columns))
        ws.column_dimensions['A'].width = 30
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 18
    else:
        ws = wb.active
        ws.title = "Операторы"
        ws['A1'] = "Нет данных по операторам"

    # ── Лист 2: Часы ──
    if h_stats is not None and not h_stats.empty:
        ws2 = wb.create_sheet("Часы")
        _write_df_to_sheet(ws2, 1, h_stats, list(h_stats.columns))
        ws2.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D']:
            ws2.column_dimensions[col].width = 20

        chart = BarChart()
        chart.type = "col"
        chart.title = "Нагрузка по часам (МСК)"
        chart.width = 25
        chart.height = 13
        cats = Reference(ws2, min_col=1, min_row=2, max_row=len(h_stats) + 1)
        chart.add_data(
            Reference(ws2, min_col=3, min_row=1, max_row=len(h_stats) + 1),
            titles_from_data=True,
        )
        chart.add_data(
            Reference(ws2, min_col=4, min_row=1, max_row=len(h_stats) + 1),
            titles_from_data=True,
        )
        chart.set_categories(cats)
        ws2.add_chart(chart, "F2")

    # ── Лист 3: Дни ──
    if day_stats is not None and not day_stats.empty:
        ws3 = wb.create_sheet("Дни")
        ws3['A1'] = 'Топ загруженных дней'
        ws3['A1'].font = Font(bold=True, size=13, color='005F6B')
        _write_df_to_sheet(ws3, 2, day_stats, list(day_stats.columns))
        ws3.column_dimensions['A'].width = 14
        ws3.column_dimensions['B'].width = 16
        ws3.column_dimensions['C'].width = 16

        sr = len(day_stats) + 5
        ws3[f'A{sr}'] = 'По дням недели'
        ws3[f'A{sr}'].font = Font(bold=True, size=13, color='005F6B')
        _write_df_to_sheet(ws3, sr + 1, wd_stats, list(wd_stats.columns))

    # ── Лист 4: Жалобы детальные ──
    if p1 is not None and not p1.empty:
        ws4 = wb.create_sheet("Жалобы детальные")
        headers = list(p1.columns)
        last_row = _write_df_to_sheet(ws4, 1, p1.copy(), headers)
        _replace_itogo_with_formula(ws4, headers, last_row)
        _add_total_row(ws4, headers, last_row)
        if city_map:
            _color_city_rows(ws4, 2, last_row - 1, city_map)
        ws4.column_dimensions['A'].width = 35

        if len(headers) > 2:
            chart = BarChart()
            chart.type = "bar"
            chart.grouping = "stacked"
            chart.title = "Детальные жалобы по ресторанам"
            chart.width = 25
            chart.height = 17
            chart.set_categories(
                Reference(ws4, min_col=1, min_row=2, max_row=last_row - 1),
            )
            for col_idx in range(2, len(headers)):
                chart.add_data(
                    Reference(ws4, min_col=col_idx, min_row=1, max_row=last_row - 1),
                    titles_from_data=True,
                )
            ws4.add_chart(chart, "H2")

    # ── Лист 5: Жалобы укрупнённые ──
    if p2 is not None and not p2.empty:
        ws5 = wb.create_sheet("Жалобы укрупнённые")
        headers = list(p2.columns)
        last_row = _write_df_to_sheet(ws5, 1, p2.copy(), headers)
        _replace_itogo_with_formula(ws5, headers, last_row)
        _add_total_row(ws5, headers, last_row)
        if city_map:
            _color_city_rows(ws5, 2, last_row - 1, city_map)
        ws5.column_dimensions['A'].width = 35

        if len(headers) > 2:
            chart = BarChart()
            chart.type = "bar"
            chart.grouping = "stacked"
            chart.title = "Укрупнённые жалобы по ресторанам"
            chart.width = 22
            chart.height = 14
            chart.set_categories(
                Reference(ws5, min_col=1, min_row=2, max_row=last_row - 1),
            )
            for col_idx in range(2, len(headers)):
                chart.add_data(
                    Reference(ws5, min_col=col_idx, min_row=1, max_row=last_row - 1),
                    titles_from_data=True,
                )
            ws5.add_chart(chart, "H2")

    # ── Лист 6: Бот ──
    if bot_df is not None and not bot_df.empty:
        ws6 = wb.create_sheet("Бот")
        _write_df_to_sheet(ws6, 1, bot_df, list(bot_df.columns))
        ws6.column_dimensions['A'].width = 15
        ws6.column_dimensions['B'].width = 20
        ws6.column_dimensions['C'].width = 30
        ws6.column_dimensions['D'].width = 60
        ws6.column_dimensions['E'].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf