# report/kr/excel.py

"""
Генерация Excel-отчёта КР (месяц / неделя).

Ответственность файла:
- лист "Оценки": блоки Сайт / Агрегаторы / Геосервисы;
- общий блок справа с формулами;
- лист "Анализ отзывов": жалобы и позитив;
- форматирование: цветные заголовки, блоки СПб/Тюмень,
  двойные границы итогов;
- формулы: =SUM, =IF, =SUMPRODUCT.

Здесь не должно быть:
- Streamlit;
- загрузки файлов;
- расчётов (они в stats.py).
"""

from __future__ import annotations

import io
from typing import TYPE_CHECKING, Any, Dict, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ==========================================================
# ИМПОРТ КОНСТАНТ
# ==========================================================
try:
    from . import constants as c
except Exception:
    try:
        from report.kr import constants as c
    except Exception:
        c = None

# ==========================================================
# ИМПОРТ МОДЕЛЕЙ — ТОЛЬКО ДЛЯ АННОТАЦИЙ ТИПОВ
# ==========================================================
if TYPE_CHECKING:  # pragma: no cover
    from .models import KRReportData, KRReportSettings


# ==========================================================
# ЛОКАЛЬНЫЕ ССЫЛКИ НА КОНСТАНТЫ
# ==========================================================
DEFAULT_SPB_ORDER = [
    "Транспортный",
    "Димитрова",
    "Шмидта",
    "Пулковская",
    "Благодатная",
    "Энтузиастов",
    "Серебристый",
    "Мурино",
    "Ветеранов",
    "Туристская",
    "Наука",
    "Ленинский",
]

DEFAULT_TMN_ORDER = [
    "Орджоникидзе",
    "Мельникайте",
]

SPB_ORDER = getattr(c, "SPB_ORDER", DEFAULT_SPB_ORDER) or DEFAULT_SPB_ORDER
TMN_ORDER = getattr(c, "TMN_ORDER", DEFAULT_TMN_ORDER) or DEFAULT_TMN_ORDER

COL_RESTAURANT = getattr(c, "COL_RESTAURANT", "Ресторан")
TOTAL_COLUMN = getattr(c, "TOTAL_COLUMN", "всего:")
AVG_COLUMN = getattr(c, "AVG_COLUMN", "Средний рейтинг")

LOW_REVIEW_COLUMN = getattr(c, "LOW_REVIEW_COLUMN", "Отзывы ≤ порог")

HEADER_CITY_SPB = getattr(c, "HEADER_CITY_SPB", "СПБ")
HEADER_CITY_TMN = getattr(c, "HEADER_CITY_TMN", "Тюмень")

CITY_SPB = getattr(c, "CITY_SPB", "СПб")
CITY_TMN = getattr(c, "CITY_TMN", "Тюмень")

CITY_TOTAL_LABEL_TEMPLATE = getattr(c, "CITY_TOTAL_LABEL_TEMPLATE", "Итого {city}:")
SPB_TOTAL_LABEL = CITY_TOTAL_LABEL_TEMPLATE.format(city=CITY_SPB)
TMN_TOTAL_LABEL = CITY_TOTAL_LABEL_TEMPLATE.format(city=CITY_TMN)

BLOCK_TITLE_SITE = getattr(c, "BLOCK_TITLE_SITE", "Сайт/приложение")
BLOCK_TITLE_AGG = getattr(c, "BLOCK_TITLE_AGG", "Агрегаторы")
BLOCK_TITLE_GEO = getattr(
    c,
    "BLOCK_TITLE_GEO",
    "Геосервисы (Я.Карты, 2ГИС, Гугл карты)",
)
BLOCK_TITLE_COMMON = getattr(
    c,
    "BLOCK_TITLE_COMMON",
    "Общий (сайт+агрегаторы+геосервисы)",
)

STAR_SUBTITLE = getattr(c, "STAR_SUBTITLE", "Кол-во поставленных звезд")

EXCEL_SHEET_RATINGS = getattr(c, "EXCEL_SHEET_RATINGS", "Оценки")
EXCEL_SHEET_ANALYSIS = getattr(c, "EXCEL_SHEET_ANALYSIS", "Анализ отзывов")

ANALYSIS_TITLE_COMPLAINTS = getattr(
    c,
    "ANALYSIS_TITLE_COMPLAINTS",
    "Анализ отзывов (жалобы)",
)
ANALYSIS_TITLE_POSITIVE = getattr(
    c,
    "ANALYSIS_TITLE_POSITIVE",
    "Анализ отзывов (позитив)",
)


# ==========================================================
# СТИЛИ EXCEL
# ==========================================================
THIN_SIDE = Side(style="thin")
DOUBLE_SIDE = Side(style="double")

THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)

# Двойные границы итоговых строк
TOTAL_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=DOUBLE_SIDE,
    bottom=DOUBLE_SIDE,
)

TITLE_FONT = Font(name="Calibri", size=14, bold=True)
GROUP_FONT = Font(name="Calibri", size=12, bold=True)
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True)
DATA_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)

GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="D9EAD3",
    end_color="D9EAD3",
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    start_color="FFF2CC",
    end_color="FFF2CC",
)

CENTER_ALIGN = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)

LEFT_ALIGN = Alignment(
    horizontal="left",
    vertical="center",
)


# ==========================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ==========================================================
def _to_int(value: Any) -> int:
    """Безопасно приводит значение к int."""
    try:
        if value is None or pd.isna(value):
            return 0
        return int(value)
    except Exception:
        try:
            return int(float(str(value).replace(",", ".").strip()))
        except Exception:
            return 0


def _to_text(value: Any) -> str:
    """Безопасно приводит значение к тексту."""
    try:
        if value is None or pd.isna(value):
            return ""
        return str(value)
    except Exception:
        return str(value)


def _avg_formula(row: int, start_col: int) -> str:
    """Формула среднего рейтинга через SUMPRODUCT."""
    col_b = get_column_letter(start_col + 1)
    col_f = get_column_letter(start_col + 5)
    col_g = get_column_letter(start_col + 6)

    return (
        f"=IF({col_g}{row}=0,0,"
        f"SUMPRODUCT({col_b}{row}:{col_f}{row},{{1,2,3,4,5}})/{col_g}{row})"
    )


# ==========================================================
# ЗАГОЛОВКИ И СТРОКИ ЛИСТА "ОЦЕНКИ"
# ==========================================================
def _write_headers(
    ws: Worksheet,
    row: int,
    region: str,
    start_col: int,
    has_low_reviews: bool,
    low_header: str,
) -> None:
    """Записывает строку заголовков блока."""
    headers = [
        region,
        "1",
        "2",
        "3",
        "4",
        "5",
        TOTAL_COLUMN,
        AVG_COLUMN,
    ]

    if has_low_reviews:
        headers.append(low_header)

    for col_idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = GREEN_FILL
        cell.border = THIN_BORDER


def _write_restaurant_row(
    ws: Worksheet,
    row: int,
    restaurant: str,
    start_col: int,
    df: pd.DataFrame,
    has_low_reviews: bool,
    low_count: Optional[int],
) -> None:
    """Записывает строку ресторана."""
    row_data = df[df[COL_RESTAURANT] == restaurant]

    if row_data.empty:
        return

    row_data = row_data.iloc[0]

    cell = ws.cell(row=row, column=start_col, value=restaurant)
    cell.font = DATA_FONT
    cell.alignment = LEFT_ALIGN
    cell.border = THIN_BORDER

    for i in range(1, 6):
        col_idx = start_col + i
        cell = ws.cell(
            row=row,
            column=col_idx,
            value=_to_int(row_data.get(str(i), 0)),
        )
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    col_b = get_column_letter(start_col + 1)
    col_f = get_column_letter(start_col + 5)

    cell = ws.cell(
        row=row,
        column=start_col + 6,
        value=f"=SUM({col_b}{row}:{col_f}{row})",
    )
    cell.font = DATA_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

    cell = ws.cell(
        row=row,
        column=start_col + 7,
        value=_avg_formula(row, start_col),
    )
    cell.font = DATA_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    cell.number_format = "0.00"

    if has_low_reviews:
        cell = ws.cell(
            row=row,
            column=start_col + 8,
            value=_to_int(low_count) if low_count is not None else 0,
        )
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _write_total_row(
    ws: Worksheet,
    row: int,
    total_label: str,
    start_col: int,
    first_data_row: int,
    last_data_row: int,
    has_low_reviews: bool,
) -> None:
    """Записывает итоговую строку блока с двойными границами."""
    cell = ws.cell(row=row, column=start_col, value=total_label)
    cell.font = TOTAL_FONT
    cell.alignment = LEFT_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = TOTAL_BORDER

    for i in range(5):
        col_idx = start_col + 1 + i
        col_letter = get_column_letter(col_idx)

        if first_data_row <= last_data_row:
            formula = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        else:
            formula = "0"

        cell = ws.cell(row=row, column=col_idx, value=formula)
        cell.font = TOTAL_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = YELLOW_FILL
        cell.border = TOTAL_BORDER

    col_b = get_column_letter(start_col + 1)
    col_f = get_column_letter(start_col + 5)

    cell = ws.cell(
        row=row,
        column=start_col + 6,
        value=f"=SUM({col_b}{row}:{col_f}{row})",
    )
    cell.font = TOTAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = TOTAL_BORDER

    cell = ws.cell(
        row=row,
        column=start_col + 7,
        value=_avg_formula(row, start_col),
    )
    cell.font = TOTAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = TOTAL_BORDER
    cell.number_format = "0.00"

    if has_low_reviews:
        col_i = get_column_letter(start_col + 8)

        if first_data_row <= last_data_row:
            formula = f"=SUM({col_i}{first_data_row}:{col_i}{last_data_row})"
        else:
            formula = "0"

        cell = ws.cell(row=row, column=start_col + 8, value=formula)
        cell.font = TOTAL_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = YELLOW_FILL
        cell.border = TOTAL_BORDER


def _write_common_restaurant_row(
    ws: Worksheet,
    row: int,
    restaurant: str,
    start_col: int,
    site_row: int,
    agg_row: int,
    geo_row: int,
) -> None:
    """Записывает строку ресторана в общем блоке (формулы из 3 блоков)."""
    cell = ws.cell(row=row, column=start_col, value=restaurant)
    cell.font = DATA_FONT
    cell.alignment = LEFT_ALIGN
    cell.border = THIN_BORDER

    for i in range(5):
        target_col = start_col + 1 + i
        source_col = get_column_letter(2 + i)

        formula = f"={source_col}{site_row}+{source_col}{agg_row}+{source_col}{geo_row}"

        cell = ws.cell(row=row, column=target_col, value=formula)
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    col_b = get_column_letter(start_col + 1)
    col_f = get_column_letter(start_col + 5)

    cell = ws.cell(
        row=row,
        column=start_col + 6,
        value=f"=SUM({col_b}{row}:{col_f}{row})",
    )
    cell.font = DATA_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

    cell = ws.cell(
        row=row,
        column=start_col + 7,
        value=_avg_formula(row, start_col),
    )
    cell.font = DATA_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    cell.number_format = "0.00"


def _write_common_total_row(
    ws: Worksheet,
    row: int,
    total_label: str,
    start_col: int,
    site_total_row: int,
    agg_total_row: int,
    geo_total_row: int,
) -> None:
    """Записывает итоговую строку общего блока с двойными границами."""
    cell = ws.cell(row=row, column=start_col, value=total_label)
    cell.font = TOTAL_FONT
    cell.alignment = LEFT_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = TOTAL_BORDER

    for i in range(5):
        target_col = start_col + 1 + i
        source_col = get_column_letter(2 + i)

        formula = (
            f"={source_col}{site_total_row}"
            f"+{source_col}{agg_total_row}"
            f"+{source_col}{geo_total_row}"
        )

        cell = ws.cell(row=row, column=target_col, value=formula)
        cell.font = TOTAL_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = YELLOW_FILL
        cell.border = TOTAL_BORDER

    col_b = get_column_letter(start_col + 1)
    col_f = get_column_letter(start_col + 5)

    cell = ws.cell(
        row=row,
        column=start_col + 6,
        value=f"=SUM({col_b}{row}:{col_f}{row})",
    )
    cell.font = TOTAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = TOTAL_BORDER

    cell = ws.cell(
        row=row,
        column=start_col + 7,
        value=_avg_formula(row, start_col),
    )
    cell.font = TOTAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = TOTAL_BORDER
    cell.number_format = "0.00"


# ==========================================================
# БЛОКИ ЛИСТА "ОЦЕНКИ"
# ==========================================================
def _write_block(
    ws: Worksheet,
    start_row: int,
    block_title: str,
    df: pd.DataFrame,
    period_label: str,
    has_low_reviews: bool,
    low_counts: Optional[Dict[str, int]],
    low_header: str,
) -> Tuple[Dict[str, int], int]:
    """
    Записывает один блок: Сайт / Агрегаторы / Геосервисы.
    Возвращает карту строк и следующую свободную строку.
    """
    row_map: Dict[str, int] = {}
    max_col = 9 if has_low_reviews else 8

    # Подпись периода
    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=max_col,
    )
    cell = ws.cell(row=start_row, column=1, value=period_label)
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN

    start_row += 2

    # Название блока
    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=max_col,
    )
    cell = ws.cell(row=start_row, column=1, value=block_title)
    cell.font = GROUP_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = GREEN_FILL
    cell.border = THIN_BORDER

    start_row += 1

    # Подзаголовок "Кол-во поставленных звезд"
    ws.merge_cells(
        start_row=start_row,
        start_column=2,
        end_row=start_row,
        end_column=6,
    )
    cell = ws.cell(row=start_row, column=2, value=STAR_SUBTITLE)
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

    start_row += 1

    # Санкт-Петербург
    _write_headers(ws, start_row, HEADER_CITY_SPB, 1, has_low_reviews, low_header)
    start_row += 1

    first_spb_row = start_row

    for restaurant in SPB_ORDER:
        low_count = low_counts.get(restaurant, 0) if low_counts else 0

        _write_restaurant_row(
            ws,
            start_row,
            restaurant,
            1,
            df,
            has_low_reviews,
            low_count,
        )
        row_map[restaurant] = start_row
        start_row += 1

    last_spb_row = start_row - 1

    _write_total_row(
        ws,
        start_row,
        SPB_TOTAL_LABEL,
        1,
        first_spb_row,
        last_spb_row,
        has_low_reviews,
    )
    row_map[SPB_TOTAL_LABEL] = start_row

    start_row += 2

    # Подзаголовок для Тюмени
    ws.merge_cells(
        start_row=start_row,
        start_column=2,
        end_row=start_row,
        end_column=6,
    )
    cell = ws.cell(row=start_row, column=2, value=STAR_SUBTITLE)
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

    start_row += 1

    # Тюмень
    _write_headers(ws, start_row, HEADER_CITY_TMN, 1, has_low_reviews, low_header)
    start_row += 1

    first_tmn_row = start_row

    for restaurant in TMN_ORDER:
        low_count = low_counts.get(restaurant, 0) if low_counts else 0

        _write_restaurant_row(
            ws,
            start_row,
            restaurant,
            1,
            df,
            has_low_reviews,
            low_count,
        )
        row_map[restaurant] = start_row
        start_row += 1

    last_tmn_row = start_row - 1

    _write_total_row(
        ws,
        start_row,
        TMN_TOTAL_LABEL,
        1,
        first_tmn_row,
        last_tmn_row,
        has_low_reviews,
    )
    row_map[TMN_TOTAL_LABEL] = start_row

    start_row += 1

    return row_map, start_row


def _write_common_block(
    ws: Worksheet,
    start_row: int,
    row_map_site: Dict[str, int],
    row_map_agg: Dict[str, int],
    row_map_geo: Dict[str, int],
) -> int:
    """Записывает общий блок справа (формулы из трёх блоков)."""
    start_col = 11

    # Заголовок общего блока
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row,
        end_column=start_col + 7,
    )
    cell = ws.cell(row=start_row, column=start_col, value=BLOCK_TITLE_COMMON)
    cell.font = GROUP_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = GREEN_FILL
    cell.border = THIN_BORDER

    start_row += 1

    # Подзаголовок
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col + 1,
        end_row=start_row,
        end_column=start_col + 5,
    )
    cell = ws.cell(row=start_row, column=start_col + 1, value=STAR_SUBTITLE)
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

    start_row += 1

    # Санкт-Петербург
    _write_headers(ws, start_row, HEADER_CITY_SPB, start_col, False, "")
    start_row += 1

    for restaurant in SPB_ORDER:
        _write_common_restaurant_row(
            ws,
            start_row,
            restaurant,
            start_col,
            row_map_site[restaurant],
            row_map_agg[restaurant],
            row_map_geo[restaurant],
        )
        start_row += 1

    _write_common_total_row(
        ws,
        start_row,
        SPB_TOTAL_LABEL,
        start_col,
        row_map_site[SPB_TOTAL_LABEL],
        row_map_agg[SPB_TOTAL_LABEL],
        row_map_geo[SPB_TOTAL_LABEL],
    )

    start_row += 2

    # Подзаголовок для Тюмени
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col + 1,
        end_row=start_row,
        end_column=start_col + 5,
    )
    cell = ws.cell(row=start_row, column=start_col + 1, value=STAR_SUBTITLE)
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER

    start_row += 1

    # Тюмень
    _write_headers(ws, start_row, HEADER_CITY_TMN, start_col, False, "")
    start_row += 1

    for restaurant in TMN_ORDER:
        _write_common_restaurant_row(
            ws,
            start_row,
            restaurant,
            start_col,
            row_map_site[restaurant],
            row_map_agg[restaurant],
            row_map_geo[restaurant],
        )
        start_row += 1

    _write_common_total_row(
        ws,
        start_row,
        TMN_TOTAL_LABEL,
        start_col,
        row_map_site[TMN_TOTAL_LABEL],
        row_map_agg[TMN_TOTAL_LABEL],
        row_map_geo[TMN_TOTAL_LABEL],
    )

    start_row += 1

    return start_row


# ==========================================================
# ЛИСТ "АНАЛИЗ ОТЗЫВОВ"
# ==========================================================
def _write_analysis_table(
    ws: Worksheet,
    start_row: int,
    table_title: str,
    df: Optional[pd.DataFrame],
) -> int:
    """
    Записывает одну таблицу анализа (жалобы или позитив).
    Возвращает следующую свободную строку.
    """
    if df is None or len(df.columns) == 0:
        ws.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=3,
        )
        cell = ws.cell(row=start_row, column=1, value=table_title)
        cell.font = TITLE_FONT
        cell.alignment = CENTER_ALIGN
        return start_row + 2

    max_col = max(3, len(df.columns))

    # Заголовок таблицы
    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=max_col,
    )
    cell = ws.cell(row=start_row, column=1, value=table_title)
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN

    start_row += 1

    # Заголовки колонок
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = GREEN_FILL
        cell.border = THIN_BORDER

    start_row += 1

    # Данные
    for _, row_data in df.iterrows():
        first_value = _to_text(row_data.iloc[0]) if len(row_data) > 0 else ""
        is_total_row = first_value.startswith("Всего")

        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_idx == 1:
                value: Any = _to_text(row_data[col_name])
            else:
                value = _to_int(row_data[col_name])

            cell = ws.cell(row=start_row, column=col_idx, value=value)

            if is_total_row:
                cell.font = TOTAL_FONT
                cell.fill = YELLOW_FILL
                cell.border = TOTAL_BORDER
                cell.alignment = (
                    CENTER_ALIGN if col_idx > 1 else LEFT_ALIGN
                )
            else:
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = (
                    CENTER_ALIGN if col_idx > 1 else LEFT_ALIGN
                )

        start_row += 1

    return start_row


# ==========================================================
# ПУБЛИЧНАЯ ФУНКЦИЯ
# ==========================================================
def build_kr_excel(
    report_data: KRReportData,
    settings: KRReportSettings,
) -> io.BytesIO:
    """
    Собирает итоговый Excel-файл КР и возвращает BytesIO.
    """
    period_label = getattr(settings, "period_label", "") or ""
    is_month = bool(getattr(settings, "is_month", False))

    threshold = getattr(settings, "price_threshold", None)

    if threshold is None:
        month_settings = getattr(settings, "month", None)
        threshold = getattr(month_settings, "price_threshold", 749)

    try:
        threshold = int(threshold)
    except Exception:
        threshold = 749

    if is_month:
        low_header_template = getattr(
            c,
            "LOW_REVIEW_HEADER_TEMPLATE",
            "Отзывы ≤ {threshold} руб (включительно)",
        )
        low_header = low_header_template.format(threshold=threshold)
    else:
        low_header = LOW_REVIEW_COLUMN

    stats_site = getattr(report_data, "stats_site", None)
    stats_agg = getattr(report_data, "stats_agg", None)
    stats_geo = getattr(report_data, "stats_geo", None)

    low_counts: Dict[str, int] = {}

    if (
        is_month
        and stats_site is not None
        and LOW_REVIEW_COLUMN in getattr(stats_site, "columns", [])
    ):
        low_counts = {
            str(rest): _to_int(val)
            for rest, val in zip(
                stats_site[COL_RESTAURANT],
                stats_site[LOW_REVIEW_COLUMN],
            )
        }

    workbook = Workbook()

    ws = workbook.active
    ws.title = EXCEL_SHEET_RATINGS

    # Блок "Сайт/приложение"
    start_row = 1

    row_map_site, next_row = _write_block(
        ws,
        start_row,
        BLOCK_TITLE_SITE,
        stats_site,
        period_label,
        is_month,
        low_counts,
        low_header,
    )

    # Блок "Агрегаторы"
    agg_start_row = next_row + 2

    row_map_agg, next_row = _write_block(
        ws,
        agg_start_row,
        BLOCK_TITLE_AGG,
        stats_agg,
        period_label,
        False,
        None,
        "",
    )

    # Блок "Геосервисы"
    geo_start_row = next_row + 2

    row_map_geo, next_row = _write_block(
        ws,
        geo_start_row,
        BLOCK_TITLE_GEO,
        stats_geo,
        period_label,
        False,
        None,
        "",
    )

    # Общий блок справа
    _write_common_block(
        ws,
        1,
        row_map_site,
        row_map_agg,
        row_map_geo,
    )

    # Ширины колонок листа "Оценки"
    ws.column_dimensions["A"].width = 22

    for col_idx in range(2, 10):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11

    ws.column_dimensions["J"].width = 3
    ws.column_dimensions["K"].width = 22

    for col_idx in range(12, 19):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11

    # Лист "Анализ отзывов"
    ws2 = workbook.create_sheet(EXCEL_SHEET_ANALYSIS)

    next_row = _write_analysis_table(
        ws2,
        1,
        ANALYSIS_TITLE_COMPLAINTS,
        getattr(report_data, "complaints", None),
    )

    _write_analysis_table(
        ws2,
        next_row + 2,
        ANALYSIS_TITLE_POSITIVE,
        getattr(report_data, "positives", None),
    )

    # Ширины колонок листа "Анализ отзывов"
    ws2.column_dimensions["A"].width = 24

    complaints = getattr(report_data, "complaints", None)
    positives = getattr(report_data, "positives", None)

    complaint_cols = len(complaints.columns) if complaints is not None else 0
    positive_cols = len(positives.columns) if positives is not None else 0
    analysis_max_col = max(complaint_cols, positive_cols, 1)

    for col_idx in range(2, analysis_max_col + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 18

    # ======================================================
    # СОХРАНЕНИЕ КНИГИ В БУФЕР
    # ======================================================
    output = io.BytesIO()

    workbook.save(output)
    output.seek(0)

    return output


# Алиас для удобства импорта
build_excel = build_kr_excel