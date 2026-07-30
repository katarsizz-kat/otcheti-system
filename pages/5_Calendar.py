"""Страница календаря событий."""
import streamlit as st
from datetime import datetime, date, timedelta
from streamlit_calendar import calendar
import utils.supabase_db as db
import config.calendar as cfg
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io

st.set_page_config(page_title="Календарь событий", page_icon="📅", layout="wide")

# =============================================================================
# Вспомогательные функции
# =============================================================================

def parse_date(d):
    """Безопасный парсинг даты из строки или объекта date."""
    if isinstance(d, date):
        return d
    if d is None:
        return date.today()
    
    d_str = str(d).strip()
    
    # Убираем время и часовой пояс если есть
    if 'T' in d_str:
        d_str = d_str.split('T')[0]
    if '+' in d_str[10:]:
        d_str = d_str.split('+')[0]
    if d_str.endswith('Z'):
        d_str = d_str[:-1]
    
    try:
        return date.fromisoformat(d_str)
    except ValueError:
        st.warning(f"⚠️ Не удалось распарсить дату: {d}")
        return date.today()

def get_key_by_value(d: dict, target_val, default=None):
    """Получить ключ словаря по его значению."""
    for k, v in d.items():
        if v == target_val:
            return k
    return default if default else list(d.keys())[0]

# =============================================================================
# Кэширование данных
# =============================================================================

@st.cache_data(ttl=10)
def get_cached_events():
    """Получить все события с кэшированием."""
    return db.get_all_events()

@st.cache_data(ttl=10)
def get_cached_event_by_id(event_id):
    """Получить событие по ID с кэшированием."""
    return db.get_event_by_id(event_id)

def clear_cache():
    """Очистить кэш."""
    get_cached_events.clear()
    get_cached_event_by_id.clear()

# =============================================================================
# Инициализация session state
# =============================================================================

if 'selected_event' not in st.session_state:
    st.session_state.selected_event = None
if 'edit_mode' not in st.session_state:
    st.session_state.edit_mode = False
if 'show_edit_form' not in st.session_state:
    st.session_state.show_edit_form = False
if 'events_to_delete' not in st.session_state:
    st.session_state.events_to_delete = []
if 'delete_confirm_id' not in st.session_state:
    st.session_state.delete_confirm_id = None
if 'success_message' not in st.session_state:
    st.session_state.success_message = None

form_defaults = {
    'data_title': "",
    'data_start_date': date.today(),
    'data_end_date': date.today(),
    'data_category': list(cfg.EVENT_CATEGORIES.keys())[0],
    'data_location': "Все рестораны",
    'data_location_custom': [],
    'data_reminder_on_start': True,
    'data_reminder_days_start': 0,
    'data_reminder_days_end': 0,
    'data_use_custom_date': False,
    'data_reminder_custom_date': date.today(),
    'data_recurrence': list(cfg.RECURRENCE_TYPES.values())[0] if cfg.RECURRENCE_TYPES else 'none'
}

for key, value in form_defaults.items():
    if key not in st.session_state:
        st.session_state[key] = value

# =============================================================================
# Заголовок страницы
# =============================================================================

st.title("📅 Календарь событий")

if st.session_state.success_message:
    st.success(st.session_state.success_message, icon="✅")
    st.session_state.success_message = None

# =============================================================================
# Функция для сброса формы
# =============================================================================

def reset_form():
    """Сбросить форму к значениям по умолчанию."""
    for key, value in form_defaults.items():
        st.session_state[key] = value
    st.session_state.edit_mode = False
    st.session_state.selected_event = None
    st.session_state.show_edit_form = False

# =============================================================================
# Боковая панель с фильтрами
# =============================================================================

with st.sidebar:
    st.header(" Фильтры")
    
    selected_category = st.multiselect(
        "Категория",
        options=list(cfg.EVENT_CATEGORIES.keys()),
        default=[],
        key="sidebar_filter_category"
    )
    
    selected_location = st.selectbox(
        "Локация",
        options=list(cfg.LOCATIONS.keys()),
        key="sidebar_filter_location"
    )
    
    search_query = st.text_input("Поиск по названию", key="sidebar_filter_search")
    
    date_range = st.date_input(
        "Диапазон дат",
        value=(date.today(), date.today() + timedelta(days=365)),
        key="sidebar_filter_date_range"
    )
    
    if isinstance(date_range, tuple) and len(date_range) == 2:
        start_date_filter, end_date_filter = date_range
    else:
        start_date_filter = end_date_filter = date_range if isinstance(date_range, date) else date.today()
    
    st.divider()
    st.header(" Экспорт")
    
    if st.button("Экспорт в Excel", use_container_width=True, key="btn_export"):
        with st.spinner("Генерация Excel..."):
            events = get_cached_events()
            if events:
                wb = Workbook()
                ws = wb.active
                ws.title = "События"
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                spb_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                tyumen_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                center_align = Alignment(horizontal="center", vertical="center")
                double_border = Border(
                    top=Side(style="double"), bottom=Side(style="double"),
                    left=Side(style="thin"), right=Side(style="thin")
                )
                
                headers = ["Название", "Категория", "Локация", "Дата начала", "Дата окончания", "Статус (IF)", "Вес (SUMPRODUCT)"]
                ws.append(headers)
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                sorted_events = sorted(events, key=lambda x: x.get('location_type', 'all'))
                
                for idx, event in enumerate(sorted_events, start=2):
                    loc_type = event.get('location_type', 'all')
                    loc_display = event.get('location_custom') or loc_type
                    
                    fill = spb_fill if loc_type in ['spb', 'all'] else tyumen_fill
                    status_formula = f'=IF(D{idx}<>"", "Активно", "Архив")'
                    sumproduct_formula = f'=SUMPRODUCT(--(C$2:C${idx}="{loc_display}"))'
                    
                    ws.append([
                        event['title'],
                        event['category'],
                        loc_display,
                        event['start_date'],
                        event['end_date'],
                        status_formula,
                        sumproduct_formula
                    ])
                    
                    for cell in ws[idx]:
                        cell.fill = fill
                        cell.alignment = center_align
                
                last_row = len(sorted_events) + 1
                ws.append(["ИТОГО СОБЫТИЙ", "", "", "", "", "", f'=SUM(G2:G{last_row-1})'])
                
                for cell in ws[last_row]:
                    cell.font = Font(bold=True, color="000000")
                    cell.border = double_border
                    cell.alignment = center_align
                
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
                
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                st.download_button(
                    label="📥 Скачать Excel",
                    data=buffer,
                    file_name=f"calendar_events_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_download_excel"
                )
            else:
                st.warning("Нет событий для экспорта")

# =============================================================================
# Получение и фильтрация событий
# =============================================================================

all_events = get_cached_events()
filtered_events = all_events if all_events else []

if selected_category:
    filtered_events = [e for e in filtered_events if e['category'] in selected_category]

if selected_location != "Все рестораны":
    location_code = cfg.LOCATIONS.get(selected_location, "all")
    if location_code == "spb":
        filtered_events = [e for e in filtered_events if e.get('location_type') in ['spb', 'all']]
    elif location_code == "tyumen":
        filtered_events = [e for e in filtered_events if e.get('location_type') in ['tyumen', 'all']]

if search_query:
    filtered_events = [e for e in filtered_events if search_query.lower() in str(e['title']).lower()]

filtered_events = [
    e for e in filtered_events
    if parse_date(e['start_date']) <= end_date_filter
    and parse_date(e['end_date']) >= start_date_filter
]

# =============================================================================
# Вкладки
# =============================================================================

tab_calendar, tab_list, tab_add = st.tabs(["📅 Календарь", "📋 Список", "➕ Добавить событие"])

# =============================================================================
# Вкладка 1: Календарь
# =============================================================================

with tab_calendar:
    calendar_events = []
    for event in filtered_events:
        color = cfg.EVENT_CATEGORIES.get(event['category'], '#808080')
        calendar_events.append({
            "title": event['title'],
            "start": str(event['start_date']),
            "end": str(event['end_date']),
            "id": str(event['id']),
            "color": color,
            "extendedProps": {
                "category": event['category'],
                "location": event.get('location_type', 'all')
            }
        })
    
    calendar_options = {
        "editable": False,
        "selectable": False,
        "events": calendar_events,
        "locale": "ru",
        "firstDay": 1,
        "headerToolbar": {
            "left": "prev,next today",
            "center": "title",
            "right": "dayGridMonth"
        },
        "initialView": "dayGridMonth",
        "dayMaxEventRows": 3,
    }
    
    calendar(events=calendar_events, options=calendar_options, key="main_calendar")
    st.info("️ Для управления событиями перейдите во вкладку **📋 Список**")

# =============================================================================
# Вкладка 2: Список событий
# =============================================================================

with tab_list:
    st.subheader(" Список событий")
    
    if st.session_state.show_edit_form and st.session_state.selected_event:
        st.divider()
        st.subheader(f"✏️ Редактирование: {st.session_state.selected_event['title']}")
        
        title = st.text_input("Название события", value=st.session_state.data_title, key="edit_input_title")
        
        col1, col2 = st.columns(2)
        with col1:
            start_date = st.date_input("Дата начала", value=st.session_state.data_start_date, key="edit_input_start")
        with col2:
            is_indefinite = st.checkbox("♾️ Однодневное событие", value=(st.session_state.data_start_date == st.session_state.data_end_date), key="edit_input_indefinite")
            if is_indefinite:
                end_date = start_date
            else:
                end_date = st.date_input("Дата окончания", value=st.session_state.data_end_date, key="edit_input_end")
        
        category = st.selectbox("Категория", options=list(cfg.EVENT_CATEGORIES.keys()), index=list(cfg.EVENT_CATEGORIES.keys()).index(st.session_state.data_category) if st.session_state.data_category in cfg.EVENT_CATEGORIES else 0, key="edit_input_category")
        location_type = st.selectbox("Локация", options=list(cfg.LOCATIONS.keys()), index=list(cfg.LOCATIONS.keys()).index(st.session_state.data_location) if st.session_state.data_location in cfg.LOCATIONS else 0, key="edit_input_location")
        
        location_custom = None
        if location_type == "Выбрать конкретные":
            location_custom_list = st.multiselect("📍 Выберите рестораны", options=getattr(cfg, 'ALL_RESTAURANTS', []), default=st.session_state.data_location_custom, key="edit_input_location_custom")
            location_custom = ", ".join(location_custom_list) if location_custom_list else None
        
        st.divider()
        st.subheader("🔔 Напоминания")
        reminder_on_start_day = st.checkbox("Напомнить в день начала события", value=st.session_state.data_reminder_on_start, key="edit_input_reminder_start")
        reminder_days_before_start = st.number_input("За сколько дней до НАЧАЛА напомнить", min_value=0, max_value=90, value=st.session_state.data_reminder_days_start, key="edit_input_reminder_days_start")
        reminder_days_before_end = st.number_input("За сколько дней до ОКОНЧАНИЯ напомнить", min_value=0, max_value=90, value=st.session_state.data_reminder_days_end, key="edit_input_reminder_days_end")
        use_custom_date = st.checkbox("Напомнить в конкретную дату", value=st.session_state.data_use_custom_date, key="edit_input_custom_date")
        reminder_custom_date = None
        if use_custom_date:
            reminder_custom_date = st.date_input("Выберите дату напоминания", value=st.session_state.data_reminder_custom_date, key="edit_input_reminder_custom")
        
        recurrence_type = st.selectbox("Повторяемость", options=list(cfg.RECURRENCE_TYPES.keys()), index=list(cfg.RECURRENCE_TYPES.values()).index(st.session_state.data_recurrence) if st.session_state.data_recurrence in cfg.RECURRENCE_TYPES.values() else 0, key="edit_input_recurrence")
        
        col_save, col_cancel = st.columns(2)
        with col_save:
            if st.button("💾 Сохранить изменения", use_container_width=True, type="primary", key="btn_save_edit"):
                if not title:
                    st.error("❌ Введите название события")
                elif start_date > end_date:
                    st.error("❌ Дата начала не может быть позже даты окончания")
                else:
                    location_code = cfg.LOCATIONS.get(location_type, "all")
                    if not use_custom_date:
                        reminder_custom_date = None
                    
                    try:
                        db.update_event(
                            event_id=st.session_state.selected_event['id'],
                            title=title,
                            start_date=start_date,
                            end_date=end_date,
                            category=category,
                            location_type=location_code,
                            location_custom=location_custom,
                            reminder_days_before_start=reminder_days_before_start,
                            reminder_on_start_day=1 if reminder_on_start_day else 0,
                            reminder_days_before_end=reminder_days_before_end,
                            reminder_custom_date=reminder_custom_date,
                            recurrence_type=cfg.RECURRENCE_TYPES.get(recurrence_type, 'none')
                        )
                        clear_cache()
                        st.session_state.success_message = f"✅ Событие '{title}' успешно обновлено!"
                        reset_form()
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Ошибка при сохранении: {str(e)}")
        
        with col_cancel:
            if st.button("❌ Отменить", use_container_width=True, key="btn_cancel_edit_inline"):
                reset_form()
                st.rerun()
        
        st.divider()
    
    if st.session_state.delete_confirm_id:
        event_to_delete = next((e for e in filtered_events if e['id'] == st.session_state.delete_confirm_id), None)
        if event_to_delete:
            st.warning(f"⚠️ Вы уверены, что хотите удалить событие **'{event_to_delete['title']}'**?")
            col_yes, col_no = st.columns(2)
            with col_yes:
                if st.button("🗑️ Да, удалить", use_container_width=True, type="primary", key="confirm_delete_yes"):
                    db.delete_event(st.session_state.delete_confirm_id)
                    clear_cache()
                    st.session_state.success_message = f"✅ Событие '{event_to_delete['title']}' удалено!"
                    st.session_state.delete_confirm_id = None
                    st.rerun()
            with col_no:
                if st.button("❌ Отмена", use_container_width=True, key="confirm_delete_no"):
                    st.session_state.delete_confirm_id = None
                    st.rerun()
            st.divider()
    
    today = date.today()
    future_events = [e for e in filtered_events if parse_date(e['end_date']) >= today]
    future_events.sort(key=lambda x: parse_date(x['start_date']))
    
    if future_events:
        select_all = st.checkbox("Выбрать все", key="select_all_events")
        
        if select_all:
            st.session_state.events_to_delete = [e['id'] for e in future_events]
        
        st.write(f"**Всего событий:** {len(future_events)}")
        st.divider()
        
        for event in future_events:
            col_checkbox, col_info, col_actions = st.columns([0.5, 7, 2])
            
            with col_checkbox:
                is_selected = st.checkbox(
                    "",
                    value=event['id'] in st.session_state.events_to_delete,
                    key=f"chk_{event['id']}",
                    label_visibility="collapsed"
                )
                if is_selected and event['id'] not in st.session_state.events_to_delete:
                    st.session_state.events_to_delete.append(event['id'])
                elif not is_selected and event['id'] in st.session_state.events_to_delete:
                    st.session_state.events_to_delete.remove(event['id'])
            
            with col_info:
                location_display = event.get('location_custom') or event.get('location_type', 'Не указано')
                st.markdown(f"**{event['title']}**")
                st.caption(f"📅 {event['start_date']} | {event['category']} | {location_display}")
            
            with col_actions:
                if st.button("✏️", key=f"edit_{event['id']}", help="Редактировать"):
                    st.session_state.selected_event = event
                    st.session_state.edit_mode = True
                    st.session_state.show_edit_form = True
                    
                    st.session_state.data_title = event['title']
                    st.session_state.data_start_date = parse_date(event['start_date'])
                    st.session_state.data_end_date = parse_date(event['end_date'])
                    st.session_state.data_category = event['category']
                    st.session_state.data_location = get_key_by_value(cfg.LOCATIONS, event.get('location_type', 'all'), "Все рестораны")
                    
                    if event.get('location_custom'):
                        st.session_state.data_location_custom = [r.strip() for r in event['location_custom'].split(',')]
                    else:
                        st.session_state.data_location_custom = []
                    
                    st.session_state.data_reminder_on_start = bool(event.get('reminder_on_start_day', True))
                    st.session_state.data_reminder_days_start = event.get('reminder_days_before_start', 0)
                    st.session_state.data_reminder_days_end = event.get('reminder_days_before_end', 0)
                    
                    if event.get('reminder_custom_date'):
                        st.session_state.data_use_custom_date = True
                        st.session_state.data_reminder_custom_date = parse_date(event['reminder_custom_date'])
                    else:
                        st.session_state.data_use_custom_date = False
                    
                    st.session_state.data_recurrence = get_key_by_value(cfg.RECURRENCE_TYPES, event.get('recurrence_type', 'none'), list(cfg.RECURRENCE_TYPES.values())[0])
                    st.rerun()
                
                if st.button("🗑️", key=f"delete_{event['id']}", help="Удалить"):
                    st.session_state.delete_confirm_id = event['id']
                    st.rerun()
            
            st.divider()
        
        if st.session_state.events_to_delete:
            st.warning(f"️ Выбрано событий для удаления: {len(st.session_state.events_to_delete)}")
            col_confirm, col_cancel = st.columns(2)
            with col_confirm:
                if st.button("✅ Подтвердить удаление", type="primary", key="confirm_bulk_delete"):
                    count = len(st.session_state.events_to_delete)
                    for event_id in st.session_state.events_to_delete:
                        db.delete_event(event_id)
                    clear_cache()
                    st.session_state.success_message = f"✅ Удалено событий: {count}"
                    st.session_state.events_to_delete = []
                    st.rerun()
            with col_cancel:
                if st.button("❌ Отмена", key="cancel_bulk_delete"):
                    st.session_state.events_to_delete = []
                    st.rerun()
    else:
        st.info("Нет будущих событий для отображения")

# =============================================================================
# Вкладка 3: Добавить событие
# =============================================================================

with tab_add:
    st.subheader("➕ Добавить новое событие")
    
    title = st.text_input("Название события", value=st.session_state.data_title, key="input_title")
    
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Дата начала", value=st.session_state.data_start_date, key="input_start_date")
    with col2:
        is_indefinite = st.checkbox("♾️ Однодневное событие", value=(st.session_state.data_start_date == st.session_state.data_end_date), key="input_indefinite")
        if is_indefinite:
            end_date = start_date
            st.caption("✅ Дата окончания = дате начала")
        else:
            end_date = st.date_input("Дата окончания", value=st.session_state.data_end_date, key="input_end_date")
    
    category = st.selectbox("Категория", options=list(cfg.EVENT_CATEGORIES.keys()), index=list(cfg.EVENT_CATEGORIES.keys()).index(st.session_state.data_category) if st.session_state.data_category in cfg.EVENT_CATEGORIES else 0, key="input_category")
    location_type = st.selectbox("Локация", options=list(cfg.LOCATIONS.keys()), index=list(cfg.LOCATIONS.keys()).index(st.session_state.data_location) if st.session_state.data_location in cfg.LOCATIONS else 0, key="input_location")
    
    location_custom = None
    if location_type == "Выбрать конкретные":
        location_custom_list = st.multiselect("📍 Выберите рестораны", options=getattr(cfg, 'ALL_RESTAURANTS', []), default=st.session_state.data_location_custom, key="input_location_custom")
        location_custom = ", ".join(location_custom_list) if location_custom_list else None
    
    st.divider()
    st.subheader("🔔 Напоминания")
    st.caption("Выберите один или несколько способов напоминания")
    
    reminder_on_start_day = st.checkbox("Напомнить в день начала события", value=st.session_state.data_reminder_on_start, key="input_reminder_start")
    reminder_days_before_start = st.number_input("За сколько дней до НАЧАЛА напомнить (0 = не напоминать)", min_value=0, max_value=90, value=st.session_state.data_reminder_days_start, key="input_reminder_days_start")
    reminder_days_before_end = st.number_input("За сколько дней до ОКОНЧАНИЯ напомнить (0 = не напоминать)", min_value=0, max_value=90, value=st.session_state.data_reminder_days_end, key="input_reminder_days_end")
    use_custom_date = st.checkbox("Напомнить в конкретную дату", value=st.session_state.data_use_custom_date, key="input_custom_date")
    reminder_custom_date = None
    if use_custom_date:
        reminder_custom_date = st.date_input("Выберите дату напоминания", value=st.session_state.data_reminder_custom_date, key="input_reminder_custom")
    
    recurrence_type = st.selectbox("Повторяемость", options=list(cfg.RECURRENCE_TYPES.keys()), index=list(cfg.RECURRENCE_TYPES.values()).index(st.session_state.data_recurrence) if st.session_state.data_recurrence in cfg.RECURRENCE_TYPES.values() else 0, key="input_recurrence")
    
    col_save, col_cancel = st.columns(2)
    with col_save:
        if st.button("💾 Сохранить", use_container_width=True, type="primary", key="btn_save_event"):
            if not title:
                st.error("❌ Введите название события")
            elif start_date > end_date:
                st.error(f"❌ Дата начала ({start_date}) не может быть позже даты окончания ({end_date})")
            else:
                location_code = cfg.LOCATIONS.get(location_type, "all")
                if not use_custom_date:
                    reminder_custom_date = None
                
                try:
                    st.write(f"**Отладка:** Пытаемся сохранить событие...")
                    st.write(f"- title: {title}")
                    st.write(f"- start_date: {start_date}")
                    st.write(f"- end_date: {end_date}")
                    st.write(f"- category: {category}")
                    st.write(f"- location_type: {location_code}")
                    
                    event_id = db.add_event(
                        title=title,
                        start_date=start_date,
                        end_date=end_date,
                        category=category,
                        location_type=location_code,
                        location_custom=location_custom,
                        reminder_days_before_start=reminder_days_before_start,
                        reminder_on_start_day=1 if reminder_on_start_day else 0,
                        reminder_days_before_end=reminder_days_before_end,
                        reminder_custom_date=reminder_custom_date,
                        recurrence_type=cfg.RECURRENCE_TYPES.get(recurrence_type, 'none')
                    )
                    
                    st.write(f"✅ Событие сохранено с ID: {event_id}")
                    
                    clear_cache()
                    st.session_state.success_message = f"✅ Событие '{title}' успешно добавлено!"
                    reset_form()
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Ошибка при сохранении: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())
    
    with col_cancel:
        if st.button("🔄 Очистить форму", use_container_width=True, key="btn_clear_form"):
            reset_form()
            st.rerun()
