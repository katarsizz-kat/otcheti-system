"""Страница календаря событий."""
import streamlit as st
import pandas as pd
import io
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import streamlit.components.v1 as components
from streamlit_calendar import calendar
import db
import config as cfg

# =============================================================================
# КОНФИГУРАЦИЯ СТРАНИЦЫ
# =============================================================================
st.set_page_config(page_title="📅 Календарь событий", page_icon="📅", layout="wide")

from styles import apply_subtle_theme
from config.greetings import get_current_greeting
from config.holidays import get_today_holiday

# =============================================================================
# ПРИМЕНЯЕМ ПРИГЛУШЁННУЮ ТЕМУ
# =============================================================================
greeting_data = get_current_greeting()
holiday = get_today_holiday()
holiday_effects = holiday.get("effects") if holiday and isinstance(holiday, dict) else None
apply_subtle_theme(greeting_data["theme"], holiday_effects)

# =============================================================================
# ОСВЕТЛЕНИЕ ФОНА (УСИЛЕННОЕ) + СТИЛИ
# =============================================================================
st.markdown("""
<style>
/* Усиленное осветление фона — 0.35 вместо 0.15 */
.stApp::before {
    content: '';
    position: fixed;
    top: 0; left: 0;
    width: 100%; height: 100%;
    background: rgba(255, 255, 255, 0.35);
    z-index: 0;
    pointer-events: none;
}

/* ===== ЗАГОЛОВОК СТРАНИЦЫ ===== */
.header-block {
    background: linear-gradient(135deg, #005F6B 0%, #008B9A 100%);
    padding: 32px;
    border-radius: 20px;
    margin-bottom: 24px;
    box-shadow: 0 8px 24px rgba(0,0,0,0.15);
    border: 2px solid rgba(255,255,255,0.3);
}
.header-block h1 {
    margin: 0 0 8px 0;
    color: #FFFFFF !important;
    font-size: 36px;
    font-weight: 700;
    text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
}
.header-block p {
    margin: 0 0 8px 0;
    color: #FFFFFF !important;
    font-size: 18px;
    text-shadow: 1px 1px 2px rgba(0,0,0,0.2);
}

/* ===== БЛОК КАЛЕНДАРЯ ===== */
.calendar-block {
    background: rgba(255, 255, 255, 0.95);
    padding: 20px;
    border-radius: 16px;
    margin-bottom: 24px;
    border: 2px solid rgba(0, 95, 107, 0.2);
    box-shadow: 0 4px 12px rgba(0,0,0,0.08);
}

/* ===== БЛОК ФИЛЬТРОВ (компактный, аккуратный) ===== */
.filters-block {
    background: rgba(255, 255, 255, 0.95);
    padding: 20px 24px;
    border-radius: 16px;
    margin-bottom: 24px;
    border: 2px solid #005F6B;
    box-shadow: 0 4px 16px rgba(0, 95, 107, 0.15);
}
.filters-block h3 {
    font-size: 20px;
    font-weight: 700;
    color: #005F6B !important;
    margin: 0 0 16px 0;
}
.filters-block label {
    font-size: 13px !important;
    font-weight: 600 !important;
    color: #2C3E50 !important;
    margin-bottom: 4px !important;
}

/* Компактные поля фильтров */
.filters-block .stSelectbox > div > div,
.filters-block .stMultiselect > div > div,
.filters-block .stTextInput > div > div > input,
.filters-block .stDateInput > div > div > input {
    border: 1.5px solid rgba(0, 95, 107, 0.3) !important;
    border-radius: 8px !important;
    background: #FFFFFF !important;
    font-size: 13px !important;
    min-height: 36px !important;
}
.filters-block .stSelectbox label,
.filters-block .stMultiselect label,
.filters-block .stTextInput label,
.filters-block .stDateInput label {
    font-size: 12px !important;
    color: #005F6B !important;
}

/* ===== БЛОК ЭКСПОРТА ===== */
.export-block {
    background: rgba(255, 255, 255, 0.95);
    padding: 20px 24px;
    border-radius: 16px;
    margin-bottom: 24px;
    border: 2px solid rgba(0, 95, 107, 0.2);
    box-shadow: 0 4px 12px rgba(0,0,0,0.08);
}
.export-block h3 {
    font-size: 20px;
    font-weight: 700;
    color: #005F6B !important;
    margin: 0 0 12px 0;
}

/* ===== ВКЛАДКИ ===== */
div[data-testid="stTabs"] { margin-bottom: 24px !important; }
div[data-testid="stTabs"] button[kind="secondary"] {
    background: rgba(255, 255, 255, 0.7) !important;
    color: #2C3E50 !important;
    border: 2px solid rgba(0, 95, 107, 0.3) !important;
    border-radius: 10px !important;
    padding: 10px 20px !important;
    font-size: 15px !important;
    font-weight: 600 !important;
    margin-right: 6px !important;
}
div[data-testid="stTabs"] button[kind="secondary"]:hover {
    background: rgba(255, 255, 255, 0.95) !important;
    border-color: #005F6B !important;
}
div[data-testid="stTabs"] button[aria-selected="true"] {
    background: #005F6B !important;
    color: #FFFFFF !important;
    border-color: #005F6B !important;
}

/* ===== КНОПКИ ===== */
.stButton>button {
    background: #005F6B !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-size: 14px !important;
    padding: 6px 14px !important;
}
.stButton>button:hover {
    filter: brightness(1.1) !important;
    transform: translateY(-1px) !important;
}

/* ===== РАЗДЕЛИТЕЛЬ ===== */
hr {
    border: none !important;
    border-top: 2px solid rgba(0, 95, 107, 0.2) !important;
    margin: 16px 0 !important;
}

/* ===== КАРТОЧКИ СОБЫТИЙ ===== */
.event-card {
    background: #FFFFFF;
    border-left: 4px solid #005F6B;
    padding: 12px 16px;
    margin-bottom: 10px;
    border-radius: 8px;
    box-shadow: 0 2px 6px rgba(0,0,0,0.06);
}
.event-card-completed {
    background: #F5F5F5;
    border-left: 4px solid #9E9E9E;
    padding: 12px 16px;
    margin-bottom: 10px;
    border-radius: 8px;
    opacity: 0.75;
}
</style>
""", unsafe_allow_html=True)

# =============================================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =============================================================================
def parse_date(d):
    if isinstance(d, date): return d
    try: return pd.to_datetime(str(d)).date()
    except: return date.today()

def get_key_by_value(d, value, default):
    for k, v in d.items():
        if v == value: return k
    return default

def get_recurrence_key(recurrence_value):
    mapping = {
        'none': 'Не повторяется', 'daily': 'Ежедневно', 'ежедневно': 'Ежедневно',
        'weekly': 'Еженедельно', 'еженедельно': 'Еженедельно',
        'monthly': 'Ежемесячно', 'ежемесячно': 'Ежемесячно',
        'yearly': 'Ежегодно', 'ежегодно': 'Ежегодно'
    }
    return mapping.get(recurrence_value, 'Не повторяется')

# =============================================================================
# КЭШИРОВАНИЕ
# =============================================================================
@st.cache_data(ttl=10)
def get_cached_events():
    return db.get_all_events()

@st.cache_data(ttl=10)
def get_cached_completed_events():
    return db.get_completed_events()

def clear_cache():
    get_cached_events.clear()
    get_cached_completed_events.clear()

# =============================================================================
# SESSION STATE
# =============================================================================
if 'selected_event' not in st.session_state: st.session_state.selected_event = None
if 'edit_mode' not in st.session_state: st.session_state.edit_mode = False
if 'show_edit_form' not in st.session_state: st.session_state.show_edit_form = False
if 'events_to_delete' not in st.session_state: st.session_state.events_to_delete = []
if 'delete_confirm_id' not in st.session_state: st.session_state.delete_confirm_id = None
if 'success_message' not in st.session_state: st.session_state.success_message = None

form_defaults = {
    'data_title': "", 'data_start_date': date.today(), 'data_end_date': date.today(),
    'data_category': list(cfg.EVENT_CATEGORIES.keys())[0],
    'data_location': "Все рестораны", 'data_location_custom': [],
    'data_reminder_on_start': True, 'data_reminder_days_start': 0,
    'data_reminder_days_end': 0, 'data_use_custom_date': False,
    'data_reminder_custom_date': date.today(), 'data_recurrence': 'Не повторяется'
}
for key, value in form_defaults.items():
    if key not in st.session_state: st.session_state[key] = value

# =============================================================================
# ЗАГОЛОВОК СТРАНИЦЫ
# =============================================================================
st.markdown("""
<div class="header-block">
    <h1>📅 Календарь событий</h1>
    <p>Планирование и управление событиями сети ресторанов</p>
</div>
""", unsafe_allow_html=True)

if st.session_state.success_message:
    st.success(st.session_state.success_message, icon="✅")
    st.session_state.success_message = None

def reset_form():
    for key, value in form_defaults.items():
        st.session_state[key] = value
    st.session_state.edit_mode = False
    st.session_state.selected_event = None
    st.session_state.show_edit_form = False

# =============================================================================
# ВКЛАДКА 1: КАЛЕНДАРЬ (основная область)
# =============================================================================
st.markdown('<div class="calendar-block">', unsafe_allow_html=True)
st.markdown("### 📅 Календарь")

# Получаем все события для отображения в календаре
all_events_for_calendar = get_cached_events()
calendar_events = []
for event in (all_events_for_calendar or []):
    color = cfg.EVENT_CATEGORIES.get(event['category'], '#808080')
    calendar_events.append({
        "title": event['title'],
        "start": str(event['start_date']),
        "end": str(event['end_date']),
        "id": str(event['id']),
        "color": color,
        "extendedProps": {
            "category": event['category'],
            "location": event.get('location_type', 'all')
        }
    })

calendar_options = {
    "editable": False,
    "selectable": False,
    "events": calendar_events,
    "locale": "ru",
    "firstDay": 1,
    "headerToolbar": {
        "left": "prev,next today",
        "center": "title",
        "right": "dayGridMonth"
    },
    "initialView": "dayGridMonth",
    "dayMaxEventRows": 3,
    "height": 650
}
calendar(events=calendar_events, options=calendar_options, key="main_calendar")
st.caption("ℹ️ Для управления событиями и фильтрации перейдите к вкладкам ниже")
st.markdown('</div>', unsafe_allow_html=True)

# =============================================================================
# БЛОК ФИЛЬТРОВ (перенесён из сайдбара в основную область, компактный)
# =============================================================================
st.markdown('<div class="filters-block">', unsafe_allow_html=True)
st.markdown("### 🔍 Фильтры и поиск")

f_col1, f_col2, f_col3, f_col4 = st.columns([2, 2, 2, 2])
with f_col1:
    selected_category = st.multiselect(
        "🏷️ Категория",
        options=list(cfg.EVENT_CATEGORIES.keys()),
        default=[],
        key="sidebar_filter_category"
    )
with f_col2:
    selected_location = st.selectbox(
        "📍 Локация",
        options=list(cfg.LOCATIONS.keys()),
        key="sidebar_filter_location"
    )
with f_col3:
    search_query = st.text_input("🔎 Поиск по названию", key="sidebar_filter_search")
with f_col4:
    date_range = st.date_input(
        "📅 Диапазон дат",
        value=(date.today(), date.today() + timedelta(days=365)),
        key="sidebar_filter_date_range"
    )
    if isinstance(date_range, tuple) and len(date_range) == 2:
        start_date_filter, end_date_filter = date_range
    else:
        start_date_filter = end_date_filter = date_range if isinstance(date_range, date) else date.today()

st.markdown('</div>', unsafe_allow_html=True)

# =============================================================================
# БЛОК ЭКСПОРТА И TELEGRAM (компактно, в одну строку)
# =============================================================================
exp_col1, exp_col2, exp_col3 = st.columns([2, 2, 2])

with exp_col1:
    st.markdown('<div class="export-block">', unsafe_allow_html=True)
    st.markdown("#### 📊 Экспорт в Excel")
    if st.button("📥 Скачать Excel", use_container_width=True, key="btn_export"):
        with st.spinner("Генерация..."):
            events = get_cached_events()
            if events:
                wb = Workbook()
                ws = wb.active
                ws.title = "События"
                header_fill = PatternFill(start_color="005F6B", end_color="005F6B", fill_type="solid")
                spb_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                tyumen_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                center_align = Alignment(horizontal="center", vertical="center")
                double_border = Border(
                    top=Side(style="double"), bottom=Side(style="double"),
                    left=Side(style="thin"), right=Side(style="thin")
                )
                headers = ["Название", "Категория", "Локация", "Дата начала", "Дата окончания", "Статус (IF)", "Вес (SUMPRODUCT)"]
                ws.append(headers)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                sorted_events = sorted(events, key=lambda x: x.get('location_type', 'all'))
                for idx, event in enumerate(sorted_events, start=2):
                    loc_type = event.get('location_type', 'all')
                    loc_display = event.get('location_custom') or loc_type
                    fill = spb_fill if loc_type in ['spb', 'all'] else tyumen_fill
                    status_formula = f'=IF(D{idx}<>"", "Активно", "Архив")'
                    sumproduct_formula = f'=SUMPRODUCT(--(C$2:C${idx}="{loc_display}"))'
                    ws.append([event['title'], event['category'], loc_display,
                               event['start_date'], event['end_date'],
                               status_formula, sumproduct_formula])
                    for cell in ws[idx]:
                        cell.fill = fill
                        cell.alignment = center_align
                last_row = len(sorted_events) + 1
                ws.append(["ИТОГО СОБЫТИЙ", "", "", "", "", "", f'=SUM(G2:G{last_row-1})'])
                for cell in ws[last_row]:
                    cell.font = Font(bold=True, color="000000")
                    cell.border = double_border
                    cell.alignment = center_align
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                st.download_button("💾 Скачать", data=buffer,
                                   file_name=f"calendar_events_{date.today()}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True, key="btn_download_excel")
            else:
                st.warning("Нет событий для экспорта")
    st.markdown('</div>', unsafe_allow_html=True)

with exp_col2:
    st.markdown('<div class="export-block">', unsafe_allow_html=True)
    st.markdown("#### 📄 Экспорт в PDF")
    pdf_months = st.number_input("Месяцев вперёд", min_value=1, max_value=12, value=3, key="pdf_months_input")
    if st.button("📥 Скачать PDF", use_container_width=True, key="btn_export_pdf"):
        with st.spinner("Генерация..."):
            try:
                from utils.pdf_export import create_pdf_calendar
                all_events_for_pdf = get_cached_events()
                filtered_for_pdf = all_events_for_pdf
                if selected_category:
                    filtered_for_pdf = [e for e in filtered_for_pdf if e['category'] in selected_category]
                if selected_location != "Все рестораны":
                    location_code = cfg.LOCATIONS.get(selected_location, "all")
                    if location_code == "spb":
                        filtered_for_pdf = [e for e in filtered_for_pdf if e.get('location_type') in ['spb', 'all']]
                    elif location_code == "tyumen":
                        filtered_for_pdf = [e for e in filtered_for_pdf if e.get('location_type') in ['tyumen', 'all']]
                pdf_bytes = create_pdf_calendar(filtered_for_pdf, "Календарь событий", months_ahead=pdf_months)
                st.download_button("💾 PDF", data=pdf_bytes,
                                   file_name=f"calendar_{date.today()}.pdf",
                                   mime="application/pdf", use_container_width=True, key="btn_download_pdf")
            except ImportError:
                st.error("❌ Модуль PDF не установлен")
            except Exception as e:
                st.error(f"❌ Ошибка: {str(e)}")
    st.markdown('</div>', unsafe_allow_html=True)

with exp_col3:
    st.markdown('<div class="export-block">', unsafe_allow_html=True)
    st.markdown("#### 🔔 Telegram")
    tc1, tc2 = st.columns(2)
    with tc1:
        if st.button("📱 Тест", use_container_width=True, key="btn_test_telegram"):
            try:
                from utils.telegram_bot import test_telegram_connection
                if test_telegram_connection(): st.success("✅ OK")
                else: st.error("❌ Ошибка")
            except ImportError: st.error("❌ Не установлен")
            except Exception as e: st.error(f"❌ {str(e)}")
    with tc2:
        if st.button("📨 Отпр.", use_container_width=True, key="btn_send_reminders"):
            try:
                from utils.telegram_bot import send_reminder_notification
                today = date.today()
                reminders = db.get_upcoming_reminders(today)
                if reminders:
                    count = sum(1 for e in reminders if send_reminder_notification(e))
                    st.success(f"✅ {count} напоминаний")
                else:
                    st.info("ℹ️ Нет событий")
            except ImportError: st.error("❌ Не установлен")
            except Exception as e: st.error(f"❌ {str(e)}")
    st.markdown('</div>', unsafe_allow_html=True)

# =============================================================================
# ПРИМЕНЕНИЕ ФИЛЬТРОВ К СОБЫТИЯМ
# =============================================================================
all_events = get_cached_events()
filtered_events = all_events if all_events else []

if selected_category:
    filtered_events = [e for e in filtered_events if e['category'] in selected_category]
if selected_location != "Все рестораны":
    location_code = cfg.LOCATIONS.get(selected_location, "all")
    if location_code == "spb":
        filtered_events = [e for e in filtered_events if e.get('location_type') in ['spb', 'all']]
    elif location_code == "tyumen":
        filtered_events = [e for e in filtered_events if e.get('location_type') in ['tyumen', 'all']]
if search_query:
    filtered_events = [e for e in filtered_events if search_query.lower() in str(e['title']).lower()]
filtered_events = [
    e for e in filtered_events
    if parse_date(e['start_date']) <= end_date_filter
    and parse_date(e['end_date']) >= start_date_filter
]

# =============================================================================
# ВКЛАДКИ: СПИСОК / ЗАВЕРШЁННЫЕ / ДОБАВИТЬ
# =============================================================================
tab_list, tab_completed, tab_add = st.tabs([
    "📋 Список событий",
    "✅ Завершённые",
    "➕ Добавить событие"
])

# =============================================================================
# ВКЛАДКА: СПИСОК СОБЫТИЙ
# =============================================================================
with tab_list:
    # Форма редактирования (если активна)
    if st.session_state.show_edit_form and st.session_state.selected_event:
        st.markdown("---")
        st.subheader(f"✏️ Редактирование: {st.session_state.selected_event['title']}")
        title = st.text_input("Название", value=st.session_state.data_title, key="edit_input_title")
        ec1, ec2 = st.columns(2)
        with ec1:
            start_date = st.date_input("Дата начала", value=st.session_state.data_start_date, key="edit_input_start")
        with ec2:
            is_indefinite = st.checkbox("♾️ Однодневное", value=(st.session_state.data_start_date == st.session_state.data_end_date), key="edit_input_indefinite")
            end_date = start_date if is_indefinite else st.date_input("Дата окончания", value=st.session_state.data_end_date, key="edit_input_end")
        category = st.selectbox("Категория", options=list(cfg.EVENT_CATEGORIES.keys()),
                                index=list(cfg.EVENT_CATEGORIES.keys()).index(st.session_state.data_category) if st.session_state.data_category in cfg.EVENT_CATEGORIES else 0,
                                key="edit_input_category")
        location_type = st.selectbox("Локация", options=list(cfg.LOCATIONS.keys()),
                                     index=list(cfg.LOCATIONS.keys()).index(st.session_state.data_location) if st.session_state.data_location in cfg.LOCATIONS else 0,
                                     key="edit_input_location")
        location_custom = None
        if location_type == "Выбрать конкретные":
            location_custom_list = st.multiselect("📍 Рестораны", options=getattr(cfg, 'ALL_RESTAURANTS', []),
                                                  default=st.session_state.data_location_custom, key="edit_input_location_custom")
            location_custom = ", ".join(location_custom_list) if location_custom_list else None

        st.markdown("##### 🔔 Напоминания")
        reminder_on_start_day = st.checkbox("В день начала", value=st.session_state.data_reminder_on_start, key="edit_input_reminder_start")
        rc1, rc2, rc3 = st.columns(3)
        with rc1: reminder_days_before_start = st.number_input("За сколько дней до НАЧАЛА", min_value=0, max_value=90, value=st.session_state.data_reminder_days_start, key="edit_input_reminder_days_start")
        with rc2: reminder_days_before_end = st.number_input("За сколько дней до ОКОНЧАНИЯ", min_value=0, max_value=90, value=st.session_state.data_reminder_days_end, key="edit_input_reminder_days_end")
        with rc3:
            use_custom_date = st.checkbox("В конкретную дату", value=st.session_state.data_use_custom_date, key="edit_input_custom_date")
            reminder_custom_date = st.date_input("Дата", value=st.session_state.data_reminder_custom_date, key="edit_input_reminder_custom") if use_custom_date else None
        recurrence_options = list(cfg.RECURRENCE_TYPES.keys())
        recurrence_index = recurrence_options.index(st.session_state.data_recurrence) if st.session_state.data_recurrence in recurrence_options else 0
        recurrence_type = st.selectbox("Повторяемость", options=recurrence_options, index=recurrence_index, key="edit_input_recurrence")

        sc1, sc2 = st.columns(2)
        with sc1:
            if st.button("💾 Сохранить", use_container_width=True, type="primary", key="btn_save_edit"):
                if not title: st.error("❌ Введите название")
                elif start_date > end_date: st.error("❌ Дата начала позже окончания")
                else:
                    location_code = cfg.LOCATIONS.get(location_type, "all")
                    if not use_custom_date: reminder_custom_date = None
                    try:
                        db.update_event(event_id=st.session_state.selected_event['id'], title=title,
                                        start_date=start_date, end_date=end_date, category=category,
                                        location_type=location_code, location_custom=location_custom,
                                        reminder_days_before_start=reminder_days_before_start,
                                        reminder_on_start_day=1 if reminder_on_start_day else 0,
                                        reminder_days_before_end=reminder_days_before_end,
                                        reminder_custom_date=reminder_custom_date,
                                        recurrence_type=cfg.RECURRENCE_TYPES.get(recurrence_type, 'none'))
                        clear_cache()
                        st.session_state.success_message = f"✅ '{title}' обновлено!"
                        reset_form(); st.rerun()
                    except Exception as e: st.error(f"❌ {str(e)}")
        with sc2:
            if st.button("↩️ Отмена", use_container_width=True, key="btn_cancel_edit_inline"):
                reset_form(); st.rerun()
        st.markdown("---")

    # Подтверждение удаления
    if st.session_state.delete_confirm_id:
        event_to_delete = next((e for e in filtered_events if e['id'] == st.session_state.delete_confirm_id), None)
        if event_to_delete:
            st.warning(f"⚠️ Удалить **'{event_to_delete['title']}'**?")
            dc1, dc2 = st.columns(2)
            with dc1:
                if st.button("🗑️ Да, удалить", use_container_width=True, type="primary", key="confirm_delete_yes"):
                    db.delete_event(st.session_state.delete_confirm_id)
                    clear_cache()
                    st.session_state.success_message = f"✅ '{event_to_delete['title']}' удалено!"
                    st.session_state.delete_confirm_id = None; st.rerun()
            with dc2:
                if st.button("↩️ Отмена", use_container_width=True, key="confirm_delete_no"):
                    st.session_state.delete_confirm_id = None; st.rerun()
            st.markdown("---")

    # Список будущих событий
    today = date.today()
    future_events = [e for e in filtered_events if parse_date(e['end_date']) >= today]
    future_events.sort(key=lambda x: parse_date(x['start_date']))

    if future_events:
        select_all = st.checkbox("Выбрать все", key="select_all_events")
        if select_all: st.session_state.events_to_delete = [e['id'] for e in future_events]
        st.write(f"**Всего событий:** {len(future_events)}")

        for event in future_events:
            col_checkbox, col_info, col_actions = st.columns([0.5, 7, 2])
            with col_checkbox:
                is_selected = st.checkbox("Выбрать", value=event['id'] in st.session_state.events_to_delete,
                                          key=f"chk_{event['id']}", label_visibility="visible")
                if is_selected and event['id'] not in st.session_state.events_to_delete:
                    st.session_state.events_to_delete.append(event['id'])
                elif not is_selected and event['id'] in st.session_state.events_to_delete:
                    st.session_state.events_to_delete.remove(event['id'])
            with col_info:
                location_display = event.get('location_custom') or event.get('location_type', 'Не указано')
                st.markdown(f"**{event['title']}**")
                st.caption(f"📅 {event['start_date']} | {event['category']} | {location_display}")
            with col_actions:
                ac1, ac2 = st.columns(2)
                with ac1:
                    if st.button("✏️", key=f"edit_{event['id']}", help="Редактировать"):
                        st.session_state.selected_event = event
                        st.session_state.edit_mode = True
                        st.session_state.show_edit_form = True
                        st.session_state.data_title = event['title']
                        st.session_state.data_start_date = parse_date(event['start_date'])
                        st.session_state.data_end_date = parse_date(event['end_date'])
                        st.session_state.data_category = event['category']
                        st.session_state.data_location = get_key_by_value(cfg.LOCATIONS, event.get('location_type', 'all'), "Все рестораны")
                        st.session_state.data_location_custom = [r.strip() for r in event['location_custom'].split(',')] if event.get('location_custom') else []
                        st.session_state.data_reminder_on_start = bool(event.get('reminder_on_start_day', True))
                        st.session_state.data_reminder_days_start = event.get('reminder_days_before_start', 0)
                        st.session_state.data_reminder_days_end = event.get('reminder_days_before_end', 0)
                        if event.get('reminder_custom_date'):
                            st.session_state.data_use_custom_date = True
                            st.session_state.data_reminder_custom_date = parse_date(event['reminder_custom_date'])
                        else:
                            st.session_state.data_use_custom_date = False
                        st.session_state.data_recurrence = get_recurrence_key(event.get('recurrence_type', 'none'))
                        st.rerun()
                with ac2:
                    if st.button("🗑️", key=f"delete_{event['id']}", help="Удалить"):
                        st.session_state.delete_confirm_id = event['id']; st.rerun()

        if st.session_state.events_to_delete:
            st.warning(f"🗑️ Выбрано для удаления: {len(st.session_state.events_to_delete)}")
            bc1, bc2 = st.columns(2)
            with bc1:
                if st.button("✅ Подтвердить", use_container_width=True, type="primary", key="confirm_bulk_delete"):
                    count = len(st.session_state.events_to_delete)
                    for event_id in st.session_state.events_to_delete: db.delete_event(event_id)
                    clear_cache()
                    st.session_state.success_message = f"✅ Удалено: {count}"
                    st.session_state.events_to_delete = []; st.rerun()
            with bc2:
                if st.button("❌ Отмена", use_container_width=True, key="cancel_bulk_delete"):
                    st.session_state.events_to_delete = []; st.rerun()
    else:
        st.info("Нет будущих событий для отображения")

# =============================================================================
# ВКЛАДКА: ЗАВЕРШЁННЫЕ СОБЫТИЯ
# =============================================================================
with tab_completed:
    st.subheader("✅ Завершённые события")
    st.caption("События, дата окончания которых уже прошла")
    completed_events = get_cached_completed_events()

    if selected_category:
        completed_events = [e for e in completed_events if e['category'] in selected_category]
    if selected_location != "Все рестораны":
        location_code = cfg.LOCATIONS.get(selected_location, "all")
        if location_code == "spb":
            completed_events = [e for e in completed_events if e.get('location_type') in ['spb', 'all']]
        elif location_code == "tyumen":
            completed_events = [e for e in completed_events if e.get('location_type') in ['tyumen', 'all']]
    if search_query:
        completed_events = [e for e in completed_events if search_query.lower() in str(e['title']).lower()]

    if completed_events:
        st.write(f"**Всего завершённых:** {len(completed_events)}")
        for event in completed_events:
            col_info, col_actions = st.columns([8, 2])
            with col_info:
                location_display = event.get('location_custom') or event.get('location_type', 'Не указано')
                end_date = parse_date(event['end_date'])
                days_ago = (date.today() - end_date).days
                st.markdown(f"""
                <div class="event-card-completed">
                    <div style="color: #424242; font-size: 16px;"><b>{event['title']}</b></div>
                    <div style="color: #757575; font-size: 14px; margin-top: 5px;">
                        📅 {event['start_date']} → {event['end_date']} | 🏷 {event['category']} | 📍 {location_display}
                    </div>
                    <div style="color: #9E9E9E; font-size: 12px; margin-top: 5px;">Завершено {days_ago} дн. назад</div>
                </div>
                """, unsafe_allow_html=True)
            with col_actions:
                ac1, ac2 = st.columns(2)
                with ac1:
                    if st.button("✏️", key=f"edit_completed_{event['id']}", help="Редактировать"):
                        st.session_state.selected_event = event
                        st.session_state.edit_mode = True
                        st.session_state.show_edit_form = True
                        st.session_state.data_title = event['title']
                        st.session_state.data_start_date = parse_date(event['start_date'])
                        st.session_state.data_end_date = parse_date(event['end_date'])
                        st.session_state.data_category = event['category']
                        st.session_state.data_location = get_key_by_value(cfg.LOCATIONS, event.get('location_type', 'all'), "Все рестораны")
                        st.session_state.data_location_custom = [r.strip() for r in event['location_custom'].split(',')] if event.get('location_custom') else []
                        st.session_state.data_reminder_on_start = bool(event.get('reminder_on_start_day', True))
                        st.session_state.data_reminder_days_start = event.get('reminder_days_before_start', 0)
                        st.session_state.data_reminder_days_end = event.get('reminder_days_before_end', 0)
                        if event.get('reminder_custom_date'):
                            st.session_state.data_use_custom_date = True
                            st.session_state.data_reminder_custom_date = parse_date(event['reminder_custom_date'])
                        else:
                            st.session_state.data_use_custom_date = False
                        st.session_state.data_recurrence = get_recurrence_key(event.get('recurrence_type', 'none'))
                        st.rerun()
                with ac2:
                    if st.button("🗑️", key=f"delete_completed_{event['id']}", help="Удалить"):
                        st.session_state.delete_confirm_id = event['id']; st.rerun()
    else:
        st.info("ℹ️ Пока нет завершённых событий")

# =============================================================================
# ВКЛАДКА: ДОБАВИТЬ СОБЫТИЕ
# =============================================================================
with tab_add:
    st.subheader("➕ Добавить новое событие")
    title = st.text_input("Название события", value=st.session_state.data_title, key="input_title")
    ac1, ac2 = st.columns(2)
    with ac1:
        start_date = st.date_input("Дата начала", value=st.session_state.data_start_date, key="input_start_date")
    with ac2:
        is_indefinite = st.checkbox("♾️ Однодневное", value=(st.session_state.data_start_date == st.session_state.data_end_date), key="input_indefinite")
        if is_indefinite:
            end_date = start_date
            st.caption("✅ Дата окончания = дате начала")
        else:
            end_date = st.date_input("Дата окончания", value=st.session_state.data_end_date, key="input_end_date")
    category = st.selectbox("Категория", options=list(cfg.EVENT_CATEGORIES.keys()),
                            index=list(cfg.EVENT_CATEGORIES.keys()).index(st.session_state.data_category) if st.session_state.data_category in cfg.EVENT_CATEGORIES else 0,
                            key="input_category")
    location_type = st.selectbox("Локация", options=list(cfg.LOCATIONS.keys()),
                                 index=list(cfg.LOCATIONS.keys()).index(st.session_state.data_location) if st.session_state.data_location in cfg.LOCATIONS else 0,
                                 key="input_location")
    location_custom = None
    if location_type == "Выбрать конкретные":
        location_custom_list = st.multiselect("📍 Рестораны", options=getattr(cfg, 'ALL_RESTAURANTS', []),
                                              default=st.session_state.data_location_custom, key="input_location_custom")
        location_custom = ", ".join(location_custom_list) if location_custom_list else None

    st.markdown("##### 🔔 Напоминания")
    reminder_on_start_day = st.checkbox("В день начала", value=st.session_state.data_reminder_on_start, key="input_reminder_start")
    rc1, rc2, rc3 = st.columns(3)
    with rc1: reminder_days_before_start = st.number_input("До НАЧАЛА (дней)", min_value=0, max_value=90, value=st.session_state.data_reminder_days_start, key="input_reminder_days_start")
    with rc2: reminder_days_before_end = st.number_input("До ОКОНЧАНИЯ (дней)", min_value=0, max_value=90, value=st.session_state.data_reminder_days_end, key="input_reminder_days_end")
    with rc3:
        use_custom_date = st.checkbox("В конкретную дату", value=st.session_state.data_use_custom_date, key="input_custom_date")
        reminder_custom_date = st.date_input("Дата напоминания", value=st.session_state.data_reminder_custom_date, key="input_reminder_custom") if use_custom_date else None
    recurrence_options = list(cfg.RECURRENCE_TYPES.keys())
    recurrence_index = recurrence_options.index(st.session_state.data_recurrence) if st.session_state.data_recurrence in recurrence_options else 0
    recurrence_type = st.selectbox("Повторяемость", options=recurrence_options, index=recurrence_index, key="input_recurrence")

    sc1, sc2 = st.columns(2)
    with sc1:
        if st.button("💾 Сохранить", use_container_width=True, type="primary", key="btn_save_event"):
            if not title: st.error("❌ Введите название")
            elif start_date > end_date: st.error(f"❌ Дата начала ({start_date}) позже окончания ({end_date})")
            else:
                location_code = cfg.LOCATIONS.get(location_type, "all")
                if not use_custom_date: reminder_custom_date = None
                try:
                    db.add_event(title=title, start_date=start_date, end_date=end_date, category=category,
                                 location_type=location_code, location_custom=location_custom,
                                 reminder_days_before_start=reminder_days_before_start,
                                 reminder_on_start_day=1 if reminder_on_start_day else 0,
                                 reminder_days_before_end=reminder_days_before_end,
                                 reminder_custom_date=reminder_custom_date,
                                 recurrence_type=cfg.RECURRENCE_TYPES.get(recurrence_type, 'none'))
                    clear_cache()
                    st.session_state.success_message = f"✅ '{title}' добавлено!"
                    reset_form(); st.rerun()
                except Exception as e: st.error(f"❌ {str(e)}")
    with sc2:
        if st.button("🔄 Очистить", use_container_width=True, key="btn_clear_form"):
            reset_form(); st.rerun()
