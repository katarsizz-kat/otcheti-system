# pages/universal_report.py
import streamlit as st
import pandas as pd
import numpy as np
import re
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

st.set_page_config(page_title="Универсальный отчёт", layout="wide", page_icon="📊")

# ==========================================================
# ДИНАМИЧЕСКИЕ СТИЛИ
# ==========================================================

def get_sky_gradient():
    hour = datetime.now().hour
    if 5 <= hour < 12:
        return {
            "header": "linear-gradient(135deg, #FF8C42 0%, #5DADE2 100%)",
            "block": "linear-gradient(135deg, #FFF5EB 0%, #EBF5FB 100%)",
            "text": "#FFFFFF",
            "button_text": "#FFFFFF",
            "button_bg": "#FF6B35"
        }
    elif 12 <= hour < 18:
        return {
            "header": "linear-gradient(135deg, #3498DB 0%, #EBF5FB 100%)",
            "block": "linear-gradient(135deg, #EBF5FB 0%, #FFFFFF 100%)",
            "text": "#FFFFFF",
            "button_text": "#FFFFFF",
            "button_bg": "#2E86C1"
        }
    elif 18 <= hour < 23:
        return {
            "header": "linear-gradient(135deg, #2874A6 0%, #F39C12 100%)",
            "block": "linear-gradient(135deg, #FDEBD0 0%, #D4E6F1 100%)",
            "text": "#FFFFFF",
            "button_text": "#FFFFFF",
            "button_bg": "#D68910"
        }
    else:
        return {
            "header": "linear-gradient(135deg, #1B4F72 0%, #5DADE2 100%)",
            "block": "linear-gradient(135deg, #D6EAF8 0%, #FFFFFF 100%)",
            "text": "#FFFFFF",
            "button_text": "#1B4F72",
            "button_bg": "#2E86C1"
        }

def apply_dynamic_styles():
    sky = get_sky_gradient()
    st.markdown(f"""
    <style>
        .stApp {{
            background: {sky['block']} !important;
        }}
        .header-block {{
            background: {sky['header']};
            padding: 24px;
            border-radius: 16px;
            margin-bottom: 24px;
            box-shadow: 0 6px 18px rgba(0,0,0,0.15);
            border: 2px solid rgba(255,255,255,0.3);
        }}
        .header-block h1 {{
            margin: 0;
            color: {sky['text']};
            font-size: 36px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
        }}
        .header-block p {{
            margin-top: 8px;
            margin-bottom: 0;
            font-size: 18px;
            color: {sky['text']};
            text-shadow: 1px 1px 2px rgba(0,0,0,0.2);
        }}
        .content-block {{
            background: rgba(255, 255, 255, 0.95);
            padding: 24px;
            border-radius: 12px;
            margin-bottom: 24px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
            border: 1px solid #D6EAF8;
        }}
        .stButton>button {{
            background: {sky['button_bg']} !important;
            color: {sky['button_text']} !important;
            border: none !important;
            border-radius: 8px !important;
            font-weight: bold !important;
            box-shadow: 0 4px 8px rgba(0,0,0,0.2) !important;
            transition: all 0.3s ease !important;
        }}
        .stButton>button:hover {{
            filter: brightness(1.1) !important;
            transform: translateY(-2px) !important;
        }}
        textarea {{
            font-family: 'Courier New', monospace !important;
            font-size: 14px !important;
        }}
    </style>
    """, unsafe_allow_html=True)

def greeting_by_time():
    hour = datetime.now().hour
    if 5 <= hour < 12: return "☀️ Доброе утро!"
    if 12 <= hour < 18: return "🌤 Добрый день!"
    if 18 <= hour < 23: return " Добрый вечер!"
    return "🌙 Доброй ночи!"

# ==========================================================
# ФУНКЦИИ ПАРСИНГА
# ==========================================================

def normalize_text(text):
    """Нормализует текст для сравнения."""
    if not isinstance(text, str): 
        return ""
    replacements = {
        'a': 'а', 'e': 'е', 'o': 'о', 'p': 'р', 'c': 'с', 'y': 'у', 'x': 'х',
        'A': 'А', 'E': 'Е', 'O': 'О', 'P': 'Р', 'C': 'С', 'Y': 'У', 'X': 'Х',
        'ё': 'е', 'Ё': 'Е'
    }
    for lat, cyr in replacements.items():
        text = text.replace(lat, cyr)
    text = text.lower().replace('"', '')
    text = re.sub(r'\(.*?\)', '', text)
    text = re.sub(r'^пицца\s*', '', text)
    text = re.sub(r'\bпромо\b', '', text)
    text = re.sub(r'\bподарок\b', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def map_city(legal_entity):
    """Определяет город по юридическому лицу."""
    le = str(legal_entity)
    if 'СПБ' in le:
        return 'СПБ'
    elif 'ПД' in le and 'СПБ' not in le:
        return 'Тюмень'
    return None

def extract_period_from_file(uploaded_file):
    """Извлекает период из файла."""
    df = pd.read_excel(uploaded_file, header=None, nrows=10)
    
    for idx, row in df.iterrows():
        for val in row.values:
            if pd.notna(val) and 'Период' in str(val):
                match = re.search(r'(\d{2}\.\d{2}\.\d{4})\s+по\s+(\d{2}\.\d{2}\.\d{4})', str(val))
                if match:
                    start_date = match.group(1)
                    end_date = match.group(2)
                    start_dt = datetime.strptime(start_date, '%d.%m.%Y')
                    end_dt = datetime.strptime(end_date, '%d.%m.%Y')
                    return f"{start_dt.strftime('%d.%m')}-{end_dt.strftime('%d.%m.%Y')}"
    
    now = datetime.now()
    return f"01.{now.strftime('%m')}.{now.year}"

def read_and_parse_file(uploaded_file):
    """Читает и парсит основной файл."""
    df = pd.read_excel(uploaded_file, header=None)
    
    header_row = None
    for idx, row in df.iterrows():
        vals = [str(v) for v in row.values if pd.notna(v)]
        if 'Юридическое лицо' in ' '.join(vals) and 'Блюдо' in ' '.join(vals):
            header_row = idx
            break
    
    if header_row is None:
        raise ValueError("Не найдена строка заголовков в файле!")
    
    headers = df.iloc[header_row].values
    df = df[header_row + 1:].copy()
    df.columns = headers
    df = df.reset_index(drop=True)
    
    df = df[~df['Юридическое лицо'].astype(str).str.contains('OLAP|всего|Итого', na=False)]
    df = df.dropna(how='all')
    
    df['Юридическое лицо'] = df['Юридическое лицо'].ffill()
    df['Категория блюда'] = df['Категория блюда'].ffill()
    
    df = df[~df['Блюдо'].astype(str).str.contains('\+', na=False)]
    df = df[~df['Блюдо'].astype(str).str.lower().str.contains('персонал', na=False)]
    df = df[df['Блюдо'].notna()]
    df = df[df['Блюдо'].astype(str).str.strip() != '']
    
    df['Количество блюд'] = pd.to_numeric(df['Количество блюд'], errors='coerce').fillna(0)
    df['Сумма со скидкой, р.'] = pd.to_numeric(df['Сумма со скидкой, р.'], errors='coerce').fillna(0)
    
    return df

def aggregate_by_rules(df, rules_text):
    """Агрегирует данные по введённым правилам."""
    df = df.copy()
    df['Город'] = df['Юридическое лицо'].apply(map_city)
    df = df[df['Город'].notna()]
    
    # Нормализуем названия блюд
    df['Блюдо_norm'] = df['Блюдо'].apply(normalize_text)
    
    # Парсим правила (каждая строка = одна позиция)
    rules = [line.strip() for line in rules_text.split('\n') if line.strip()]
    
    result = {}
    for city in ['СПБ', 'Тюмень']:
        city_data = df[df['Город'] == city]
        city_result = {}
        
        for rule in rules:
            rule_norm = normalize_text(rule)
            
            # Ищем все строки в выгрузке, которые содержат это название
            matching_rows = city_data[city_data['Блюдо_norm'].str.contains(rule_norm, na=False)]
            
            total_qty = int(matching_rows['Количество блюд'].sum())
            total_sum = round(matching_rows['Сумма со скидкой, р.'].sum(), 2)
            
            city_result[rule] = {
                'qty': total_qty,
                'sum': total_sum
            }
        
        result[city] = city_result
    
    return result

# ==========================================================
# СОЗДАНИЕ EXCEL
# ==========================================================

def create_excel_report(data, period_str, rules):
    """Создает Excel файл с отчётом."""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    ws = wb.create_sheet("Отчёт")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Строка 1: Период
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    header_cell = ws.cell(1, 1, period_str)
    header_cell.font = Font(bold=True, size=16)
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # Строка 2: Города
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    spb_header = ws.cell(2, 1, "СПб сейчас")
    spb_header.font = Font(bold=True, size=12)
    spb_header.alignment = Alignment(horizontal='center', vertical='center')
    spb_header.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
    
    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=8)
    tmn_header = ws.cell(2, 5, "Тюмень сейчас")
    tmn_header.font = Font(bold=True, size=12)
    tmn_header.alignment = Alignment(horizontal='center', vertical='center')
    tmn_header.fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
    
    # Строка 3: Заголовки колонок
    headers = ['Наименование', 'Количество', 'Сумма, ₽']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(3, col_idx, header)
        cell.font = Font(bold=True, size=10)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for col_idx, header in enumerate(headers, 5):
        cell = ws.cell(3, col_idx, header)
        cell.font = Font(bold=True, size=10)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Заполняем данными
    current_row = 4
    
    for rule in rules:
        # СПБ: колонки 1-3
        ws.cell(current_row, 1, rule).border = thin_border
        
        spb_data = data['СПБ'][rule]
        ws.cell(current_row, 2, spb_data['qty']).border = thin_border
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(current_row, 3, spb_data['sum']).border = thin_border
        ws.cell(current_row, 3).number_format = '#,##0.00'
        
        # Тюмень: колонки 5-7
        ws.cell(current_row, 5, rule).border = thin_border
        
        tmn_data = data['Тюмень'][rule]
        ws.cell(current_row, 6, tmn_data['qty']).border = thin_border
        ws.cell(current_row, 6).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(current_row, 7, tmn_data['sum']).border = thin_border
        ws.cell(current_row, 7).number_format = '#,##0.00'
        
        current_row += 1
    
    # Итоговая строка
    total_row = current_row + 1
    
    ws.cell(total_row, 1, 'ИТОГО').font = Font(bold=True, size=11)
    ws.cell(total_row, 1).border = thin_border
    
    ws.cell(total_row, 2, f"=SUM(B4:B{total_row-1})").font = Font(bold=True, size=11)
    ws.cell(total_row, 2).border = thin_border
    ws.cell(total_row, 2).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.cell(total_row, 3, f"=SUM(C4:C{total_row-1})").font = Font(bold=True, size=11)
    ws.cell(total_row, 3).border = thin_border
    ws.cell(total_row, 3).number_format = '#,##0.00'
    
    ws.cell(total_row, 5, 'ИТОГО').font = Font(bold=True, size=11)
    ws.cell(total_row, 5).border = thin_border
    
    ws.cell(total_row, 6, f"=SUM(F4:F{total_row-1})").font = Font(bold=True, size=11)
    ws.cell(total_row, 6).border = thin_border
    ws.cell(total_row, 6).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.cell(total_row, 7, f"=SUM(G4:G{total_row-1})").font = Font(bold=True, size=11)
    ws.cell(total_row, 7).border = thin_border
    ws.cell(total_row, 7).number_format = '#,##0.00'
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 3
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    
    return wb

# ==========================================================
# ИНТЕРФЕЙС
# ==========================================================

apply_dynamic_styles()

st.markdown(f"""
<div class="header-block">
    <h1>📊 Универсальный отчёт</h1>
    <p>{greeting_by_time()}</p>
    <p style="margin-top:10px; margin-bottom:0; font-size:16px;">
        Гибкий отчёт с настраиваемым списком позиций
    </p>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="content-block">
    <h3> Как работает отчёт</h3>
    <p><b>Принцип работы:</b></p>
    <ul>
        <li>Введите список позиций в текстовое поле (каждая строка = одна позиция в отчёте)</li>
        <li>Если позиция точно совпадает с названием в выгрузке — показывается как есть</li>
        <li>Если позиция является частью названия — суммируются все подходящие строки</li>
    </ul>
    <p><b>Пример:</b></p>
    <ul>
        <li><code>Картофель из печи 150 гр</code> + <code>Картофель из печи 200 гр</code> → две отдельные строки</li>
        <li><code>Картофель из печи</code> → одна строка с суммой обоих видов</li>
    </ul>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="content-block">', unsafe_allow_html=True)
st.markdown("###  Загрузка файла")
file_main = st.file_uploader(
    "Загрузите файл рейтинг_продукт (Excel)",
    type=['xlsx'],
    key="universal_report",
    help="Загрузите OLAP-отчет с продажами"
)
st.markdown('</div>', unsafe_allow_html=True)

if file_main:
    st.markdown('<div class="content-block">', unsafe_allow_html=True)
    
    # Извлекаем период
    try:
        period_str = extract_period_from_file(file_main)
        file_main.seek(0)
        st.info(f"📅 Обнаруженный период: **{period_str}**")
    except Exception as e:
        period_str = f"01.{datetime.now().strftime('%m')}.{datetime.now().year}"
        st.warning(f"⚠️ Не удалось определить период из файла.")
    
    st.markdown("### ✏️ Введите список позиций")
    st.markdown("<small>Каждая строка = одна позиция в отчёте. Позиции суммируются, если название является частью названия в выгрузке.</small>", unsafe_allow_html=True)
    
    default_rules = """Картофель из печи 150 гр
Картофель из печи 200 гр
Рогалики с колбасками
Рогалики с ветчиной
Сырные палочки
Чикен Джонер
Пепперони
Супер Папа
Мясная"""
    
    rules_text = st.text_area(
        "Список позиций:",
        value=default_rules,
        height=300,
        help="Введите по одной позиции на строку"
    )
    
    if st.button("📊 Сгенерировать отчёт", type="primary", use_container_width=True):
        with st.spinner("⏳ Формирую отчёт..."):
            try:
                # Читаем файл
                df = read_and_parse_file(file_main)
                
                # Парсим правила
                rules = [line.strip() for line in rules_text.split('\n') if line.strip()]
                
                if not rules:
                    st.error(" Введите хотя бы одну позицию!")
                else:
                    # Агрегируем данные
                    data = aggregate_by_rules(df, rules_text)
                    
                    # Создаём Excel
                    wb = create_excel_report(data, period_str, rules)
                    
                    # Сохраняем в BytesIO
                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)
                    
                    st.success("✅ Отчёт сформирован!")
                    
                    # Показываем превью для СПБ
                    st.subheader("📋 Превью данных (СПБ)")
                    
                    preview_data = []
                    for rule in rules:
                        row = {
                            'Позиция': rule,
                            'Количество': data['СПБ'][rule]['qty'],
                            'Сумма, ₽': data['СПБ'][rule]['sum']
                        }
                        preview_data.append(row)
                    
                    df_preview = pd.DataFrame(preview_data)
                    st.dataframe(df_preview, use_container_width=True)
                    
                    # Кнопка скачивания
                    st.download_button(
                        label=" Скачать отчёт Excel",
                        data=output,
                        file_name=f"Универсальный_отчёт_{period_str.replace('.', '-')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
            except Exception as e:
                st.error(f"❌ Ошибка при формировании отчёта: {str(e)}")
                import traceback
                st.code(traceback.format_exc())
    st.markdown('</div>', unsafe_allow_html=True)
else:
    st.info("👆 Загрузите файл для начала работы")
