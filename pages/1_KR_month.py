import streamlit as st
import pandas as pd
import io
import warnings
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

# ВАЖНО: st.set_page_config должен быть ПЕРВЫМ вызовом streamlit в файле!
st.set_page_config(page_title="КР месяц", page_icon="📅", layout="wide")

# ==========================================================
# ДИНАМИЧЕСКИЕ СТИЛИ (НЕБО ПО ВРЕМЕНИ СУТОК)
# ==========================================================
def get_sky_gradient():
    """Возвращает градиент неба в зависимости от времени суток."""
    hour = datetime.now().hour
    # Утро (5:00 - 12:00) - восход: оранжевый → голубой
    if 5 <= hour < 12:
        return {
            "header": "linear-gradient(135deg, #FF8C42 0%, #5DADE2 100%)",
            "block": "linear-gradient(135deg, #FFF5EB 0%, #EBF5FB 100%)",
            "text": "#FFFFFF",
            "button_text": "#FFFFFF",
            "button_bg": "#FF6B35"
        }
    # День (12:00 - 18:00) - ясное небо: голубой → белый
    elif 12 <= hour < 18:
        return {
            "header": "linear-gradient(135deg, #3498DB 0%, #EBF5FB 100%)",
            "block": "linear-gradient(135deg, #EBF5FB 0%, #FFFFFF 100%)",
            "text": "#FFFFFF",
            "button_text": "#FFFFFF",
            "button_bg": "#2E86C1"
        }
    # Вечер (18:00 - 23:00) - закат: сине-голубой → оранжевый
    elif 18 <= hour < 23:
        return {
            "header": "linear-gradient(135deg, #2874A6 0%, #F39C12 100%)",
            "block": "linear-gradient(135deg, #FDEBD0 0%, #D4E6F1 100%)",
            "text": "#FFFFFF",
            "button_text": "#FFFFFF",
            "button_bg": "#D68910"
        }
    # Ночь (23:00 - 5:00) - ночь: тёмно-синий → светло-голубой (СВЕТЛЫЙ БЛОК!)
    else:
        return {
            "header": "linear-gradient(135deg, #1B4F72 0%, #5DADE2 100%)",
            "block": "linear-gradient(135deg, #D6EAF8 0%, #FFFFFF 100%)",  # СВЕТЛЫЙ!
            "text": "#FFFFFF",
            "button_text": "#1B4F72",  # Тёмный текст на светлом фоне
            "button_bg": "#2E86C1"
        }

def apply_dynamic_styles():
    """Применяет динамические стили с учётом времени суток."""
    sky = get_sky_gradient()
    st.markdown(f"""
    <style>
        /* Основной фон приложения */
        .stApp {{
            background: {sky['block']} !important;
        }}
        /* Заголовок */
        .header-block {{
            background: {sky['header']};
            padding: 24px;
            border-radius: 16px;
            margin-bottom: 24px;
            box-shadow: 0 6px 18px rgba(0,0,0,0.15);
            border: 2px solid rgba(255,255,255,0.3);
        }}
        .header-block h1 {{
            margin: 0;
            color: {sky['text']};
            font-size: 36px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
        }}
        .header-block p {{
            margin-top: 8px;
            margin-bottom: 0;
            font-size: 18px;
            color: {sky['text']};
            text-shadow: 1px 1px 2px rgba(0,0,0,0.2);
        }}
        /* Блоки с контентом */
        .content-block {{
            background: rgba(255, 255, 255, 0.95);
            padding: 24px;
            border-radius: 12px;
            margin-bottom: 24px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
            border: 1px solid #D6EAF8;
        }}
        /* Кнопки */
        .stButton>button {{
            background: {sky['button_bg']} !important;
            color: {sky['button_text']} !important;
            border: none !important;
            border-radius: 8px !important;
            font-weight: bold !important;
            box-shadow: 0 4px 8px rgba(0,0,0,0.2) !important;
            transition: all 0.3s ease !important;
        }}
        .stButton>button:hover {{
            filter: brightness(1.1) !important;
            transform: translateY(-2px) !important;
        }}
        /* Поля ввода и селекторы */
        .stSelectbox > div > div, .stNumberInput > div > div > input {{
            border: 1px solid #D6EAF8 !important;
            border-radius: 8px !important;
        }}
        /* Загрузчики файлов */
        .stFileUploader {{
            background: #F8F9F9;
            padding: 15px;
            border-radius: 10px;
            border: 2px dashed #D6EAF8;
        }}
        hr {{
            border: none;
            border-top: 2px solid #D6EAF8;
            margin: 24px 0;
        }}
    </style>
    """, unsafe_allow_html=True)

# ==========================================================
# НАСТРОЙКИ
# ==========================================================
SPB_ORDER = [
    "Транспортный", "Димитрова", "Шмидта", "Пулковская", "Благодатная",
    "Энтузиастов", "Серебристый", "Мурино", "Ветеранов", "Туристская",
    "Наука", "Ленинский",
]
TMN_ORDER = ["Орджоникидзе", "Мельникайте"]
ALL_RESTAURANTS = SPB_ORDER + TMN_ORDER
DISPLAY_MAP_SITE = {"Шмидта": "Соверен"}
RESTAURANT_MAP_FILE1 = {
    "Санкт-Петербург №1": "Транспортный", "Санкт-Петербург №2": "Димитрова",
    "Санкт-Петербург №4": "Шмидта", "Санкт-Петербург №5": "Пулковская",
    "Санкт-Петербург №6": "Благодатная", "Санкт-Петербург №7": "Энтузиастов",
    "Санкт-Петербург №8": "Серебристый", "Санкт-Петербург №13": "Мурино",
    "Санкт-Петербург №15": "Ветеранов", "Санкт-Петербург №16": "Туристская",
    "Санкт-Петербург №18": "Наука", "Санкт-Петербург №20": "Ленинский",
    "Тюмень №2 – Орджоникидзе": "Орджоникидзе", "Тюмень №3 – Мельникайте": "Мельникайте",
}
ADDRESS_MAP = {
    "транспортный": "Транспортный", "димитрова": "Димитрова", "шмидта": "Шмидта",
    "13-я линия": "Шмидта", "васильевск": "Шмидта", "пулковская": "Пулковская",
    "благодатная": "Благодатная", "энтузиастов": "Энтузиастов", "серебристый": "Серебристый",
    "мурино": "Мурино", "петровский бульвар": "Мурино", "ветеранов": "Ветеранов",
    "туристская": "Туристская", "науки": "Наука", "ленинский": "Ленинский",
    "орджоникидзе": "Орджоникидзе", "мельникайте": "Мельникайте",
}
COMPLAINT_KEYWORDS = {
    "Жалоба на продукт": r"сух|холодн|остыл|невкусн|плох|ужас|отврат|прогоркл|испорч|стар|жестк|резин|волос|металл|гряз|воняет|гнил|сырой",
    "Ошибки приготовления": r"пережар|недожар|сыро|сгорел|пригорел|пересолен|недосолен|мало соли|много соли|непропеч",
    "Перепутанные/недоложенные позиции": r"не положили|не доложили|забыли|не было|не дали|перепутали|не тот|вместо|заменя|замен|не хватает|недоложили|недодали|не пришло|не привезли",
    "Жалобы на сервис": r"хам|груб|невежлив|неприятн|игнор|сбросил|отказал|не ответил|не помог|не решил|проблем|жалоб|администратор",
    "Опоздание": r"опозда|опоздал|опоздание|долго|задерж|задержк|не вовремя|не успел|час|полчаса|30 минут|40 минут|50 минут|1 час|1.5 часа|полтора часа",
}
POSITIVE_KEYWORDS = {
    "Вкус": r"вкусн|отличн|превосходн|восхитит|бомб|огонь|пушк|шедевр|идеальн|сочн",
    "Быстрая доставка": r"быстр|оперативн|вовремя|минут|горяч|с пылу с жару|молниеносн",
    "Вежливый персонал": r"вежлив|приятн|доброжелат|внимател|учтив|милый|отзывчив|профессионал",
    "Качество": r"качеств|свеж|хорош|отличн|превосходн|много начинк|много сыра",
    "Атмосфера": r"уютн|атмосфер|чист|комфортн|приятн|красив|интерьер",
}

# ==========================================================
# СТИЛИ EXCEL (ИСПРАВЛЕНО - убраны лишние пробелы)
# ==========================================================
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
GROUP_FONT = Font(name="Calibri", size=12, bold=True)
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True)
DATA_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)

# ✅ ИСПРАВЛЕНО: убраны пробелы внутри кавычек
GREEN_FILL = PatternFill(fill_type="solid", start_color="D9EAD3", end_color="D9EAD3")
YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")

# ✅ ИСПРАВЛЕНО: убраны пробелы внутри кавычек
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")

# ==========================================================
# ФУНКЦИИ ОБРАБОТКИ
# ==========================================================
def parse_price(val):
    if pd.isna(val): return None
    s = str(val).replace(",", "").replace(" ", "").strip()
    try: return float(s)
    except Exception: return None

def map_restaurant_file1(val):
    if pd.isna(val): return None
    return RESTAURANT_MAP_FILE1.get(str(val).strip(), None)

def map_restaurant_address(val):
    if pd.isna(val): return None
    value = str(val).lower()
    for key, restaurant in ADDRESS_MAP.items():
        if key in value: return restaurant
    return None

def greeting_by_time():
    hour = datetime.now().hour
    if 5 <= hour < 12: return "️ Доброе утро!"
    if 12 <= hour < 18: return "🌤 Добрый день!"
    if 18 <= hour < 23: return "🌙 Добрый вечер!"
    return "🌜 Доброй ночи!"

def calc_stats_site(df, threshold):
    results = []
    for restaurant in ALL_RESTAURANTS:
        sub = df[df["Ресторан"] == restaurant]
        five_low = sub[(sub["Рейтинг"] == 5) & (sub["Сумма"] <= threshold)]
        low_count = len(five_low)
        sub = sub[~((sub["Рейтинг"] == 5) & (sub["Сумма"] <= threshold))]
        row = {"Ресторан": restaurant}
        for i in range(1, 6): row[str(i)] = int((sub["Рейтинг"] == i).sum())
        row["всего:"] = sum(row[str(i)] for i in range(1, 6))
        row["Средний рейтинг"] = round(sum(i * row[str(i)] for i in range(1, 6)) / row["всего:"], 2) if row["всего:"] > 0 else 0
        row["Отзывы ≤ порог"] = low_count
        results.append(row)
    return pd.DataFrame(results)

def calc_stats_standard(df):
    results = []
    for restaurant in ALL_RESTAURANTS:
        sub = df[df["Ресторан"] == restaurant]
        row = {"Ресторан": restaurant}
        for i in range(1, 6): row[str(i)] = int((sub["Рейтинг"] == i).sum())
        row["всего:"] = sum(row[str(i)] for i in range(1, 6))
        row["Средний рейтинг"] = round(sum(i * row[str(i)] for i in range(1, 6)) / row["всего:"], 2) if row["всего:"] > 0 else 0
        results.append(row)
    return pd.DataFrame(results)

def calc_summary_fast(df, keywords_dict):
    results = []
    for restaurant in ALL_RESTAURANTS:
        sub = df[df["Ресторан"] == restaurant].copy()
        counts = {cat: 0 for cat in keywords_dict.keys()}
        if not sub.empty:
            for category, pattern in keywords_dict.items():
                counts[category] = int(sub["Текст"].str.contains(pattern, case=False, na=False, regex=True).sum())
        row_result = {"Ресторан": restaurant}
        row_result.update(counts)
        row_result["Всего:"] = sum(counts.values())
        results.append(row_result)
    return pd.DataFrame(results)

def add_summary_totals(df, region, keywords):
    subset = df[df["Ресторан"].isin(SPB_ORDER)] if region == "СПб" else df[df["Ресторан"].isin(TMN_ORDER)]
    total = {"Ресторан": f"Всего {region}:"}
    for category in keywords.keys(): total[category] = int(subset[category].sum())
    total["Всего:"] = sum(total[category] for category in keywords.keys())
    return pd.concat([df, pd.DataFrame([total])], ignore_index=True)

# ==========================================================
# ЗАПИСЬ EXCEL С ФОРМУЛАМИ
# ==========================================================
def write_headers(ws, row, region, start_col, has_low_reviews):
    headers = [region] + [str(i) for i in range(1, 6)] + ["всего:", "Средний рейтинг"]
    if has_low_reviews:
        headers.append("Отзывы менее 749 руб (включительно)")
    for col_idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = GREEN_FILL
        cell.border = THIN_BORDER

def write_restaurant_row(ws, row, restaurant, start_col, df, has_low_reviews=False, low_count=None, use_display_name=True):
    row_data = df[df["Ресторан"] == restaurant]
    if row_data.empty: return row
    row_data = row_data.iloc[0]
    display_name = DISPLAY_MAP_SITE.get(restaurant, restaurant) if use_display_name else restaurant
    cell = ws.cell(row=row, column=start_col, value=display_name)
    cell.font = DATA_FONT
    cell.alignment = LEFT_ALIGN
    cell.border = THIN_BORDER
    for i, col_name in enumerate(["1", "2", "3", "4", "5"]):
        col_idx = start_col + 1 + i
        cell = ws.cell(row=row, column=col_idx, value=int(row_data[col_name]))
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    col_b = get_column_letter(start_col + 1)
    col_f = get_column_letter(start_col + 5)
    cell = ws.cell(row=row, column=start_col + 6, value=f"=SUM({col_b}{row}:{col_f}{row})")
    cell.font = DATA_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    col_c = get_column_letter(start_col + 2)
    col_d = get_column_letter(start_col + 3)
    col_e = get_column_letter(start_col + 4)
    col_g = get_column_letter(start_col + 6)
    formula = f'=IF({col_g}{row}=0,0,({col_b}{row}*1+{col_c}{row}*2+{col_d}{row}*3+{col_e}{row}*4+{col_f}{row}*5)/{col_g}{row})'
    cell = ws.cell(row=row, column=start_col + 7, value=formula)
    cell.font = DATA_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    cell.number_format = '0.00'
    if has_low_reviews:
        cell = ws.cell(row=row, column=start_col + 8, value=int(low_count) if low_count is not None else 0)
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    return row

def write_total_row(ws, row, region, start_col, first_rest_row, last_rest_row, has_low_reviews=False):
    cell = ws.cell(row=row, column=start_col, value=f"Итого {region}:")
    cell.font = TOTAL_FONT
    cell.alignment = LEFT_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = THIN_BORDER
    for i in range(5):
        col_idx = start_col + 1 + i
        col_letter = get_column_letter(col_idx)
        formula = f"=SUM({col_letter}{first_rest_row}:{col_letter}{last_rest_row})"
        cell = ws.cell(row=row, column=col_idx, value=formula)
        cell.font = TOTAL_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = YELLOW_FILL
        cell.border = THIN_BORDER
    col_b = get_column_letter(start_col + 1)
    col_f = get_column_letter(start_col + 5)
    cell = ws.cell(row=row, column=start_col + 6, value=f"=SUM({col_b}{row}:{col_f}{row})")
    cell.font = TOTAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = THIN_BORDER
    col_c = get_column_letter(start_col + 2)
    col_d = get_column_letter(start_col + 3)
    col_e = get_column_letter(start_col + 4)
    col_g = get_column_letter(start_col + 6)
    formula = f'=IF({col_g}{row}=0,0,({col_b}{row}*1+{col_c}{row}*2+{col_d}{row}*3+{col_e}{row}*4+{col_f}{row}*5)/{col_g}{row})'
    cell = ws.cell(row=row, column=start_col + 7, value=formula)
    cell.font = TOTAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = THIN_BORDER
    cell.number_format = '0.00'
    if has_low_reviews:
        col_i = get_column_letter(start_col + 8)
        formula = f"=SUM({col_i}{first_rest_row}:{col_i}{last_rest_row})"
        cell = ws.cell(row=row, column=start_col + 8, value=formula)
        cell.font = TOTAL_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = YELLOW_FILL
        cell.border = THIN_BORDER
    return row

def write_common_restaurant_row(ws, row, restaurant, start_col, site_row, agg_row, geo_row):
    cell = ws.cell(row=row, column=start_col, value=restaurant)
    cell.font = DATA_FONT
    cell.alignment = LEFT_ALIGN
    cell.border = THIN_BORDER
    for i in range(5):
        col_idx = start_col + 1 + i
        site_col = get_column_letter(2 + i)
        agg_col = get_column_letter(2 + i)
        geo_col = get_column_letter(2 + i)
        formula = f"={site_col}{site_row}+{agg_col}{agg_row}+{geo_col}{geo_row}"
        cell = ws.cell(row=row, column=col_idx, value=formula)
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    col_b = get_column_letter(start_col + 1)
    col_f = get_column_letter(start_col + 5)
    cell = ws.cell(row=row, column=start_col + 6, value=f"=SUM({col_b}{row}:{col_f}{row})")
    cell.font = DATA_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    col_c = get_column_letter(start_col + 2)
    col_d = get_column_letter(start_col + 3)
    col_e = get_column_letter(start_col + 4)
    col_g = get_column_letter(start_col + 6)
    formula = f'=IF({col_g}{row}=0,0,({col_b}{row}*1+{col_c}{row}*2+{col_d}{row}*3+{col_e}{row}*4+{col_f}{row}*5)/{col_g}{row})'
    cell = ws.cell(row=row, column=start_col + 7, value=formula)
    cell.font = DATA_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    cell.number_format = '0.00'
    return row

def write_common_total_row(ws, row, region, start_col, site_total_row, agg_total_row, geo_total_row):
    cell = ws.cell(row=row, column=start_col, value=f"Итого {region}:")
    cell.font = TOTAL_FONT
    cell.alignment = LEFT_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = THIN_BORDER
    for i in range(5):
        col_idx = start_col + 1 + i
        site_col = get_column_letter(2 + i)
        agg_col = get_column_letter(2 + i)
        geo_col = get_column_letter(2 + i)
        formula = f"={site_col}{site_total_row}+{agg_col}{agg_total_row}+{geo_col}{geo_total_row}"
        cell = ws.cell(row=row, column=col_idx, value=formula)
        cell.font = TOTAL_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = YELLOW_FILL
        cell.border = THIN_BORDER
    col_b = get_column_letter(start_col + 1)
    col_f = get_column_letter(start_col + 5)
    cell = ws.cell(row=row, column=start_col + 6, value=f"=SUM({col_b}{row}:{col_f}{row})")
    cell.font = TOTAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = THIN_BORDER
    col_c = get_column_letter(start_col + 2)
    col_d = get_column_letter(start_col + 3)
    col_e = get_column_letter(start_col + 4)
    col_g = get_column_letter(start_col + 6)
    formula = f'=IF({col_g}{row}=0,0,({col_b}{row}*1+{col_c}{row}*2+{col_d}{row}*3+{col_e}{row}*4+{col_f}{row}*5)/{col_g}{row})'
    cell = ws.cell(row=row, column=start_col + 7, value=formula)
    cell.font = TOTAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = THIN_BORDER
    cell.number_format = '0.00'
    return row

def write_block(ws, start_row, block_title, df, month_year, has_low_reviews=True, low_counts=None):
    row_map = {}
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9 if has_low_reviews else 8)
    cell = ws.cell(row=start_row, column=1, value=month_year)
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN
    start_row += 2
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9 if has_low_reviews else 8)
    cell = ws.cell(row=start_row, column=1, value=block_title)
    cell.font = GROUP_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = GREEN_FILL
    cell.border = THIN_BORDER
    start_row += 1
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=6)
    cell = ws.cell(row=start_row, column=2, value="Кол-во поставленных звезд")
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    start_row += 1
    write_headers(ws, start_row, "СПБ", 1, has_low_reviews)
    start_row += 1
    first_spb_row = start_row
    for restaurant in SPB_ORDER:
        low_count = low_counts.get(restaurant, 0) if low_counts else 0
        write_restaurant_row(ws, start_row, restaurant, 1, df, has_low_reviews, low_count, use_display_name=True)
        row_map[restaurant] = start_row
        start_row += 1
    last_spb_row = start_row - 1
    write_total_row(ws, start_row, "СПб", 1, first_spb_row, last_spb_row, has_low_reviews)
    row_map["Итого СПб:"] = start_row
    start_row += 2
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=6)
    cell = ws.cell(row=start_row, column=2, value="Кол-во поставленных звезд")
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    start_row += 1
    write_headers(ws, start_row, "Тюмень", 1, has_low_reviews)
    start_row += 1
    first_tmn_row = start_row
    for restaurant in TMN_ORDER:
        low_count = low_counts.get(restaurant, 0) if low_counts else 0
        write_restaurant_row(ws, start_row, restaurant, 1, df, has_low_reviews, low_count, use_display_name=True)
        row_map[restaurant] = start_row
        start_row += 1
    last_tmn_row = start_row - 1
    write_total_row(ws, start_row, "Тюмень", 1, first_tmn_row, last_tmn_row, has_low_reviews)
    row_map["Итого Тюмень:"] = start_row
    start_row += 1
    return row_map, start_row

def write_common_block(ws, start_row, row_map_site, row_map_agg, row_map_geo):
    ws.merge_cells(start_row=start_row, start_column=11, end_row=start_row, end_column=18)
    cell = ws.cell(row=start_row, column=11, value="Общий (сайт+агрегаторы+геосервисы)")
    cell.font = GROUP_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = GREEN_FILL
    cell.border = THIN_BORDER
    start_row += 1
    ws.merge_cells(start_row=start_row, start_column=12, end_row=start_row, end_column=16)
    cell = ws.cell(row=start_row, column=12, value="Кол-во поставленных звезд")
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    start_row += 1
    write_headers(ws, start_row, "СПБ", 11, has_low_reviews=False)
    start_row += 1
    for restaurant in SPB_ORDER:
        site_row = row_map_site[restaurant]
        agg_row = row_map_agg[restaurant]
        geo_row = row_map_geo[restaurant]
        write_common_restaurant_row(ws, start_row, restaurant, 11, site_row, agg_row, geo_row)
        start_row += 1
    write_common_total_row(ws, start_row, "СПб", 11, 
                           row_map_site["Итого СПб:"], 
                           row_map_agg["Итого СПб:"], 
                           row_map_geo["Итого СПб:"])
    start_row += 2
    ws.merge_cells(start_row=start_row, start_column=12, end_row=start_row, end_column=16)
    cell = ws.cell(row=start_row, column=12, value="Кол-во поставленных звезд")
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    start_row += 1
    write_headers(ws, start_row, "Тюмень", 11, has_low_reviews=False)
    start_row += 1
    for restaurant in TMN_ORDER:
        site_row = row_map_site[restaurant]
        agg_row = row_map_agg[restaurant]
        geo_row = row_map_geo[restaurant]
        write_common_restaurant_row(ws, start_row, restaurant, 11, site_row, agg_row, geo_row)
        start_row += 1
    write_common_total_row(ws, start_row, "Тюмень", 11,
                           row_map_site["Итого Тюмень:"],
                           row_map_agg["Итого Тюмень:"],
                           row_map_geo["Итого Тюмень:"])
    return start_row

def write_analysis_sheet(writer, comp_all, pos_all):
    comp_all.to_excel(writer, sheet_name="Анализ отзывов", index=False, startrow=1)
    pos_all.to_excel(writer, sheet_name="Анализ отзывов", index=False, startrow=len(comp_all) + 4)
    ws = writer.sheets["Анализ отзывов"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(comp_all.columns))
    cell = ws.cell(row=1, column=1, value="Анализ отзывов (жалобы)")
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN
    for col_idx in range(1, len(comp_all.columns) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = GREEN_FILL
        cell.border = THIN_BORDER
    for row_idx in range(3, len(comp_all) + 3):
        for col_idx in range(1, len(comp_all.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN if col_idx > 1 else LEFT_ALIGN
            cell.border = THIN_BORDER
    pos_start_row = len(comp_all) + 4
    ws.merge_cells(start_row=pos_start_row, start_column=1, end_row=pos_start_row, end_column=len(pos_all.columns))
    cell = ws.cell(row=pos_start_row, column=1, value="Анализ отзывов (позитив)")
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN
    for col_idx in range(1, len(pos_all.columns) + 1):
        cell = ws.cell(row=pos_start_row + 1, column=col_idx)
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = GREEN_FILL
        cell.border = THIN_BORDER
    for row_idx in range(pos_start_row + 2, pos_start_row + 2 + len(pos_all)):
        for col_idx in range(1, len(pos_all.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN if col_idx > 1 else LEFT_ALIGN
            cell.border = THIN_BORDER
    for col_idx in range(1, max(len(comp_all.columns), len(pos_all.columns)) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

# ==========================================================
# ПРИМЕНЯЕМ СТИЛИ И РИСУЕМ ИНТЕРФЕЙС
# ==========================================================
apply_dynamic_styles()
sky = get_sky_gradient()
st.markdown("""
<style>
/* Контейнеры загрузчиков */
.element-container:has( > .stFileUploader) {
    background: white;
    border-radius: 12px;
    padding: 20px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.08);
    border: 2px solid transparent;
    transition: all 0.3s ease;
}

.element-container:has( > .stFileUploader):hover {
    transform: translateY(-4px);
    box-shadow: 0 8px 20px rgba(0,0,0,0.12);
}

/* Сайт/приложение - голубой */
.element-container:nth-child(2):has( > .stFileUploader) {
    border-color: #3498DB;
    background: linear-gradient(135deg, #EBF5FB 0%, #FFFFFF 100%);
}

/* Агрегаторы - ЖЁЛТЫЙ */
.element-container:nth-child(3):has( > .stFileUploader) {
    border-color: #F1C40F;
    background: linear-gradient(135deg, #FEF9E7 0%, #FFFFFF 100%);
}

/* Геосервисы - зелёный */
.element-container:nth-child(4):has( > .stFileUploader) {
    border-color: #27AE60;
    background: linear-gradient(135deg, #E8F8F5 0%, #FFFFFF 100%);
}

/* Кнопка Upload */
.stFileUploader [data-testid="stFileUploader"] {
    background: white;
    border-radius: 8px;
    padding: 10px;
}

.stFileUploader label {
    font-weight: bold;
    color: #2C3E50;
}

/* Зона drag & drop */
.stFileUploader > div > div {
    border: 2px dashed rgba(0,0,0,0.2) !important;
    border-radius: 8px !important;
    background: rgba(255,255,255,0.8) !important;
    transition: all 0.3s ease !important;
}

.stFileUploader > div > div:hover {
    border-color: #3498DB !important;
    background: rgba(255,255,255,1) !important;
}

/* Текст под кнопкой */
.stFileUploader small {
    color: #7F8C8D;
    font-size: 12px;
}
</style>
""", unsafe_allow_html=True)
st.markdown(f"""
<div class="header-block">
    <h1>📅 КР месяц</h1>
    <p>{greeting_by_time()}</p>
    <p style="margin-top:10px; margin-bottom:0; font-size:16px;">Формирование ежемесячного отчёта по отзывам из трёх источников</p>
</div>
""", unsafe_allow_html=True)
st.markdown('<div class="content-block">', unsafe_allow_html=True)
st.markdown("### 📂 Загрузка файлов")
col1, col2, col3 = st.columns(3)
with col1:
    st.markdown("#### 📱 Сайт / приложение")
    file1 = st.file_uploader("Загрузите Excel", type=["xlsx", "xls"], key="site", label_visibility="collapsed")
with col2:
    st.markdown("#### 🛵 Агрегаторы")
    file2 = st.file_uploader("Загрузите Excel", type=["xlsx", "xls"], key="agg", label_visibility="collapsed")
with col3:
    st.markdown("#### 📍 Геосервисы")
    file3 = st.file_uploader("Загрузите Excel", type=["xlsx", "xls"], key="geo", label_visibility="collapsed")
st.markdown('</div>', unsafe_allow_html=True)
st.markdown('<div class="content-block">', unsafe_allow_html=True)
st.subheader("⚙️ Настройки")
col_a, col_b, col_c = st.columns(3)
with col_a:
    month_names = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                   "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
    current_month_idx = datetime.now().month - 1
    if current_month_idx == 0:
        default_month = "Декабрь"
        default_year = datetime.now().year - 1
    else:
        default_month = month_names[current_month_idx - 1]
        default_year = datetime.now().year
    selected_month = st.selectbox("Месяц отчёта", month_names, index=month_names.index(default_month))
with col_b:
    selected_year = st.number_input("Год", value=default_year, min_value=2020, max_value=2030, step=1)
with col_c:
    price_threshold = st.number_input("Минимальная сумма заказа", value=749, step=10)
st.markdown("")
generate_report = st.button("🚀 Сформировать отчёт", use_container_width=True, type="primary")
st.markdown('</div>', unsafe_allow_html=True)

# ==========================================================
# ОБРАБОТКА ДАННЫХ
# ==========================================================
if generate_report:
    if not (file1 and file2 and file3):
        st.error("⚠️ Пожалуйста, загрузите все три Excel-файла.")
        st.stop()
    month_year = f"{selected_month} ({selected_year}г)"
    with st.spinner("⏳ Обработка данных..."):
        try:
            # --- Сайт ---
            df1 = pd.read_excel(file1)
            df1["Ресторан"] = df1["Ресторан"].apply(map_restaurant_file1)
            df1["Сумма"] = df1["Сумма заказа со скидкой"].apply(parse_price)
            df1["Рейтинг"] = pd.to_numeric(df1["Рейтинг"], errors="coerce")
            df1["Текст"] = df1["Комментарий"].fillna("")
            df1 = df1[["Ресторан", "Рейтинг", "Текст", "Сумма"]].dropna(subset=["Ресторан", "Рейтинг"])
            # --- Агрегаторы ---
            df2 = pd.read_excel(file2)
            df2["Ресторан"] = df2["Адрес"].apply(map_restaurant_address)
            df2["Рейтинг"] = pd.to_numeric(df2["Оценка"], errors="coerce")
            df2["Текст"] = df2["Отзыв"].fillna("")
            df2 = df2[["Ресторан", "Рейтинг", "Текст"]].dropna(subset=["Ресторан", "Рейтинг"])
            # --- Геосервисы ---
            df3 = pd.read_excel(file3)
            if "Статус отзыва" in df3.columns:
                df3 = df3[df3["Статус отзыва"] != "Удален"].copy()
            if "Название филиала" in df3.columns and "Адрес филиала" in df3.columns:
                df3["Ресторан"] = df3.apply(lambda r: map_restaurant_address(str(r["Название филиала"])) or map_restaurant_address(str(r["Адрес филиала"])), axis=1)
            elif "Адрес филиала" in df3.columns:
                df3["Ресторан"] = df3["Адрес филиала"].apply(map_restaurant_address)
            else:
                df3["Ресторан"] = None
            df3["Рейтинг"] = pd.to_numeric(df3["Оценка"], errors="coerce")
            df3["Текст"] = df3["Текст отзыва"].fillna("")
            df3 = df3[["Ресторан", "Рейтинг", "Текст"]].dropna(subset=["Ресторан", "Рейтинг"])
            # --- Таблицы оценок ---
            stats1 = calc_stats_site(df1, price_threshold)
            stats2 = calc_stats_standard(df2)
            stats3 = calc_stats_standard(df3)
            low_counts = {rest: len(df1[(df1["Ресторан"] == rest) & (df1["Рейтинг"] == 5) & (df1["Сумма"] <= price_threshold)]) for rest in ALL_RESTAURANTS}
            # --- Анализ отзывов ---
            df_all_texts = pd.concat([df1[["Ресторан", "Текст"]], df2[["Ресторан", "Текст"]], df3[["Ресторан", "Текст"]]], ignore_index=True)
            comp_all = calc_summary_fast(df_all_texts, COMPLAINT_KEYWORDS)
            pos_all = calc_summary_fast(df_all_texts, POSITIVE_KEYWORDS)
            comp_all = add_summary_totals(add_summary_totals(comp_all, "СПб", COMPLAINT_KEYWORDS), "Тюмень", COMPLAINT_KEYWORDS)
            pos_all = add_summary_totals(add_summary_totals(pos_all, "СПб", POSITIVE_KEYWORDS), "Тюмень", POSITIVE_KEYWORDS)
            # =====================================================
            # СОХРАНЕНИЕ EXCEL С ФОРМУЛАМИ
            # =====================================================
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                workbook = writer.book
                ws = workbook.create_sheet("Оценки")
                start_row = 1
                row_map_site, next_row = write_block(
                    ws, start_row, "Сайт/приложение", stats1, month_year,
                    has_low_reviews=True, low_counts=low_counts
                )
                agg_start_row = next_row + 2
                row_map_agg, next_row = write_block(
                    ws, agg_start_row, "Агрегаторы", stats2, month_year,
                    has_low_reviews=False, low_counts=None
                )
                geo_start_row = next_row + 2
                row_map_geo, next_row = write_block(
                    ws, geo_start_row, "Геосервисы (Я.Карты, 2ГИС, Гугл карты)", stats3, month_year,
                    has_low_reviews=False, low_counts=None
                )
                write_common_block(ws, start_row, row_map_site, row_map_agg, row_map_geo)
                ws.column_dimensions['A'].width = 18
                for col_idx in range(2, 10):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 10
                ws.column_dimensions['J'].width = 3
                ws.column_dimensions['K'].width = 18
                for col_idx in range(12, 19):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 10
                write_analysis_sheet(writer, comp_all, pos_all)
                if "Sheet" in workbook.sheetnames:
                    del workbook["Sheet"]
            output.seek(0)
            st.success(f"✅ Отчёт за {month_year} успешно сформирован!")
            st.download_button(
                "📥 Скачать Excel", output,
                file_name=f"КР_{selected_month}_{selected_year}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.markdown('<div class="content-block">', unsafe_allow_html=True)
            st.subheader(" Общий итог")
            df_all_filtered = pd.concat([
                df1[~((df1["Рейтинг"] == 5) & (df1["Сумма"] <= price_threshold))][["Ресторан", "Рейтинг", "Текст"]],
                df2, df3
            ], ignore_index=True)
            stats_all_rows = []
            for restaurant in ALL_RESTAURANTS:
                sub = df_all_filtered[df_all_filtered["Ресторан"] == restaurant]
                row = {"Ресторан": restaurant}
                for i in range(1, 6): row[str(i)] = int((sub["Рейтинг"] == i).sum())
                row["всего:"] = sum(row[str(i)] for i in range(1, 6))
                row["Средний рейтинг"] = round(sum(i * row[str(i)] for i in range(1, 6)) / row["всего:"], 2) if row["всего:"] > 0 else 0
                row["Отзывы ≤ порог"] = low_counts[restaurant]
                stats_all_rows.append(row)
            stats_all = pd.DataFrame(stats_all_rows)
            st.dataframe(stats_all, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        except Exception as e:
            st.error(f"❌ Произошла ошибка при обработке: {e}")
            st.exception(e)
