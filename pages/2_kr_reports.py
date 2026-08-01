import streamlit as st
import pandas as pd
import io
import warnings
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

st.set_page_config(page_title=" КР отчёты", page_icon="", layout="wide")

from styles import apply_theme
from config.greetings import get_current_greeting
from config.holidays import get_today_holiday

# ==========================================================
# ПРИМЕНЯЕМ ОБЩУЮ ТЕМУ
# ==========================================================
greeting_data = get_current_greeting()
holiday = get_today_holiday()
holiday_effects = holiday.get("effects") if holiday and isinstance(holiday, dict) else None
apply_subtle_theme(greeting_data["theme"], holiday_effects)

# ==========================================================
# УВЕЛИЧЕННЫЙ ПЕРЕКЛЮЧАТЕЛЬ ТИПА ОТЧЁТА (2x больше)
# ==========================================================
st.markdown("""
<style>
/* Увеличенный переключатель типа отчёта */
div[data-testid="stRadio"] > div {
    transform: scale(2);
    transform-origin: left center;
    margin-bottom: 40px !important;
    margin-top: 20px !important;
}

div[data-testid="stRadio"] label {
    font-size: 18px !important;
    font-weight: 600 !important;
}

div[data-testid="stRadio"] > div > div {
    gap: 12px !important;
}

/* Увеличенные кружочки radio-кнопок */
div[data-testid="stRadio"] div[data-testid="stWidget"] {
    width: 24px !important;
    height: 24px !important;
    min-width: 24px !important;
    min-height: 24px !important;
}

/* Заголовок секции */
div[data-testid="stRadio"] > div > div > div:first-child {
    font-size: 20px !important;
    font-weight: 700 !important;
    margin-bottom: 15px !important;
}
</style>
""", unsafe_allow_html=True)

# ==========================================================
# ОСВЕТЛЕНИЕ ФОНА
# ==========================================================
st.markdown("""
<style>
.stApp::before {
    content: '';
    position: fixed;
    top: 0;
    left: 0;
    width: 100%;
    height: 100%;
    background: rgba(255, 255, 255, 0.15);
    z-index: 0;
    pointer-events: none;
}

/* Стили загрузчиков */
.element-container:has( > .stFileUploader) {
    background: transparent !important;
    box-shadow: none !important;
    border: none !important;
    padding: 0 !important;
}

.stFileUploader {
    background: rgba(255, 255, 255, 0.9);
    padding: 20px;
    border-radius: 12px;
    border: 2px dashed rgba(0, 0, 0, 0.15);
    transition: all 0.3s ease;
}

.stFileUploader:hover {
    border-color: #3498DB;
    background: rgba(255, 255, 255, 1);
}

.stFileUploader > div > div {
    border: none !important;
    background: transparent !important;
}

.stFileUploader button {
    background: #3498DB !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 8px 16px !important;
    font-weight: 600 !important;
}

.stFileUploader button:hover {
    background: #2980B9 !important;
}

.stFileUploader small {
    color: #7F8C8D;
    font-size: 12px;
}

.stMarkdown h4 {
    margin-bottom: 12px;
    color: var(--text-primary, #2C3E50);
}
</style>
""", unsafe_allow_html=True)

# ==========================================================
# НАСТРОЙКИ
# ==========================================================
SPB_ORDER = [
    "Транспортный", "Димитрова", "Шмидта", "Пулковская", "Благодатная",
    "Энтузиастов", "Серебристый", "Мурино", "Ветеранов", "Туристская",
    "Наука", "Ленинский",
]
TMN_ORDER = ["Орджоникидзе", "Мельникайте"]
ALL_RESTAURANTS = SPB_ORDER + TMN_ORDER

# Маппинги для разных типов отчётов
RESTAURANT_MAP_MONTH = {
    "Санкт-Петербург №1": "Транспортный", "Санкт-Петербург №2": "Димитрова",
    "Санкт-Петербург №4": "Шмидта", "Санкт-Петербург №5": "Пулковская",
    "Санкт-Петербург №6": "Благодатная", "Санкт-Петербург №7": "Энтузиастов",
    "Санкт-Петербург №8": "Серебристый", "Санкт-Петербург №13": "Мурино",
    "Санкт-Петербург №15": "Ветеранов", "Санкт-Петербург №16": "Туристская",
    "Санкт-Петербург №18": "Наука", "Санкт-Петербург №20": "Ленинский",
    "Тюмень №2 – Орджоникидзе": "Орджоникидзе", "Тюмень №3 – Мельникайте": "Мельникайте",
}

RESTAURANT_MAP_WEEK = {
    "Санкт-Петербург №1": "01 Транспортный", "Санкт-Петербург №2": "02 Димитрова",
    "Санкт-Петербург №4": "04 Шмидта", "Санкт-Петербург №5": "05 Пулковская",
    "Санкт-Петербург №6": "06 Благодатная", "Санкт-Петербург №7": "07 Энтузиастов",
    "Санкт-Петербург №8": "08 Серебристый", "Санкт-Петербург №13": "13 Мурино",
    "Санкт-Петербург №15": "15 Ветеранов", "Санкт-Петербург №16": "16 Туристская",
    "Санкт-Петербург №18": "18 Наука", "Санкт-Петербург №20": "20 Ленинский",
    "Тюмень №2 – Орджоникидзе": "ТМН 2 (Орджоникидзе)", "Тюмень №3 – Мельникайте": "ТМН 3 (Мельникайте)"
}

ADDRESS_MAP = {
    "транспортный": "Транспортный", "димитрова": "Димитрова", "шмидта": "Шмидта",
    "13-я линия": "Шмидта", "васильевск": "Шмидта", "пулковская": "Пулковская",
    "благодатная": "Благодатная", "энтузиастов": "Энтузиастов", "серебристый": "Серебристый",
    "мурино": "Мурино", "петровский бульвар": "Мурино", "ветеранов": "Ветеранов",
    "туристская": "Туристская", "науки": "Наука", "ленинский": "Ленинский",
    "орджоникидзе": "Орджоникидзе", "мельникайте": "Мельникайте",
}

COMPLAINT_KEYWORDS = {
    "Жалоба на продукт": r"сух|холодн|остыл|невкусн|плох|ужас|отврат|прогоркл|испорч|стар|жестк|резин|волос|металл|гряз|воняет|гнил|сырой",
    "Ошибки приготовления": r"пережар|недожар|сыро|сгорел|пригорел|пересолен|недосолен|мало соли|много соли|непропеч",
    "Перепутанные/недоложенные позиции": r"не положили|не доложили|забыли|не было|не дали|перепутали|не тот|вместо|заменя|замен|не хватает|недоложили|недодали|не пришло|не привезли",
    "Жалобы на сервис": r"хам|груб|невежлив|неприятн|игнор|сбросил|отказал|не ответил|не помог|не решил|проблем|жалоб|администратор",
    "Опоздание": r"опозда|опоздал|опоздание|долго|задерж|задержк|не вовремя|не успел|час|полчаса|30 минут|40 минут|50 минут|1 час|1.5 часа|полтора часа",
}

POSITIVE_KEYWORDS = {
    "Вкус": r"вкусн|отличн|превосходн|восхитит|бомб|огонь|пушк|шедевр|идеальн|сочн",
    "Быстрая доставка": r"быстр|оперативн|вовремя|минут|горяч|с пылу с жару|молниеносн",
    "Вежливый персонал": r"вежлив|приятн|доброжелат|внимател|учтив|милый|отзывчив|профессионал",
    "Качество": r"качеств|свеж|хорош|отличн|превосходн|много начинк|много сыра",
    "Атмосфера": r"уютн|атмосфер|чист|комфортн|приятн|красив|интерьер",
}

# Стили Excel
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
GROUP_FONT = Font(name="Calibri", size=12, bold=True)
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=True)
DATA_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)
GREEN_FILL = PatternFill(fill_type="solid", start_color="D9EAD3", end_color="D9EAD3")
YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")

# ==========================================================
# ФУНКЦИИ ОБРАБОТКИ
# ==========================================================
def parse_price(val):
    if pd.isna(val): return None
    s = str(val).replace(",", "").replace(" ", "").strip()
    try: return float(s)
    except Exception: return None

def map_restaurant_file1(val, report_type):
    if pd.isna(val): return None
    mapper = RESTAURANT_MAP_MONTH if report_type == "month" else RESTAURANT_MAP_WEEK
    return mapper.get(str(val).strip(), None)

def map_restaurant_address(val):
    if pd.isna(val): return None
    value = str(val).lower()
    for key, restaurant in ADDRESS_MAP.items():
        if key in value: return restaurant
    return None

def greeting_by_time():
    hour = datetime.now().hour
    if 5 <= hour < 12: return "🌅 Доброе утро!"
    if 12 <= hour < 18: return "🌤 Добрый день!"
    if 18 <= hour < 23: return "🌙 Добрый вечер!"
    return "🌜 Доброй ночи!"

def calc_stats(df, threshold=None, filter_low_reviews=False):
    """Универсальная функция расчёта статистики."""
    results = []
    restaurants = ALL_RESTAURANTS
    
    for restaurant in restaurants:
        sub = df[df["Ресторан"] == restaurant]
        
        # Для месячного отчёта: фильтруем 5★ с низкой суммой
        if filter_low_reviews and threshold is not None:
            five_low = sub[(sub["Рейтинг"] == 5) & (sub["Сумма"] <= threshold)]
            low_count = len(five_low)
            sub = sub[~((sub["Рейтинг"] == 5) & (sub["Сумма"] <= threshold))]
        else:
            low_count = 0
        
        row = {"Ресторан": restaurant}
        for i in range(1, 6): 
            row[str(i)] = int((sub["Рейтинг"] == i).sum())
        row["всего:"] = sum(row[str(i)] for i in range(1, 6))
        row["Средний рейтинг"] = round(sum(i * row[str(i)] for i in range(1, 6)) / row["всего:"], 2) if row["всего:"] > 0 else 0
        
        if filter_low_reviews:
            row["Отзывы ≤ порог"] = low_count
        
        results.append(row)
    
    return pd.DataFrame(results)

def calc_summary_fast(df, keywords_dict):
    results = []
    for restaurant in ALL_RESTAURANTS:
        sub = df[df["Ресторан"] == restaurant].copy()
        counts = {cat: 0 for cat in keywords_dict.keys()}
        if not sub.empty:
            for category, pattern in keywords_dict.items():
                counts[category] = int(sub["Текст"].str.contains(pattern, case=False, na=False, regex=True).sum())
        row_result = {"Ресторан": restaurant}
        row_result.update(counts)
        row_result["Всего:"] = sum(counts.values())
        results.append(row_result)
    return pd.DataFrame(results)

def add_summary_totals(df, region, keywords):
    subset = df[df["Ресторан"].isin(SPB_ORDER)] if region == "СПб" else df[df["Ресторан"].isin(TMN_ORDER)]
    total = {"Ресторан": f"Всего {region}:"}
    for category in keywords.keys(): 
        total[category] = int(subset[category].sum())
    total["Всего:"] = sum(total[category] for category in keywords.keys())
    return pd.concat([df, pd.DataFrame([total])], ignore_index=True)

# ==========================================================
# ЗАПИСЬ EXCEL
# ==========================================================
def write_headers(ws, row, region, start_col, has_low_reviews):
    headers = [region] + [str(i) for i in range(1, 6)] + ["всего:", "Средний рейтинг"]
    if has_low_reviews:
        headers.append("Отзывы менее 749 руб (включительно)")
    for col_idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = GREEN_FILL
        cell.border = THIN_BORDER

def write_restaurant_row(ws, row, restaurant, start_col, df, has_low_reviews=False, low_count=None):
    row_data = df[df["Ресторан"] == restaurant]
    if row_data.empty: return row
    
    row_data = row_data.iloc[0]
    cell = ws.cell(row=row, column=start_col, value=restaurant)
    cell.font = DATA_FONT
    cell.alignment = LEFT_ALIGN
    cell.border = THIN_BORDER
    
    for i, col_name in enumerate(["1", "2", "3", "4", "5"]):
        col_idx = start_col + 1 + i
        cell = ws.cell(row=row, column=col_idx, value=int(row_data[col_name]))
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    col_b = get_column_letter(start_col + 1)
    col_f = get_column_letter(start_col + 5)
    cell = ws.cell(row=row, column=start_col + 6, value=f"=SUM({col_b}{row}:{col_f}{row})")
    cell.font = DATA_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    
    col_c = get_column_letter(start_col + 2)
    col_d = get_column_letter(start_col + 3)
    col_e = get_column_letter(start_col + 4)
    col_g = get_column_letter(start_col + 6)
    formula = f'=IF({col_g}{row}=0,0,({col_b}{row}*1+{col_c}{row}*2+{col_d}{row}*3+{col_e}{row}*4+{col_f}{row}*5)/{col_g}{row})'
    cell = ws.cell(row=row, column=start_col + 7, value=formula)
    cell.font = DATA_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    cell.number_format = '0.00'
    
    if has_low_reviews:
        cell = ws.cell(row=row, column=start_col + 8, value=int(low_count) if low_count is not None else 0)
        cell.font = DATA_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    return row

def write_total_row(ws, row, region, start_col, first_rest_row, last_rest_row, has_low_reviews=False):
    cell = ws.cell(row=row, column=start_col, value=f"Итого {region}:")
    cell.font = TOTAL_FONT
    cell.alignment = LEFT_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = THIN_BORDER
    
    for i in range(5):
        col_idx = start_col + 1 + i
        col_letter = get_column_letter(col_idx)
        formula = f"=SUM({col_letter}{first_rest_row}:{col_letter}{last_rest_row})"
        cell = ws.cell(row=row, column=col_idx, value=formula)
        cell.font = TOTAL_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = YELLOW_FILL
        cell.border = THIN_BORDER
    
    col_b = get_column_letter(start_col + 1)
    col_f = get_column_letter(start_col + 5)
    cell = ws.cell(row=row, column=start_col + 6, value=f"=SUM({col_b}{row}:{col_f}{row})")
    cell.font = TOTAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = THIN_BORDER
    
    col_c = get_column_letter(start_col + 2)
    col_d = get_column_letter(start_col + 3)
    col_e = get_column_letter(start_col + 4)
    col_g = get_column_letter(start_col + 6)
    formula = f'=IF({col_g}{row}=0,0,({col_b}{row}*1+{col_c}{row}*2+{col_d}{row}*3+{col_e}{row}*4+{col_f}{row}*5)/{col_g}{row})'
    cell = ws.cell(row=row, column=start_col + 7, value=formula)
    cell.font = TOTAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = YELLOW_FILL
    cell.border = THIN_BORDER
    cell.number_format = '0.00'
    
    if has_low_reviews:
        col_i = get_column_letter(start_col + 8)
        formula = f"=SUM({col_i}{first_rest_row}:{col_i}{last_rest_row})"
        cell = ws.cell(row=row, column=start_col + 8, value=formula)
        cell.font = TOTAL_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = YELLOW_FILL
        cell.border = THIN_BORDER
    
    return row

def write_block(ws, start_row, block_title, df, period_text, has_low_reviews=True, low_counts=None):
    row_map = {}
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9 if has_low_reviews else 8)
    cell = ws.cell(row=start_row, column=1, value=period_text)
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN
    start_row += 2
    
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9 if has_low_reviews else 8)
    cell = ws.cell(row=start_row, column=1, value=block_title)
    cell.font = GROUP_FONT
    cell.alignment = CENTER_ALIGN
    cell.fill = GREEN_FILL
    cell.border = THIN_BORDER
    start_row += 1
    
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=6)
    cell = ws.cell(row=start_row, column=2, value="Кол-во поставленных звезд")
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    start_row += 1
    
    write_headers(ws, start_row, "СПБ", 1, has_low_reviews)
    start_row += 1
    
    first_spb_row = start_row
    for restaurant in SPB_ORDER:
        low_count = low_counts.get(restaurant, 0) if low_counts else 0
        write_restaurant_row(ws, start_row, restaurant, 1, df, has_low_reviews, low_count)
        row_map[restaurant] = start_row
        start_row += 1
    last_spb_row = start_row - 1
    
    write_total_row(ws, start_row, "СПб", 1, first_spb_row, last_spb_row, has_low_reviews)
    row_map["Итого СПб:"] = start_row
    start_row += 2
    
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=6)
    cell = ws.cell(row=start_row, column=2, value="Кол-во поставленных звезд")
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    start_row += 1
    
    write_headers(ws, start_row, "Тюмень", 1, has_low_reviews)
    start_row += 1
    
    first_tmn_row = start_row
    for restaurant in TMN_ORDER:
        low_count = low_counts.get(restaurant, 0) if low_counts else 0
        write_restaurant_row(ws, start_row, restaurant, 1, df, has_low_reviews, low_count)
        row_map[restaurant] = start_row
        start_row += 1
    last_tmn_row = start_row - 1
    
    write_total_row(ws, start_row, "Тюмень", 1, first_tmn_row, last_tmn_row, has_low_reviews)
    row_map["Итого Тюмень:"] = start_row
    start_row += 1
    
    return row_map, start_row

def write_analysis_sheet(writer, comp_all, pos_all):
    comp_all.to_excel(writer, sheet_name="Анализ отзывов", index=False, startrow=1)
    pos_all.to_excel(writer, sheet_name="Анализ отзывов", index=False, startrow=len(comp_all) + 4)
    ws = writer.sheets["Анализ отзывов"]
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(comp_all.columns))
    cell = ws.cell(row=1, column=1, value="Анализ отзывов (жалобы)")
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN
    
    for col_idx in range(1, len(comp_all.columns) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = GREEN_FILL
        cell.border = THIN_BORDER
    
    for row_idx in range(3, len(comp_all) + 3):
        for col_idx in range(1, len(comp_all.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN if col_idx > 1 else LEFT_ALIGN
            cell.border = THIN_BORDER
    
    pos_start_row = len(comp_all) + 4
    ws.merge_cells(start_row=pos_start_row, start_column=1, end_row=pos_start_row, end_column=len(pos_all.columns))
    cell = ws.cell(row=pos_start_row, column=1, value="Анализ отзывов (позитив)")
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN
    
    for col_idx in range(1, len(pos_all.columns) + 1):
        cell = ws.cell(row=pos_start_row + 1, column=col_idx)
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = GREEN_FILL
        cell.border = THIN_BORDER
    
    for row_idx in range(pos_start_row + 2, pos_start_row + 2 + len(pos_all)):
        for col_idx in range(1, len(pos_all.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN if col_idx > 1 else LEFT_ALIGN
            cell.border = THIN_BORDER
    
    for col_idx in range(1, max(len(comp_all.columns), len(pos_all.columns)) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

# ==========================================================
# ИНТЕРФЕЙС
# ==========================================================
st.markdown(f"""
<div class="header-block">
    <h1> КР отчёты</h1>
    <p>{greeting_by_time()}</p>
    <p style="margin-top:10px; margin-bottom:0; font-size:16px;">
        Формирование отчётов по отзывам из трёх источников
    </p>
</div>
""", unsafe_allow_html=True)

# Переключатель типа отчёта
report_type = st.radio(
    "📋 Тип отчёта",
    ["📅 КР месяц", "📆 КР неделя"],
    horizontal=True
)

st.markdown("### 📂 Загрузка файлов")
col1, col2, col3 = st.columns(3)

with col1:
    st.markdown("####  Сайт")
    file1 = st.file_uploader("Загрузите Excel", type=["xlsx", "xls"], key="site", label_visibility="collapsed")

with col2:
    st.markdown("#### 🛵 Агрегаторы")
    file2 = st.file_uploader("Загрузите Excel", type=["xlsx", "xls"], key="agg", label_visibility="collapsed")

with col3:
    st.markdown("#### 📍 Геосервисы")
    file3 = st.file_uploader("Загрузите Excel", type=["xlsx", "xls"], key="geo", label_visibility="collapsed")

st.markdown("---")
st.subheader("⚙️ Настройки")
col_a, col_b, col_c = st.columns(3)

with col_a:
    month_names = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                   "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
    current_month_idx = datetime.now().month - 1
    if current_month_idx == 0:
        default_month = "Декабрь"
        default_year = datetime.now().year - 1
    else:
        default_month = month_names[current_month_idx - 1]
        default_year = datetime.now().year
    selected_month = st.selectbox("Месяц отчёта", month_names, index=month_names.index(default_month))

with col_b:
    selected_year = st.number_input("Год", value=default_year, min_value=2020, max_value=2030, step=1)

with col_c:
    price_threshold = st.number_input("Минимальная сумма заказа", value=749, step=10)

st.markdown(" ")
generate_report = st.button("🚀 Сформировать отчёт", use_container_width=True, type="primary")

# ==========================================================
# ОБРАБОТКА ДАННЫХ
# ==========================================================
if generate_report:
    if not (file1 and file2 and file3):
        st.error("⚠️ Пожалуйста, загрузите все три Excel-файла.")
        st.stop()
    
    is_month = report_type == "📅 КР месяц"
    
    if is_month:
        period_text = f"{selected_month} ({selected_year}г)"
    else:
        today = datetime.now()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        period_text = f"{start_of_week.strftime('%d.%m')} - {end_of_week.strftime('%d.%m.%Y')}"
    
    with st.spinner("⏳ Обработка данных..."):
        try:
            # --- Сайт ---
            df1 = pd.read_excel(file1)
            df1["Ресторан"] = df1["Ресторан"].apply(lambda x: map_restaurant_file1(x, "month" if is_month else "week"))
            
            if is_month:
                df1["Сумма"] = df1["Сумма заказа со скидкой"].apply(parse_price)
                df1["Рейтинг"] = pd.to_numeric(df1["Рейтинг"], errors="coerce")
                df1["Текст"] = df1["Комментарий"].fillna(" ")
                df1 = df1[["Ресторан", "Рейтинг", "Текст", "Сумма"]].dropna(subset=["Ресторан", "Рейтинг"])
            else:
                df1["Рейтинг"] = pd.to_numeric(df1["Рейтинг"], errors="coerce")
                df1["Текст"] = df1["Комментарий"].fillna(" ")
                df1 = df1[["Ресторан", "Рейтинг", "Текст"]].dropna(subset=["Ресторан", "Рейтинг"])
            
            # --- Агрегаторы ---
            df2 = pd.read_excel(file2)
            df2["Ресторан"] = df2["Адрес"].apply(map_restaurant_address)
            df2["Рейтинг"] = pd.to_numeric(df2["Оценка"], errors="coerce")
            df2["Текст"] = df2["Отзыв"].fillna(" ")
            df2 = df2[["Ресторан", "Рейтинг", "Текст"]].dropna(subset=["Ресторан", "Рейтинг"])
            
            # --- Геосервисы ---
            df3 = pd.read_excel(file3)
            
            # КР Месяц: удаляем удалённые отзывы, КР Неделя: НЕ удаляем
            if "Статус отзыва" in df3.columns and is_month:
                df3 = df3[df3["Статус отзыва"] != "Удален"].copy()
            
            if "Название филиала" in df3.columns and "Адрес филиала" in df3.columns:
                df3["Ресторан"] = df3.apply(
                    lambda r: map_restaurant_address(str(r["Название филиала"]), is_month) or map_restaurant_address(str(r["Адрес филиала"]), is_month), 
                    axis=1
                )
            elif "Адрес филиала" in df3.columns:
                df3["Ресторан"] = df3["Адрес филиала"].apply(lambda x: map_restaurant_address(x, is_month))
            else:
                df3["Ресторан"] = None
            
            df3["Рейтинг"] = pd.to_numeric(df3["Оценка"], errors="coerce")
            df3["Текст"] = df3["Текст отзыва"].fillna(" ")
            df3 = df3[["Ресторан", "Рейтинг", "Текст"]].dropna(subset=["Ресторан", "Рейтинг"])
            
            # --- Таблицы оценок ---
            threshold = price_threshold if is_month else None
            filter_low = is_month
            
            stats1 = calc_stats(df1, threshold=threshold, filter_low_reviews=filter_low)
            stats2 = calc_stats(df2, filter_low_reviews=False)
            stats3 = calc_stats(df3, filter_low_reviews=False)
            
            # Объединяем все данные для общей статистики
            df_all_for_stats = pd.concat([
                df1[~((df1["Рейтинг"] == 5) & (df1["Сумма"] <= threshold))] if is_month else df1,
                df2, df3
            ], ignore_index=True)
            stats_all = calc_stats(df_all_for_stats, filter_low_reviews=False)
            
            # Считаем low_counts для месяца
            if is_month:
                low_counts = {
                    rest: len(df1[(df1["Ресторан"] == rest) & (df1["Рейтинг"] == 5) & (df1["Сумма"] <= threshold)]) 
                    for rest in ALL_RESTAURANTS
                }
            else:
                low_counts = None
            
            # --- Анализ отзывов ---
            df_all_texts = pd.concat([
                df1[["Ресторан", "Текст"]], 
                df2[["Ресторан", "Текст"]], 
                df3[["Ресторан", "Текст"]]
            ], ignore_index=True)
            
            comp_all = calc_summary_fast(df_all_texts, COMPLAINT_KEYWORDS)
            pos_all = calc_summary_fast(df_all_texts, POSITIVE_KEYWORDS)
            
            if is_month:
                spb_list = SPB_ORDER
                tmn_list = TMN_ORDER
            else:
                all_restaurants_in_data = sorted(df_all_texts["Ресторан"].unique())
                spb_list = [r for r in all_restaurants_in_data if "ТМН" not in str(r)]
                tmn_list = [r for r in all_restaurants_in_data if "ТМН" in str(r)]
            
            comp_all = add_summary_totals(add_summary_totals(comp_all, "СПб", COMPLAINT_KEYWORDS, spb_list, tmn_list), "Тюмень", COMPLAINT_KEYWORDS, spb_list, tmn_list)
            pos_all = add_summary_totals(add_summary_totals(pos_all, "СПб", POSITIVE_KEYWORDS, spb_list, tmn_list), "Тюмень", POSITIVE_KEYWORDS, spb_list, tmn_list)
            
            # =====================================================
            # СОХРАНЕНИЕ EXCEL
            # =====================================================
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                workbook = writer.book
                ws = workbook.create_sheet("Оценки")
                start_row = 1
                
                row_map_site, next_row = write_block(
                    ws, start_row, "Сайт/приложение", stats1, period_text,
                    has_low_reviews=is_month, low_counts=low_counts
                )
                agg_start_row = next_row + 2
                row_map_agg, next_row = write_block(
                    ws, agg_start_row, "Агрегаторы", stats2, period_text,
                    has_low_reviews=False, low_counts=None
                )
                geo_start_row = next_row + 2
                row_map_geo, next_row = write_block(
                    ws, geo_start_row, "Геосервисы (Я.Карты, 2ГИС, Гугл карты)", stats3, period_text,
                    has_low_reviews=False, low_counts=None
                )
                
                # Настройка ширины столбцов
                ws.column_dimensions['A'].width = 18
                for col_idx in range(2, 10):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 10
                ws.column_dimensions['J'].width = 3
                ws.column_dimensions['K'].width = 18
                for col_idx in range(12, 19):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 10
                
                write_analysis_sheet(writer, comp_all, pos_all)
                
                if "Sheet" in workbook.sheetnames:
                    del workbook["Sheet"]
            
            output.seek(0)
            
            st.success(f"✅ Отчёт за {period_text} успешно сформирован!")
            
            if is_month:
                file_name = f"КР_{selected_month}_{selected_year}.xlsx"
            else:
                file_name = f"КР_неделя_{start_of_week.strftime('%d%m')}-{end_of_week.strftime('%d%m')}.xlsx"
            
            st.download_button(
                "📥 Скачать Excel", output,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            
            st.subheader("📊 Общий итог")
            st.dataframe(stats_all, use_container_width=True)
            
        except Exception as e:
            st.error(f" Произошла ошибка при обработке: {e}")
            st.exception(e)
