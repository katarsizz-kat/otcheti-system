import io
import math
import re
from collections import OrderedDict
from datetime import datetime, timedelta, timezone
from difflib import get_close_matches

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# =========================
# ТЕМА И БЛОК "ОФОРМЛЕНИЕ" (v2.0)
# =========================
from styles import apply_subtle_theme
from config.greetings import get_current_greeting
from config.holidays import get_today_holiday

try:
    from components import render_theme_controls
except Exception:
    render_theme_controls = None

# =========================
# Настройки времени UTC+3
# =========================
MSK = timezone(timedelta(hours=3))

# Стандартный список номеров ресторанов, который вы указали
DEFAULT_RESTAURANT_NUMBERS = [1, 2, 4, 5, 6, 7, 8, 13, 15, 16, 18, 20]

# Цвета оформления Excel
HEADER_COLOR = "1F4E79"
BLOCK_COLOR = "DCE6F1"
TOTAL_COLOR = "FFF2CC"
VIOLATION_COLOR = "C00000"

# Ширины колонок
CATEGORY_WIDTH = 42
ITEM_WIDTH = 95
RESTAURANT_WIDTH = 7
TOTAL_WIDTH = 14

# Служебные строки, которые не являются пунктами чек-листа
IGNORE_EXACT = {
    "Пункт Результат",
    "Пункт Заметка Результат",
    "ФИО Общий",
    "Название пиццы Телега",
    "Название пиццы Общий",
    "Результаты чек-листа",
    "Блоки:",
    "Результат по контекстам выполнения:",
}

IGNORE_PREFIXES = (
    "Название чек-листа",
    "Имя проверяющего",
    "Проверяемый объект",
    "Время начала",
    "Время завершения",
    "Маркетинг",
    "Офис",
    "Ремонтные работы",
    "ЦКК",
    "Чек-лист пройден",
    "Баллов набрано",
    "Результат по контекстам",
)

HEADER_PREFIXES = (
    "ФИО Общий",
    "Название пиццы Телега",
    "Название пиццы Общий",
)

RESULT_RE = re.compile(r"\b(Да|Нет)\b[\.\s]*$")


# =========================
# Канонические категории без точек
# =========================
CANONICAL_BLOCKS = [
    "Внешний вид здания",
    "Зал",
    "Уголок потребителя",
    "Обслуживание гостей в зале",
    "Внешний вид сотрудников",
    "Производственная зона",
    "Производственная зона. Тесто",
    "Мейклайн",
    "Станция Слэп",
    "Раскатка традиционного теста",
    "Раскатка тонкого теста",
    "Оценка готовой пиццы",
    "Кат-Тейбл",
    "Оборудование кухни",
    "Станция для водителей",
    "Холодильная камера(основной)",
    "Холодильное оборудование",
    "Маркировка",
    "Овощная-баночная",
    "Сухой склад",
    "Мойка посуды",
    "Раковины и сантехника",
    "Уборочный инвентарь и моющие средства",
    "Раздевалка для персонала",
    "Туалет для персонала",
    "Туалет для гостей",
    "Риски",
    "Стоп-лист",
    "Документация",
    "Медикаменты",
    "Критические нарушения",
]


def canonical_block_key(value: str) -> str:
    """
    Делает ключ блока для сопоставления:
    убирает пробелы, точки, скобки, дефисы и другие небуквенно-цифровые символы.
    """
    value = str(value or "")
    value = value.lower().replace("ё", "е")
    value = re.sub(r"[^a-zа-яё0-9]+", "", value)
    return value


CANONICAL_BLOCK_MAP = {
    canonical_block_key(name): name
    for name in CANONICAL_BLOCKS
}


# =========================
# Корректировка категорий для проблемных мест PDF
# =========================
CATEGORY_OVERRIDE_RULES_RAW = [
    # Уголок потребителя
    (r"книга отзывов и предложений", "Уголок потребителя"),
    (r"список аллергенов", "Уголок потребителя"),

    # Обслуживание гостей в зале
    (r"шаг\s+[1-5]\s*й", "Обслуживание гостей в зале"),

    # Внешний вид сотрудников
    (r"пиццамейкер в бандане", "Внешний вид сотрудников"),
    (r"все сотрудники в поло", "Внешний вид сотрудников"),
    (r"на сотрудниках надет фартук", "Внешний вид сотрудников"),
    (r"именной бейдж", "Внешний вид сотрудников"),
    (r"брюки джинсы полной длины", "Внешний вид сотрудников"),
    (r"обувь закрытая", "Внешний вид сотрудников"),
    (r"на теле отсутствуют украшения", "Внешний вид сотрудников"),
    (r"мужчины выбриты", "Внешний вид сотрудников"),
    (r"длина ногтей соответствует стандартам", "Внешний вид сотрудников"),
    (r"серьги тоннели", "Внешний вид сотрудников"),
    (r"ремень черного или коричневого цвета", "Внешний вид сотрудников"),
    (r"носки для всех должностей", "Внешний вид сотрудников"),

    # Производственная зона
    (r"тяжелый докер", "Производственная зона"),
    (r"мелкий инвентарь убирают в лексан", "Производственная зона"),
    (r"подставка для скринов", "Производственная зона"),
    (r"термометры для теста", "Производственная зона"),
    (r"полы стены потолки и источники освещения чистые", "Производственная зона"),

    # Уборочный инвентарь и моющие средства
    (r"комплект для зоны туалет гостевой", "Уборочный инвентарь и моющие средства"),
    (r"комплект для зоны туалет для персонала", "Уборочный инвентарь и моющие средства"),
    (r"каждый комплект уборочного инвентаря хранится", "Уборочный инвентарь и моющие средства"),
    (r"уборочный инвентарь для зон туалет гостевой туалет персонала", "Уборочный инвентарь и моющие средства"),
    (r"отсутствует грязная вода в уборочном ведре", "Уборочный инвентарь и моющие средства"),

    # Раздевалка / туалеты
    (r"в наличии табличка раздевалка для персонала", "Раздевалка для персонала"),
    (r"в наличии табличка туалет для персонала", "Туалет для персонала"),
    (r"в наличии табличка обозначение туалет", "Туалет для гостей"),
    (r"график уборок туалетной комнаты", "Туалет для гостей"),

    # Риски
    (r"наличие насекомых не обнаружено", "Риски"),
    (r"ресторан не использует химические препараты", "Риски"),
    (r"кухонный инвентарь и оборудование технически исправно", "Риски"),

    # Стоп-лист
    (r"стоп лист в айко", "Стоп-лист"),

    # Документация
    (r"журнал лист контроля температуры влажности холодильного оборудования", "Документация"),
    (r"журнал генеральных уборок", "Документация"),
    (r"журнал контроля здоровья", "Документация"),
    (r"наличие действующей медицинской книжки", "Документация"),
    (r"личная подпись в лмк", "Документация"),
    (r"наличие в лмк печати", "Документация"),

    # Медикаменты
    (r"наличие аптечки", "Медикаменты"),
]

CATEGORY_OVERRIDE_RULES = [
    (re.compile(pattern), category)
    for pattern, category in CATEGORY_OVERRIDE_RULES_RAW
]


# =========================
# Проверка доступности pdfplumber
# =========================
try:
    import pdfplumber  # noqa: F401

    PDF_LIB_OK = True
except Exception:
    PDF_LIB_OK = False


# =========================
# Вспомогательные функции
# =========================
def clean_line(line: str) -> str:
    """Очищает строку от лишних пробелов и служебных символов."""
    line = str(line or "")
    line = line.replace("\u00a0", " ")
    line = line.replace("\\*", "*")
    line = re.sub(r"\s+", " ", line)
    return line.strip()


def remove_table_header_prefix(s: str) -> str:
    """Убирает случайно приклеившиеся заголовки таблиц к тексту пункта."""
    for prefix in HEADER_PREFIXES:
        if s.startswith(prefix):
            s = s[len(prefix):].strip()
    return s


def strip_score_prefix(s: str) -> str:
    """Убирает баллы вида 5/5 из начала строки."""
    s = re.sub(r"^\s*\d+\s*/\s*\d+\s*", "", s)
    return s.strip()


def is_noise_line(s: str) -> bool:
    """Проверяет, является ли строка служебной, а не пунктом чек-листа."""
    if not s:
        return True

    if s in IGNORE_EXACT:
        return True

    if re.fullmatch(r"\d+\s*/\s*\d+", s):
        return True

    for prefix in IGNORE_PREFIXES:
        if s.startswith(prefix):
            return True

    if re.match(r"^(Маркетинг|Офис|Ремонтные работы|ЦКК)\b", s):
        return True

    return False


def normalize_block(name: str) -> str:
    """Приводит название блока к аккуратному виду."""
    name = name.replace("\\*", "*")
    name = re.sub(r"\s+", " ", name).strip()
    return name


def canonicalize_block(raw_block: str) -> str:
    """
    Приводит распознанный блок к каноническому виду без точек в конце.
    """
    raw_block = normalize_block(raw_block)

    if not raw_block:
        return raw_block

    key = canonical_block_key(raw_block)

    # Точное совпадение
    if key in CANONICAL_BLOCK_MAP:
        return CANONICAL_BLOCK_MAP[key]

    # Частичное совпадение по началу ключа
    best_name = None
    best_len = 0

    for canon_key, canon_name in CANONICAL_BLOCK_MAP.items():
        if not canon_key:
            continue

        if key.startswith(canon_key) or canon_key.startswith(key):
            if len(canon_key) > best_len:
                best_name = canon_name
                best_len = len(canon_key)

    if best_name:
        return best_name

    # Дополнительная страховка от небольших отличий
    close_keys = get_close_matches(
        key,
        list(CANONICAL_BLOCK_MAP.keys()),
        n=1,
        cutoff=0.88,
    )

    if close_keys:
        return CANONICAL_BLOCK_MAP[close_keys[0]]

    # Если блок неизвестный, оставляем как есть, чтобы не потерять данные
    return raw_block


def detect_block(s: str):
    """Определяет строку как заголовок блока с процентом."""
    if not s or len(s) > 120:
        return None

    if "/" in s:
        return None

    if is_noise_line(s):
        return None

    if RESULT_RE.search(s):
        return None

    # Пример: Внешний вид здания.( 85,15%)
    m = re.match(r"^(?P<name>.+?)\s*\(\s*\d+(?:[.,]\d+)?\s*%\)\s*$", s)

    # Пример: Маркировка  16,67%
    if not m:
        m = re.match(r"^(?P<name>.+?)\s+\d+(?:[.,]\d+)?\s*%\s*$", s)

    if not m:
        return None

    name = m.group("name").strip()

    if not name or is_noise_line(name) or len(name) < 3:
        return None

    return canonicalize_block(name)


def beautify_item(s: str) -> str:
    """
    Лёгкое приведение текста к читаемому виду.
    Не меняет смысл пункта.
    """
    # Пробел перед скобкой, если его нет
    s = re.sub(r"(?<=[^\s])\(", " (", s)

    # Убираем лишние пробелы сразу после скобки и перед закрывающей скобкой
    s = re.sub(r"\(\s+", "(", s)
    s = re.sub(r"\s+\)", ")", s)

    # Добавляем пробел после запятой, если его нет
    s = re.sub(r",(?=\S)", ", ", s)

    # Сжимаем множественные пробелы
    s = re.sub(r"\s+", " ", s).strip()

    return s


def clean_item(s: str) -> str:
    """
    Очищает текст пункта чек-листа.

    Важно:
    - текст пункта сохраняется максимально полностью;
    - служебные баллы вида 5/5 убираются;
    - случайно приклеившиеся заголовки таблиц убираются;
    - теги вида *М*, *ОФ* сохраняются в отображении.
    """
    s = str(s or "")
    s = s.replace("\\*", "*")
    s = re.sub(r"\s+", " ", s).strip()

    s = strip_score_prefix(s)
    s = remove_table_header_prefix(s)

    s = beautify_item(s)

    return s


def normalize_for_rules(s: str) -> str:
    """
    Нормализация текста только для работы правил корректировки категории.
    """
    s = str(s or "")
    s = s.lower().replace("ё", "е")

    # Удаляем служебные теги вида *М*, *ОФ*, *Ц*, *Р*
    s = re.sub(r"\*[а-яёa-z]{1,4}\*", "", s)
    s = s.replace("*", "")

    # Оставляем только буквы, цифры и пробелы
    s = re.sub(r"[^a-zа-яё0-9\s]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()

    return s


def apply_category_override(block: str, item_text: str) -> str:
    """
    Применяет корректировку категории для известных проблемных пунктов.
    """
    if not item_text:
        return block

    normalized = normalize_for_rules(item_text)

    for pattern, target_category in CATEGORY_OVERRIDE_RULES:
        if pattern.search(normalized):
            return target_category

    return block


def normalize_item_key(s: str) -> str:
    """
    Нормализация пункта для сопоставления между ресторанами.

    Здесь уже убираем служебные теги и лишние символы,
    чтобы одинаковые пункты из разных файлов совпадали.
    """
    s = clean_item(s)
    s = s.lower().replace("ё", "е")

    # Удаляем служебные теги вида *М*, *ОФ*, *Ц*, *Р*
    s = re.sub(r"\*[а-яёa-z]{1,4}\*", "", s)

    # Удаляем оставшиеся звездочки, чтобы не мешали сопоставлению
    s = s.replace("*", "")

    # Агрессивная нормализация для сопоставления:
    # оставляем только буквы и цифры
    s = re.sub(r"[^a-zа-яё0-9]+", "", s)

    return s


def extract_restaurant(text: str, filename: str):
    """Достает номер и имя ресторана из PDF или имени файла."""
    m = re.search(r"Проверяемый объект:\s*(\d+)\s*([^\n]*)", text)
    if m:
        num = int(m.group(1))
        name = m.group(2).strip()
        return num, name

    m = re.match(r"^\s*(\d+)", filename)
    if m:
        num = int(m.group(1))
        name = re.sub(r"^\s*\d+\s*", "", filename)
        name = re.sub(r"\.pdf$", "", name, flags=re.IGNORECASE).strip()
        return num, name

    return None, filename


def estimate_row_height(text: str, width: float, min_height: int = 18, max_height: int = 240) -> int:
    """Примерная автоподстройка высоты строки Excel для переноса текста."""
    if not text:
        return min_height

    chars_per_line = max(int(width * 1.05), 12)
    lines = 0

    for part in str(text).split("\n"):
        lines += max(1, math.ceil(len(part) / chars_per_line))

    return min(max_height, max(min_height, lines * 15))


def order_blocks(blocks: OrderedDict) -> OrderedDict:
    """Ставит блоки в каноническом порядке, неизвестные блоки добавляет в конец."""
    ordered = OrderedDict()

    for name in CANONICAL_BLOCKS:
        if name in blocks:
            ordered[name] = blocks[name]

    for name, items in blocks.items():
        if name not in ordered:
            ordered[name] = items

    return ordered


# =========================
# Чтение PDF / TXT
# =========================
def load_pdf_text(file_obj) -> str:
    """Читает текст из PDF через pdfplumber."""
    if not PDF_LIB_OK:
        raise RuntimeError(
            "Не установлена библиотека pdfplumber. "
            "Добавьте в requirements.txt строку: pdfplumber==0.11.4"
        )

    file_obj.seek(0)
    parts = []

    with pdfplumber.open(file_obj) as pdf:
        for page in pdf.pages:
            txt = page.extract_text()
            if txt:
                parts.append(txt)

    return "\n".join(parts)


def load_any_text(file_obj) -> str:
    """Читает PDF или TXT."""
    name = str(getattr(file_obj, "name", "")).lower()

    if name.endswith(".txt"):
        file_obj.seek(0)
        return file_obj.read().decode("utf-8", errors="ignore")

    return load_pdf_text(file_obj)


# =========================
# Парсер VOC
# =========================
def parse_voc_text(text: str, filename: str):
    """
    Разбирает текст чек-листа VOC.

    Возвращает:
        restaurant_num,
        restaurant_name,
        results: {
            (block, normalized_item_key): {
                "value": 0/1,
                "display": "Текст пункта"
            }
        }
    """
    restaurant_num, restaurant_name = extract_restaurant(text, filename)

    results = {}
    block = None
    pending = []
    has_notes = False

    def add_result(block_name: str, item_text: str, result_word: str):
        if not block_name or not item_text:
            return

        item_clean = clean_item(item_text)
        if not item_clean or is_noise_line(item_clean):
            return

        norm_key = normalize_item_key(item_clean)
        if not norm_key:
            return

        # Корректируем категорию для известных проблемных пунктов
        final_block = apply_category_override(block_name, item_clean)

        value = 1 if result_word.strip().lower() == "нет" else 0
        key = (final_block, norm_key)

        if key not in results:
            results[key] = {
                "value": value,
                "display": item_clean,
            }
        else:
            if value == 1:
                results[key]["value"] = 1

            if not results[key].get("display"):
                results[key]["display"] = item_clean

    def commit_pending(result_word: str):
        nonlocal pending

        if pending and block:
            add_result(block, " ".join(pending), result_word)

        pending = []

    for raw_line in text.splitlines():
        line = clean_line(raw_line)

        if not line:
            continue

        # Отдельные заголовки таблиц
        if line in IGNORE_EXACT:
            if "Заметка" in line:
                has_notes = True
            elif line == "Пункт Результат":
                has_notes = False
            continue

        # Заголовок блока
        detected_block = detect_block(line)
        if detected_block:
            pending = []
            block = detected_block
            has_notes = False
            continue

        # Служебные строки
        if is_noise_line(line):
            continue

        # Убираем баллы из начала строки и случайно приклеившиеся заголовки
        s = strip_score_prefix(line)
        s = remove_table_header_prefix(s)

        if not s:
            continue

        # Если после очистки снова получился заголовок таблицы
        if s in IGNORE_EXACT:
            if "Заметка" in s:
                has_notes = True
            elif s == "Пункт Результат":
                has_notes = False
            continue

        # Убираем баллы в конце строки, если они есть
        s_no_tail_score = re.sub(r"\s*\d+\s*/\s*\d+\s*$", "", s).strip()

        # Ищем результат в конце строки
        m_res = RESULT_RE.search(s_no_tail_score)

        if m_res:
            item_part = s_no_tail_score[: m_res.start()].strip()
            result_word = m_res.group(1)

            if item_part:
                if pending:
                    full_item = " ".join(pending + [item_part])
                    add_result(block, full_item, result_word)
                    pending = []
                else:
                    add_result(block, item_part, result_word)
            else:
                commit_pending(result_word)

            continue

        # Если блок еще не найден, пункт некуда относить
        if block is None:
            continue

        # Эвристика: пропускаем свободные заметки в таблицах с колонкой "Заметка"
        if has_notes and pending:
            prev = pending[-1]

            if (
                s
                and s[0].isupper()
                and not s.startswith("*")
                and re.search(r"[.:;\)\"]\s*$", prev)
            ):
                continue

        pending.append(s)

    return restaurant_num, restaurant_name, results


def parse_files(files):
    """Парсит загруженные файлы."""
    parsed = []
    errors = []

    for f in files:
        try:
            text = load_any_text(f)
            num, name, results = parse_voc_text(text, f.name)

            parsed.append(
                {
                    "num": num,
                    "name": name,
                    "results": results,
                    "file": f.name,
                }
            )

        except Exception as e:
            errors.append(f"{f.name}: {e}")

    return parsed, errors


# =========================
# Сборка сводной таблицы
# =========================
def build_summary(parsed, include_default=False, only_uploaded=True):
    """
    Собирает свод:
    - блоки;
    - рестораны;
    - матрицу нарушений;
    - фильтрует только те пункты, где есть хотя бы один НЕТ.
    """
    blocks = OrderedDict()
    rest_info = OrderedDict()
    matrix = {}
    loaded_labels = set()

    def ensure_restaurant(num, name):
        if num is not None:
            label = str(int(num))
        else:
            label = (name or "Файл").strip()

        if label not in rest_info:
            rest_info[label] = {
                "num": num,
                "name": name,
            }
        else:
            if name and not rest_info[label].get("name"):
                rest_info[label]["name"] = name

            if num is not None and rest_info[label].get("num") is None:
                rest_info[label]["num"] = num

        return label

    # Наполняем данные из файлов
    for rec in parsed:
        label = ensure_restaurant(rec["num"], rec["name"])
        loaded_labels.add(label)

        for (block, norm_key), info in rec["results"].items():
            if block not in blocks:
                blocks[block] = OrderedDict()

            if norm_key not in blocks[block]:
                blocks[block][norm_key] = info.get("display", norm_key)

            key = (label, block, norm_key)
            val = int(info.get("value", 0))

            if key not in matrix:
                matrix[key] = val
            else:
                if val == 1:
                    matrix[key] = 1
                elif matrix[key] == 0 and val == 0:
                    matrix[key] = 0

    # При необходимости добавляем стандартные рестораны
    if include_default:
        for num in DEFAULT_RESTAURANT_NUMBERS:
            label = str(num)
            if label not in rest_info:
                rest_info[label] = {
                    "num": num,
                    "name": "",
                }

    # Сортировка ресторанов
    def sort_key(label):
        info = rest_info[label]

        if info.get("num") is not None:
            return (0, int(info["num"]), label)

        return (1, 0, label.lower())

    if only_uploaded:
        rest_labels = sorted(
            [label for label in rest_info.keys() if label in loaded_labels],
            key=sort_key,
        )
    else:
        rest_labels = sorted(rest_info.keys(), key=sort_key)

    # Оставляем только пункты, где есть хотя бы один НЕТ
    filtered_blocks = OrderedDict()

    for block, items in blocks.items():
        kept_items = OrderedDict()

        for norm_key, display in items.items():
            has_no = any(
                matrix.get((label, block, norm_key), 0) == 1
                for label in rest_labels
            )

            if has_no:
                kept_items[norm_key] = display

        if kept_items:
            filtered_blocks[block] = kept_items

    # Приводим блоки к нужному порядку
    filtered_blocks = order_blocks(filtered_blocks)

    return filtered_blocks, rest_labels, rest_info, matrix, loaded_labels


# =========================
# Генерация Excel
# =========================
def generate_excel(blocks, rest_labels, rest_info, matrix, loaded_labels):
    """Создает Excel-файл через openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = "VOC свод"

    thin = Side(style="thin", color="BFBFBF")
    double_black = Side(style="double", color="000000")

    base_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_col_border = Border(left=thin, right=double_black, top=thin, bottom=thin)
    total_row_border = Border(left=thin, right=thin, top=thin, bottom=double_black)
    total_corner_border = Border(left=thin, right=double_black, top=thin, bottom=double_black)

    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    block_fill = PatternFill("solid", fgColor=BLOCK_COLOR)
    total_fill = PatternFill("solid", fgColor=TOTAL_COLOR)

    wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Если нет данных
    if not blocks:
        ws.cell(row=1, column=1, value="Нет пунктов с результатом НЕТ")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    # Если нет ресторанов
    if not rest_labels:
        ws.cell(row=1, column=1, value="Нет ресторанов для отображения")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    # Ширины колонок
    ws.column_dimensions["A"].width = CATEGORY_WIDTH
    ws.column_dimensions["B"].width = ITEM_WIDTH

    start_col = 3

    for i, _ in enumerate(rest_labels):
        col_letter = get_column_letter(start_col + i)
        ws.column_dimensions[col_letter].width = RESTAURANT_WIDTH

    total_col = start_col + len(rest_labels)
    ws.column_dimensions[get_column_letter(total_col)].width = TOTAL_WIDTH

    last_rest_col = start_col + len(rest_labels) - 1
    last_rest_col_letter = get_column_letter(last_rest_col)

    # Заголовок таблицы
    headers = ["Категория", "Пункт"] + list(rest_labels) + ["Итого"]

    for idx, header in enumerate(headers, start=1):
        c = ws.cell(row=1, column=idx, value=header)
        c.fill = header_fill
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = center
        c.border = base_border

    ws.freeze_panes = "C2"

    row = 2
    item_rows = []

    # Тело таблицы
    for block, items in blocks.items():
        # Заголовок блока
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=total_col,
        )

        block_cell = ws.cell(row=row, column=1, value=block)
        block_cell.fill = block_fill
        block_cell.font = Font(bold=True)
        block_cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in range(1, total_col + 1):
            ws.cell(row=row, column=col).border = base_border

        ws.row_dimensions[row].height = 24
        row += 1

        # Пункты внутри блока
        for norm_key, display in items.items():
            ws.cell(row=row, column=1, value=block).alignment = wrap_left
            ws.cell(row=row, column=2, value=display).alignment = wrap_left

            for i, label in enumerate(rest_labels):
                col = start_col + i
                value = matrix.get((label, block, norm_key), None)

                cell = ws.cell(row=row, column=col)

                if value is not None:
                    cell.value = int(value)
                    cell.number_format = "0"

                    if value == 1:
                        cell.font = Font(color=VIOLATION_COLOR, bold=True)

                cell.alignment = center

            # Итого по строке
            formula_cell = ws.cell(
                row=row,
                column=total_col,
                value=f"=SUM({get_column_letter(start_col)}{row}:{last_rest_col_letter}{row})",
            )
            formula_cell.alignment = center
            formula_cell.font = Font(bold=True)
            formula_cell.number_format = "0"

            # Границы
            for col in range(1, total_col + 1):
                cell = ws.cell(row=row, column=col)

                if col == total_col:
                    cell.border = total_col_border
                else:
                    cell.border = base_border

            # Автоподбор высоты строки
            ws.row_dimensions[row].height = estimate_row_height(display, ITEM_WIDTH)

            item_rows.append(row)
            row += 1

    # Итоговая строка
    if item_rows:
        total_row = row
        first_item_row = item_rows[0]
        last_item_row = item_rows[-1]

        ws.merge_cells(
            start_row=total_row,
            start_column=1,
            end_row=total_row,
            end_column=2,
        )

        ws.cell(row=total_row, column=1, value="Итого").font = Font(bold=True)
        ws.cell(row=total_row, column=1).alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        for i, _ in enumerate(rest_labels):
            col = start_col + i
            col_letter = get_column_letter(col)

            cell = ws.cell(
                row=total_row,
                column=col,
                value=f"=SUM({col_letter}{first_item_row}:{col_letter}{last_item_row})",
            )
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.number_format = "0"

        total_sum_cell = ws.cell(
            row=total_row,
            column=total_col,
            value=f"=SUM({get_column_letter(start_col)}{total_row}:{last_rest_col_letter}{total_row})",
        )
        total_sum_cell.font = Font(bold=True)
        total_sum_cell.alignment = center
        total_sum_cell.number_format = "0"

        # Стили итоговой строки
        for col in range(1, total_col + 1):
            cell = ws.cell(row=total_row, column=col)
            cell.fill = total_fill

            if col == total_col:
                cell.border = total_corner_border
            else:
                cell.border = total_row_border

        ws.row_dimensions[total_row].height = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# =========================
# Предпросмотр в Streamlit
# =========================
def make_preview(blocks, rest_labels, matrix, limit=500):
    """Делает список словарей для предпросмотра."""
    rows = []

    for block, items in blocks.items():
        for norm_key, display in items.items():
            row = {
                "Блок": block,
                "Пункт": display,
            }

            total = 0

            for label in rest_labels:
                val = matrix.get((label, block, norm_key), None)
                row[label] = val if val is not None else ""

                if val == 1:
                    total += 1

            row["Итого"] = total
            rows.append(row)

            if len(rows) >= limit:
                return rows

    return rows


# =========================
# Интерфейс страницы
# =========================
def main():
    st.set_page_config(page_title="VOC", page_icon="📋", layout="wide")

    greeting_data = get_current_greeting() or {}
    holiday = get_today_holiday() or {}
    apply_subtle_theme(
        greeting_data.get("theme") if isinstance(greeting_data, dict) else None,
        holiday.get("effects") if isinstance(holiday, dict) else None,
    )
    if render_theme_controls is not None:
        try:
            render_theme_controls()
        except Exception:
            pass

    st.markdown(
        """
        <div class="header-block">
            <h1>📋 VOC: сводная таблица нарушений</h1>
            <p>Загрузите PDF-файлы VOC. В свод попадут только пункты,
            по которым хотя бы в одном ресторане есть результат НЕТ.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if not PDF_LIB_OK:
        st.error(
            "Не установлен pdfplumber. "
            "Добавьте в requirements.txt строку: pdfplumber==0.11.4"
        )

    files = st.file_uploader(
        "Загрузить файлы VOC PDF или TXT",
        type=["pdf", "txt"],
        accept_multiple_files=True,
    )

    if not files:
        st.info("Загрузите от 1 до 14 файлов.")
        return

    if len(files) > 14:
        st.warning("Можно обработать не более 14 файлов. Будут обработаны первые 14.")

    files = files[:14]

    st.markdown("---")

    include_all = st.checkbox(
        "Добавить все стандартные рестораны "
        "(1, 2, 4, 5, 6, 7, 8, 13, 15, 16, 18, 20), "
        "даже если файл по ресторану не загружен",
        value=False,
    )

    only_uploaded = not include_all

    if st.button("Обработать файлы", type="primary"):
        with st.spinner("Читаем файлы и извлекаем пункты..."):
            parsed, errors = parse_files(files)

        st.session_state["voc_parsed"] = parsed
        st.session_state["voc_errors"] = errors
        st.session_state["voc_file_names"] = [f.name for f in files]

    if "voc_parsed" not in st.session_state:
        return

    parsed = st.session_state.get("voc_parsed", [])
    errors = st.session_state.get("voc_errors", [])
    saved_names = st.session_state.get("voc_file_names", [])
    current_names = [f.name for f in files]

    if saved_names != current_names:
        st.info("Список файлов изменен. Нажмите «Обработать файлы» заново.")

    if errors:
        with st.expander("Ошибки обработки файлов"):
            for err in errors:
                st.write(err)

    if not parsed:
        st.error("Не удалось распознать ни одного файла.")
        return

    blocks, rest_labels, rest_info, matrix, loaded_labels = build_summary(
        parsed=parsed,
        include_default=include_all,
        only_uploaded=only_uploaded,
    )

    st.markdown("---")

    if not blocks:
        st.warning(
            "Нарушения с результатом НЕТ не найдены. "
            "Проверьте, что файлы являются текстовыми PDF-выгрузками VOC."
        )
        return

    # Метрики
    total_items = sum(len(items) for items in blocks.values())
    total_no = sum(
        1
        for block, items in blocks.items()
        for norm_key in items.keys()
        for label in rest_labels
        if matrix.get((label, block, norm_key), 0) == 1
    )

    c1, c2, c3, c4 = st.columns(4)

    c1.metric("Файлов обработано", len(parsed))
    c2.metric("Ресторанов в столбцах", len(rest_labels))
    c3.metric("Пунктов с нарушениями", total_items)
    c4.metric("Всего отметок НЕТ", total_no)

    # Файл Excel
    now_msk = datetime.now(MSK)
    file_name = f"VOC_Summary_Report_{now_msk.strftime('%Y-%m-%d')}.xlsx"

    excel_bytes = generate_excel(
        blocks=blocks,
        rest_labels=rest_labels,
        rest_info=rest_info,
        matrix=matrix,
        loaded_labels=loaded_labels,
    )

    st.download_button(
        label="Скачать сводный Excel",
        data=excel_bytes,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # Предпросмотр
    with st.expander("Предпросмотр таблицы"):
        preview = make_preview(blocks, rest_labels, matrix, limit=500)
        st.dataframe(preview, use_container_width=True)

        if total_items > len(preview):
            st.caption(
                "В предпросмотре показаны первые 500 строк. "
                "В Excel выгружаются все строки."
            )


main()