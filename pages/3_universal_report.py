# pages/3_universal_report.py
import streamlit as st
import pandas as pd
import re
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

st.set_page_config(page_title="Универсальный отчёт", layout="wide", page_icon="📊")

# Стили
st.markdown("""
<style>
.stApp { background: #f0f2f6 !important; }
.header-block {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    padding: 24px;
    border-radius: 16px;
    margin-bottom: 24px;
    color: white;
}
.content-block {
    background: white;
    padding: 24px;
    border-radius: 12px;
    margin-bottom: 24px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
}
.input-box {
    background: #fff3cd;
    border: 3px solid #ffc107;
    border-radius: 12px;
    padding: 20px;
    margin-bottom: 24px;
    box-shadow: 0 4px 15px rgba(255, 193, 7, 0.4);
}

/* Стили для текстового поля ввода */
textarea {
    border: 2px solid #4a90e2 !important;
    border-radius: 8px !important;
    padding: 12px !important;
    font-size: 14px !important;
    background-color: #ffffff !important;
    min-height: 300px !important;
}

textarea:focus {
    border-color: #2c5aa0 !important;
    box-shadow: 0 0 8px rgba(74, 144, 226, 0.3) !important;
    outline: none !important;
}

/* Стили для placeholder */
textarea::placeholder {
    color: #6c757d !important;
    font-style: italic;
    opacity: 0.8;
}

/* Стили для label поля ввода */
[data-testid="stTextArea"] label {
    font-weight: 600;
    color: #333;
    margin-bottom: 8px;
}
</style>
""", unsafe_allow_html=True)

# Функции
def normalize_text(text):
    if not isinstance(text, str): return ""
    # Латиница в кириллицу
    replacements = {'a': 'а', 'e': 'е', 'o': 'о', 'p': 'р', 'c': 'с', 'y': 'у', 'x': 'х',
                   'A': 'А', 'E': 'Е', 'O': 'О', 'P': 'Р', 'C': 'С', 'Y': 'У', 'X': 'Х', 'ё': 'е', 'Ё': 'Е'}
    for lat, cyr in replacements.items(): text = text.replace(lat, cyr)
    
    # Базовая очистка
    text = text.lower().replace('"', '').replace("'", "")
    text = re.sub(r'\(.*?\)', '', text)
    
    # Убираем слова-паразиты для пицц
    text = re.sub(r'\bпицца\b|\bpizza\b', '', text)
    text = re.sub(r'\bтонкое\b|\bтрад\b|\bтрадиционное\b|\bтесто\b|\bсм\b|\bnew\b', '', text)
    
    # Унификация чисел
    num_words = {'четыре': '4', 'пять': '5', 'шесть': '6', 'семь': '7', 'восемь': '8', 'девять': '9', 'десять': '10'}
    for word, digit in num_words.items():
        text = text.replace(word, digit)
        
    # Убираем лишние пробелы
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def extract_size(text):
    """Извлекает размер пиццы (15, 23, 30, 35, 40) из строки"""
    if not isinstance(text, str): return None
    match = re.search(r'\b(15|23|30|35|40)\s*см\b', str(text))
    return match.group(1) if match else None

def extract_pizza_name(text):
    """Извлекает название пиццы без размера и типа теста"""
    if not isinstance(text, str): return ""
    name = str(text)
    # Убираем размер
    name = re.sub(r'\s*\b(15|23|30|35|40)\s*см\b', '', name)
    # Убираем тип теста в скобках
    name = re.sub(r'\s*\(.*?(тонкое|трад|традиционное).*?\)', '', name)
    # Убираем слово "пицца"
    name = re.sub(r'\bпицца\b', '', name, flags=re.IGNORECASE)
    return name.strip()

def map_city(legal_entity):
    le = str(legal_entity)
    if 'СПБ' in le: return 'СПБ'
    elif 'ПД' in le and 'СПБ' not in le: return 'Тюмень'
    return None

def extract_period(uploaded_file):
    df = pd.read_excel(uploaded_file, header=None, nrows=10)
    for idx, row in df.iterrows():
        for val in row.values:
            if pd.notna(val) and 'Период' in str(val):
                match = re.search(r'(\d{2}\.\d{2}\.\d{4})\s+по\s+(\d{2}\.\d{2}\.\d{4})', str(val))
                if match:
                    start = datetime.strptime(match.group(1), '%d.%m.%Y')
                    end = datetime.strptime(match.group(2), '%d.%m.%Y')
                    return f"{start.strftime('%d.%m')}-{end.strftime('%d.%m.%Y')}"
    return f"01.{datetime.now().strftime('%m')}.{datetime.now().year}"

def read_file(uploaded_file):
    df = pd.read_excel(uploaded_file, header=None)
    header_row = None
    for idx, row in df.iterrows():
        vals = [str(v) for v in row.values if pd.notna(v)]
        if 'Юридическое лицо' in ' '.join(vals) and 'Блюдо' in ' '.join(vals):
            header_row = idx
            break
    if header_row is None: raise ValueError("Не найдены заголовки!")
    headers = df.iloc[header_row].values
    df = df[header_row + 1:].copy()
    df.columns = headers
    df = df.reset_index(drop=True)
    df = df[~df['Юридическое лицо'].astype(str).str.contains('OLAP|всего|Итого', na=False)]
    df = df.dropna(how='all')
    df['Юридическое лицо'] = df['Юридическое лицо'].ffill()
    df['Категория блюда'] = df['Категория блюда'].ffill()
    df = df[~df['Блюдо'].astype(str).str.contains(r'\+', na=False)]
    df = df[~df['Блюдо'].astype(str).str.contains('персонал', case=False, na=False)]
    df = df[df['Блюдо'].notna()]
    df['Количество блюд'] = pd.to_numeric(df['Количество блюд'], errors='coerce').fillna(0)
    df['Сумма со скидкой, р.'] = pd.to_numeric(df['Сумма со скидкой, р.'], errors='coerce').fillna(0)
    return df

def aggregate_data(df, rules_text):
    df = df.copy()
    df['Город'] = df['Юридическое лицо'].apply(map_city)
    df = df[df['Город'].notna()]
    
    # Предварительно вычисляем нормализованные имена и размеры
    df['Блюдо_norm'] = df['Блюдо'].apply(normalize_text)
    df['Размер'] = df['Блюдо'].apply(extract_size)
    
    rules = [line.strip() for line in rules_text.split('\n') if line.strip()]
    result = {}
    
    for city in ['СПБ', 'Тюмень']:
        city_data = df[df['Город'] == city]
        city_result = {}
        
        for rule in rules:
            rule_norm = normalize_text(rule)
            rule_size = extract_size(rule)
            rule_tokens = set(rule_norm.split())
            
            # Фильтрация по размеру (если он указан в правиле)
            if rule_size:
                subset = city_data[city_data['Размер'] == rule_size]
            else:
                subset = city_data.copy()
            
            # Проверка: все ли слова из правила есть в названии блюда
            def check_match(dish_norm):
                dish_tokens = set(dish_norm.split())
                return rule_tokens.issubset(dish_tokens)
            
            matches = subset[subset['Блюдо_norm'].apply(check_match)]
            
            city_result[rule] = {
                'qty': int(matches['Количество блюд'].sum()),
                'sum': round(matches['Сумма со скидкой, р.'].sum(), 2)
            }
        result[city] = city_result
    return result

def aggregate_pizza_by_size(df, rules_text):
    """Агрегирует данные по пиццам с разбивкой по размерам"""
    df = df.copy()
    df['Город'] = df['Юридическое лицо'].apply(map_city)
    df = df[df['Город'].notna()]
    
    # Добавляем поля для анализа
    df['Блюдо_norm'] = df['Блюдо'].apply(normalize_text)
    df['Размер'] = df['Блюдо'].apply(extract_size)
    df['Название_пиццы'] = df['Блюдо'].apply(extract_pizza_name)
    df['Название_пиццы_norm'] = df['Название_пиццы'].apply(normalize_text)
    
    rules = [line.strip() for line in rules_text.split('\n') if line.strip()]
    
    # Структура результата: {city: {pizza_name: {size: {qty, sum}}}}
    result = {}
    
    for city in ['СПБ', 'Тюмень']:
        city_data = df[df['Город'] == city]
        city_result = {}
        
        for rule in rules:
            rule_norm = normalize_text(rule)
            rule_tokens = set(rule_norm.split())
            
            # Находим все пиццы, соответствующие правилу
            def check_match(pizza_norm):
                pizza_tokens = set(pizza_norm.split())
                return rule_tokens.issubset(pizza_tokens)
            
            matches = city_data[city_data['Название_пиццы_norm'].apply(check_match)]
            
            # Группируем по размерам
            size_data = {}
            for size in ['15', '23', '30', '35', '40']:
                size_matches = matches[matches['Размер'] == size]
                size_data[size] = {
                    'qty': int(size_matches['Количество блюд'].sum()),
                    'sum': round(size_matches['Сумма со скидкой, р.'].sum(), 2)
                }
            
            # Итого
            total_qty = sum(size_data[s]['qty'] for s in size_data)
            total_sum = sum(size_data[s]['sum'] for s in size_data)
            size_data['Всего'] = {'qty': total_qty, 'sum': total_sum}
            
            city_result[rule] = size_data
        
        result[city] = city_result
    
    return result

def create_excel(data, period_str, rules):
    wb = Workbook()
    if 'Sheet' in wb.sheetnames: del wb['Sheet']
    ws = wb.create_sheet("Отчёт")
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    blue_fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
    
    # Период
    ws.merge_cells('A1:H1')
    c = ws.cell(1, 1, period_str)
    c.font = Font(bold=True, size=16)
    c.alignment = Alignment(horizontal='center')
    c.fill = yellow_fill
    
    # Города
    ws.merge_cells('A2:D2')
    c = ws.cell(2, 1, "СПб сейчас")
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal='center')
    c.fill = blue_fill
    
    ws.merge_cells('E2:H2')
    c = ws.cell(2, 5, "Тюмень сейчас")
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal='center')
    c.fill = blue_fill
    
    # Заголовки
    for col, header in [(1, 'Наименование'), (2, 'Количество'), (3, 'Сумма, ₽'),
                        (5, 'Наименование'), (6, 'Количество'), (7, 'Сумма, ₽')]:
        c = ws.cell(3, col, header)
        c.font = Font(bold=True, size=10)
        c.border = border
        c.alignment = Alignment(horizontal='center')
    
    # Данные
    row = 4
    for rule in rules:
        ws.cell(row, 1, rule).border = border
        ws.cell(row, 2, data['СПБ'][rule]['qty']).border = border
        ws.cell(row, 2).alignment = Alignment(horizontal='center')
        ws.cell(row, 3, data['СПБ'][rule]['sum']).border = border
        ws.cell(row, 3).number_format = '#,##0.00'
        
        ws.cell(row, 5, rule).border = border
        ws.cell(row, 6, data['Тюмень'][rule]['qty']).border = border
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        ws.cell(row, 7, data['Тюмень'][rule]['sum']).border = border
        ws.cell(row, 7).number_format = '#,##0.00'
        row += 1
    
    # Итого
    total_row = row + 1
    for col in [1, 5]:
        c = ws.cell(total_row, col, 'ИТОГО')
        c.font = Font(bold=True, size=11)
        c.border = border
    
    for col in [2, 6]:
        c = ws.cell(total_row, col, f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_row-1})")
        c.font = Font(bold=True, size=11)
        c.border = border
        c.alignment = Alignment(horizontal='center')
    
    for col in [3, 7]:
        c = ws.cell(total_row, col, f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_row-1})")
        c.font = Font(bold=True, size=11)
        c.border = border
        c.number_format = '#,##0.00'
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 3
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    
    return wb

def create_pizza_excel(data, period_str, rules):
    """Создает Excel с разбивкой пицц по размерам как на скриншоте"""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames: del wb['Sheet']
    ws = wb.create_sheet("Отчёт")
    
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    thick_border = Border(left=Side('thick'), right=Side('thick'), top=Side('thick'), bottom=Side('thick'))
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    blue_fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    # Заголовок периода
    ws.merge_cells('A1:G1')
    c = ws.cell(1, 1, period_str)
    c.font = Font(bold=True, size=16)
    c.alignment = Alignment(horizontal='center')
    c.fill = yellow_fill
    
    # Заголовок города (СПБ)
    ws.merge_cells('A2:G2')
    c = ws.cell(2, 1, "СПБ")
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal='center')
    c.fill = blue_fill
    
    # Заголовки колонок
    headers = ['Пиццы', '15 см', '23 см', '30 см', '35 см', '40 см', 'Всего']
    for col, header in enumerate(headers, 1):
        c = ws.cell(3, col, header)
        c.font = Font(bold=True, size=11)
        c.border = thick_border
        c.alignment = Alignment(horizontal='center')
        c.fill = gray_fill
    
    # Данные
    row = 4
    for rule in rules:
        # Название пиццы
        c = ws.cell(row, 1, rule)
        c.border = border
        c.alignment = Alignment(horizontal='left')
        
        # Размеры
        for col_idx, size in enumerate(['15', '23', '30', '35', '40'], 2):
            qty = data['СПБ'][rule][size]['qty']
            c = ws.cell(row, col_idx, qty)
            c.border = border
            c.alignment = Alignment(horizontal='center')
        
        # Итого
        total_qty = data['СПБ'][rule]['Всего']['qty']
        c = ws.cell(row, 7, total_qty)
        c.font = Font(bold=True)
        c.border = thick_border
        c.alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Итого строка
    ws.cell(row, 1, 'Итого')
    ws.cell(row, 1).font = Font(bold=True, size=11)
    ws.cell(row, 1).border = thick_border
    
    for col_idx, size in enumerate(['15', '23', '30', '35', '40'], 2):
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row, col_idx, f"=SUM({col_letter}4:{col_letter}{row-1})")
        c.font = Font(bold=True, size=11)
        c.border = thick_border
        c.alignment = Alignment(horizontal='center')
    
    # Общая сумма
    ws.cell(row, 7, f"=SUM(G4:G{row-1})")
    ws.cell(row, 7).font = Font(bold=True, size=11)
    ws.cell(row, 7).border = thick_border
    ws.cell(row, 7).alignment = Alignment(horizontal='center')
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    
    return wb

# Интерфейс
st.markdown("""
<div class="header-block">
    <h1>📊 Универсальный отчёт</h1>
    <p>Умный отчёт с разбивкой пицц по размерам и нечётким поиском</p>
</div>
""", unsafe_allow_html=True)

# Описание работы
st.markdown("""
<div class="content-block">
    <h3>📋 Как работает отчёт</h3>
    
    <h4> Стандартный режим:</h4>
    <ul>
        <li><b>Гибкий поиск:</b> Введите название позиции — программа найдёт все совпадения (порядок слов не важен).</li>
        <li><b>Примеры:</b> <code>Картофель из печи 150 гр</code>, <code>Рогалики с колбасками</code>, <code>Сырная пицца</code>.</li>
        <li><b>Авто-очистка:</b> Игнорируются слова <i>"пицца", "тонкое", "традиционное", "тесто", "см", "new"</i>.</li>
        <li><b>Унификация:</b> <code>4 сыра</code> и <code>четыре сыра</code> считаются одной позицией.</li>
        <li><b>Результат:</b> Количество и сумма по СПб и Тюмени для каждой позиции.</li>
    </ul>
    
    <h4>🍕 Пиццы по размерам:</h4>
    <ul>
        <li><b>Если нужен отчёт по пиццам с разбивкой по размерам</b> — переключитесь на раздел <b>" Пиццы по размерам"</b> сверху.</li>
        <li>Программа автоматически сгруппирует все размеры (15, 23, 30, 35, 40 см) в отдельные колонки.</li>
        <li>Тонкое и традиционное тесто суммируются внутри каждого размера.</li>
    </ul>
</div>
""", unsafe_allow_html=True)

# Переключатель режима
report_mode = st.radio(
    "📋 Выберите режим отчёта:",
    ["📝 Стандартный отчёт", "🍕 Пиццы по размерам"],
    horizontal=True,
    help="Стандартный - обычный отчёт с количеством и суммой\nПиццы по размерам - отчёт с разбивкой по размерам 15, 23, 30, 35, 40 см"
)

# ПОЛЕ ВВОДА
st.markdown('<div class="input-box">', unsafe_allow_html=True)
st.markdown("### ️ ВВЕДИТЕ СПИСОК ПОЗИЦИЙ")
st.markdown("<p><b>Каждая строка = одна позиция в отчёте.</b></p>", unsafe_allow_html=True)

if report_mode == "🍕 Пиццы по размерам":
    default_rules = """Сырная
Пепперони
Четыре Сыра
Мясная
Супер Папа
Гавайская
Цыпленок Барбекю
Маргарита
Мексиканская
Ветчина и Грибы"""
else:
    default_rules = """Сырная 30
Четыре сыра 35
Пепперони 30
Картофель из печи 150 гр
Рогалики с колбасками"""

rules_text = st.text_area(
    "Список позиций:",
    value="",
    height=400 if report_mode == "🍕 Пиццы по размерам" else 300,
    key="rules_input",
    label_visibility="visible",
    placeholder=default_rules
)
st.markdown('</div>', unsafe_allow_html=True)

# ЗАГРУЗКА ФАЙЛА
st.markdown('<div class="content-block">', unsafe_allow_html=True)
st.markdown("### 📂 Загрузка файла")
file_main = st.file_uploader("Загрузите файл рейтинг_продукт (Excel)", type=['xlsx'], key="universal")
st.markdown('</div>', unsafe_allow_html=True)

# ГЕНЕРАЦИЯ
if file_main:
    st.markdown('<div class="content-block">', unsafe_allow_html=True)
    
    try:
        period_str = extract_period(file_main)
        file_main.seek(0)
        st.info(f"📅 Период: **{period_str}**")
    except:
        period_str = f"01.{datetime.now().strftime('%m')}.{datetime.now().year}"
        st.warning("⚠️ Не удалось определить период")
    
    if st.button("📊 Сгенерировать отчёт", type="primary", use_container_width=True):
        with st.spinner("Формирую отчёт..."):
            try:
                df = read_file(file_main)
                rules = [line.strip() for line in rules_text.split('\n') if line.strip()]
                
                if not rules:
                    st.error("❌ Введите хотя бы одну позицию в поле выше!")
                else:
                    if report_mode == " Пиццы по размерам":
                        # Режим отчёта по пиццам с разбивкой по размерам
                        data = aggregate_pizza_by_size(df, rules_text)
                        wb = create_pizza_excel(data, period_str, rules)
                        
                        # Превью
                        st.success("✅ Отчёт сформирован!")
                        st.subheader(" Превью (СПБ)")
                        
                        preview_data = []
                        for rule in rules:
                            row = {'Позиция': rule}
                            for size in ['15', '23', '30', '35', '40']:
                                row[f'{size} см'] = data['СПБ'][rule][size]['qty']
                            row['Всего'] = data['СПБ'][rule]['Всего']['qty']
                            preview_data.append(row)
                        
                        preview = pd.DataFrame(preview_data)
                        st.dataframe(preview, use_container_width=True)
                    else:
                        # Стандартный режим
                        data = aggregate_data(df, rules_text)
                        wb = create_excel(data, period_str, rules)
                        
                        st.success("✅ Отчёт сформирован!")
                        st.subheader("📋 Превью (СПБ)")
                        preview = pd.DataFrame([
                            {'Позиция': rule, 'Количество': data['СПБ'][rule]['qty'], 'Сумма, ₽': data['СПБ'][rule]['sum']}
                            for rule in rules
                        ])
                        st.dataframe(preview, use_container_width=True)
                    
                    # Скачивание
                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)
                    
                    file_name = f"Отчёт_{'pizza_' if report_mode == '🍕 Пиццы по размерам' else ''}{period_str.replace('.', '-')}.xlsx"
                    
                    st.download_button(
                        label="📥 Скачать Excel",
                        data=output,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            except Exception as e:
                st.error(f"❌ Ошибка: {str(e)}")
                import traceback
                st.code(traceback.format_exc())
    st.markdown('</div>', unsafe_allow_html=True)
else:
    st.info("👆 Загрузите файл для начала работы")
