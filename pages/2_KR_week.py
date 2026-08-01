import streamlit as st
import pandas as pd
import io
import warnings
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ВАЖНО: st.set_page_config должен быть ПЕРВЫМ вызовом streamlit в файле!
st.set_page_config(page_title="📆 КР неделя", page_icon="📈", layout="wide")

# ==========================================================
# ДИНАМИЧЕСКИЕ СТИЛИ (НЕБО ПО ВРЕМЕНИ СУТОК)
# ==========================================================

def get_sky_gradient():
    """Возвращает градиент неба в зависимости от времени суток."""
    hour = datetime.now().hour
    
    if 5 <= hour < 12:  # Утро - восход
        return {
            "header": "linear-gradient(135deg, #FF8C42 0%, #5DADE2 100%)",
            "block": "linear-gradient(135deg, #FFF5EB 0%, #EBF5FB 100%)",
            "text": "#FFFFFF",
            "button_text": "#FFFFFF",
            "button_bg": "#FF6B35"
        }
    elif 12 <= hour < 18:  # День - ясное небо
        return {
            "header": "linear-gradient(135deg, #3498DB 0%, #EBF5FB 100%)",
            "block": "linear-gradient(135deg, #EBF5FB 0%, #FFFFFF 100%)",
            "text": "#FFFFFF",
            "button_text": "#FFFFFF",
            "button_bg": "#2E86C1"
        }
    elif 18 <= hour < 23:  # Вечер - закат
        return {
            "header": "linear-gradient(135deg, #2874A6 0%, #F39C12 100%)",
            "block": "linear-gradient(135deg, #FDEBD0 0%, #D4E6F1 100%)",
            "text": "#FFFFFF",
            "button_text": "#FFFFFF",
            "button_bg": "#D68910"
        }
    else:  # Ночь
        return {
            "header": "linear-gradient(135deg, #1B4F72 0%, #5DADE2 100%)",
            "block": "linear-gradient(135deg, #D6EAF8 0%, #FFFFFF 100%)",
            "text": "#FFFFFF",
            "button_text": "#1B4F72",
            "button_bg": "#2E86C1"
        }

def apply_dynamic_styles():
    """Применяет динамические стили."""
    sky = get_sky_gradient()
    
    st.markdown(f"""
    <style>
        .stApp {{
            background: {sky['block']} !important;
        }}
        .header-block {{
            background: {sky['header']};
            padding: 24px;
            border-radius: 16px;
            margin-bottom: 24px;
            box-shadow: 0 6px 18px rgba(0,0,0,0.15);
            border: 2px solid rgba(255,255,255,0.3);
        }}
        .header-block h1 {{
            margin: 0;
            color: {sky['text']};
            font-size: 36px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
        }}
        .header-block p {{
            margin-top: 8px;
            margin-bottom: 0;
            font-size: 18px;
            color: {sky['text']};
            text-shadow: 1px 1px 2px rgba(0,0,0,0.2);
        }}
        .content-block {{
            background: rgba(255, 255, 255, 0.95);
            padding: 24px;
            border-radius: 12px;
            margin-bottom: 24px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
            border: 1px solid #D6EAF8;
        }}
        .stButton>button {{
            background: {sky['button_bg']} !important;
            color: {sky['button_text']} !important;
            border: none !important;
            border-radius: 8px !important;
            font-weight: bold !important;
            box-shadow: 0 4px 8px rgba(0,0,0,0.2) !important;
            transition: all 0.3s ease !important;
        }}
        .stButton>button:hover {{
            filter: brightness(1.1) !important;
            transform: translateY(-2px) !important;
        }}
        .stSelectbox > div > div, .stNumberInput > div > div > input {{
            border: 1px solid #D6EAF8 !important;
            border-radius: 8px !important;
        }}
        hr {{
            border: none;
            border-top: 2px solid #D6EAF8;
            margin: 24px 0;
        }}
    </style>
    """, unsafe_allow_html=True)

def apply_file_uploader_styles():
    """Применяет стили для блоков загрузки файлов."""
    st.markdown("""
    <style>
        .element-container:has(> .stFileUploader) {
            background: white;
            border-radius: 12px;
            padding: 20px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
            border: 2px solid transparent;
            transition: all 0.3s ease;
        }
        .element-container:has(> .stFileUploader):hover {
            transform: translateY(-4px);
            box-shadow: 0 8px 20px rgba(0,0,0,0.12);
        }
        /* Сайт/приложение - голубой */
        .element-container:nth-child(2):has(> .stFileUploader) {
            border-color: #3498DB;
            background: linear-gradient(135deg, #EBF5FB 0%, #FFFFFF 100%);
        }
        /* Агрегаторы - ЖЁЛТЫЙ */
        .element-container:nth-child(3):has(> .stFileUploader) {
            border-color: #F1C40F;
            background: linear-gradient(135deg, #FEF9E7 0%, #FFFFFF 100%);
        }
        /* Геосервисы - зелёный */
        .element-container:nth-child(4):has(> .stFileUploader) {
            border-color: #27AE60;
            background: linear-gradient(135deg, #E8F8F5 0%, #FFFFFF 100%);
        }
        .stFileUploader [data-testid="stFileUploader"] {
            background: white;
            border-radius: 8px;
            padding: 10px;
        }
        .stFileUploader label {
            font-weight: bold;
            color: #2C3E50;
        }
        .stFileUploader > div > div {
            border: 2px dashed rgba(0,0,0,0.2) !important;
            border-radius: 8px !important;
            background: rgba(255,255,255,0.8) !important;
            transition: all 0.3s ease !important;
        }
        .stFileUploader > div > div:hover {
            border-color: #3498DB !important;
            background: rgba(255,255,255,1) !important;
        }
        .stFileUploader small {
            color: #7F8C8D;
            font-size: 12px;
        }
    </style>
    """, unsafe_allow_html=True)

# ==========================================================
# НАСТРОЙКИ
# ==========================================================

RESTAURANT_MAP = {
    "Санкт-Петербург №1": "01 Транспортный", "Санкт-Петербург №2": "02 Димитрова",
    "Санкт-Петербург №4": "04 Шмидта", "Санкт-Петербург №5": "05 Пулковская",
    "Санкт-Петербург №6": "06 Благодатная", "Санкт-Петербург №7": "07 Энтузиастов",
    "Санкт-Петербург №8": "08 Серебристый", "Санкт-Петербург №13": "13 Мурино",
    "Санкт-Петербург №15": "15 Ветеранов", "Санкт-Петербург №16": "16 Туристская",
    "Санкт-Петербург №18": "18 Наука", "Санкт-Петербург №20": "20 Ленинский",
    "Тюмень №2 – Орджоникидзе": "ТМН 2 (Орджоникидзе)", "Тюмень №3 – Мельникайте": "ТМН 3 (Мельникайте)"
}

COMPLAINT_KEYWORDS = {
    "Жалоба на продукт": r"сух|холодн|остыл|невкусн|плох|ужас|отврат|прогоркл|испорч|стар|жестк|резин|волос|металл|гряз|воняет|гнил|сырой",
    "Ошибки приготовления": r"пережар|недожар|сыро|сгорел|пригорел|пересолен|недосолен|мало соли|много соли|непропеч",
    "Перепутанные / недоложенные позиции": r"не положили|не доложили|забыли|не было|не дали|перепутали|не тот|вместо|заменя|замен|не хватает|недоложили|недодали|не пришло|не привезли",
    "Жалобы на сервис": r"хам|груб|невежлив|неприятн|игнор|сбросил|отказал|не ответил|не помог|не решил|проблем|жалоб|администратор",
    "Опоздание": r"опозда|опоздал|опоздание|долго|задерж|задержк|не вовремя|не успел|час|полчаса|30 минут|40 минут|50 минут|1 час|1\.5 часа|полтора часа"
}

POSITIVE_KEYWORDS = {
    "Вкус": r"вкусн|отличн|превосходн|восхитит|бомб|огонь|пушк|шедевр|идеальн|сочн",
    "Быстрая доставка": r"быстр|оперативн|вовремя|минут|горяч|с пылу с жару|молниеносн",
    "Вежливый персонал": r"вежлив|приятн|доброжелат|внимател|учтив|милый|отзывчив|профессионал",
    "Качество": r"качеств|свеж|хорош|отличн|превосходн|много начинк|много сыра",
    "Атмосфера": r"уютн|атмосфер|чист|комфортн|приятн|красив|интерьер"
}

# ==========================================================
# ФУНКЦИИ
# ==========================================================

def greeting_by_time():
    hour = datetime.now().hour
    if 5 <= hour < 12: return "☀️ Доброе утро!"
    if 12 <= hour < 18: return "🌤 Добрый день!"
    if 18 <= hour < 23: return "🌙 Добрый вечер!"
    return "🌜 Доброй ночи!"

def map_restaurant_file1(val):
    if pd.isna(val): return None
    return RESTAURANT_MAP.get(str(val).strip(), None)

def map_restaurant_address(val):
    if pd.isna(val): return None
    val_lower = str(val).lower()
    address_map = {
        "транспортный": "01 Транспортный", "димитрова": "02 Димитрова",
        "шмидта": "04 Шмидта", "13-я линия": "04 Шмидта", "васильевск": "04 Шмидта",
        "пулковская": "05 Пулковская", "благодатная": "06 Благодатная",
        "энтузиастов": "07 Энтузиастов", "серебристый": "08 Серебристый",
        "мурино": "13 Мурино", "петровский бульвар": "13 Мурино",
        "ветеранов": "15 Ветеранов", "туристская": "16 Туристская",
        "науки": "18 Наука", "ленинский": "20 Ленинский",
        "орджоникидзе": "ТМН 2 (Орджоникидзе)", "мельникайте": "ТМН 3 (Мельникайте)"
    }
    for key, value in address_map.items():
        if key in val_lower: return value
    return None

def calc_stats(df):
    results = []
    for rest in sorted(df["Ресторан"].unique()):
        sub = df[df["Ресторан"] == rest]
        row = {"Ресторан": rest}
        for i in range(1, 6): row[str(i)] = int((sub["Рейтинг"] == i).sum())
        row["всего:"] = sum(row[str(i)] for i in range(1, 6))
        row["Средний рейтинг"] = round(sum(i * row[str(i)] for i in range(1, 6)) / row["всего:"], 2) if row["всего:"] > 0 else 0
        results.append(row)
    return pd.DataFrame(results)

def calc_summary_fast(df, keywords_dict):
    """Векторизованный анализ отзывов (быстрый)."""
    results = []
    for rest in sorted(df["Ресторан"].unique()):
        sub = df[df["Ресторан"] == rest].copy()
        counts = {cat: 0 for cat in keywords_dict.keys()}
        if not sub.empty:
            for category, pattern in keywords_dict.items():
                counts[category] = int(sub["Текст"].str.contains(pattern, case=False, na=False, regex=True).sum())
        row_data = {"Ресторан": rest, **counts}
        row_data["Всего:"] = sum(counts.values())
        results.append(row_data)
    return pd.DataFrame(results)

# ==========================================================
# ФУНКЦИЯ СОЗДАНИЯ EXCEL С ФОРМУЛАМИ
# ==========================================================

def create_excel_with_formulas(stats1, stats2, stats3, stats_all, complaints_all, positives_all, spb, tmn):
    """Создаёт Excel-файл с формулами и форматированием по шаблону."""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wb = writer.book
        
        # ==========================================================
        # ЛИСТ 1: ОЦЕНКИ
        # ==========================================================
        ws = wb.create_sheet("Оценки", 0)
        
        # Стили
        header_font = Font(bold=True, size=12, color="FFFFFF")
        subheader_font = Font(bold=True, size=10)
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        double_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='double'),
            bottom=Side(style='double')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Цвета заголовков
        site_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        agg_fill = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
        geo_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        total_fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        
        # Заголовок периода
        ws.merge_cells('A1:H1')
        ws['A1'] = "1-5 Апреля"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align
        
        # Функция для добавления блока с данными
        def add_data_block(ws, start_row, title, title_fill, stats_df, spb_list, tmn_list):
            row = start_row
            
            # Заголовок источника
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = title
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = title_fill
            ws[f'A{row}'].alignment = center_align
            
            row += 1
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = "Кол-во поставленных звезд"
            ws[f'A{row}'].font = subheader_font
            ws[f'A{row}'].alignment = center_align
            
            row += 1
            # Заголовки столбцов
            headers = ["СПБ", "1", "2", "3", "4", "5", "всего:", "Средний рейтинг"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = bold_font
                cell.alignment = center_align
                cell.border = thin_border
            
            # Данные по СПб
            row += 1
            spb_start_row = row
            for rest in spb_list:
                rest_data = stats_df[stats_df["Ресторан"] == rest]
                if rest_data.empty:
                    continue
                rest_data = rest_data.iloc[0]
                ws.cell(row=row, column=1, value=rest)
                for i in range(1, 6):
                    ws.cell(row=row, column=i+1, value=int(rest_data[str(i)]))
                # Формулы
                ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
                ws.cell(row=row, column=8, value=f'=IF(G{row}=0,0,SUMPRODUCT(B{row}:F{row},{{1,2,3,4,5}})/G{row})')
                ws.cell(row=row, column=8).number_format = '0.00'
                
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=col).alignment = center_align
                row += 1
            
            spb_end_row = row - 1
            
            # Итого СПб
            ws.cell(row=row, column=1, value="Итого СПб:")
            ws.cell(row=row, column=1).font = bold_font
            for i in range(2, 7):
                col_letter = get_column_letter(i)
                ws.cell(row=row, column=i, value=f"=SUM({col_letter}{spb_start_row}:{col_letter}{spb_end_row})")
            ws.cell(row=row, column=7, value=f"=SUM(G{spb_start_row}:G{spb_end_row})")
            ws.cell(row=row, column=8, value=f'=IF(G{row}=0,0,SUMPRODUCT(B{row}:F{row},{{1,2,3,4,5}})/G{row})')
            ws.cell(row=row, column=8).number_format = '0.00'
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).font = bold_font
                ws.cell(row=row, column=col).border = double_border
                ws.cell(row=row, column=col).alignment = center_align
            
            row += 2
            
            # Заголовки для Тюмени
            headers_tmn = ["Тюмень", "1", "2", "3", "4", "5", "всего:", "Средний рейтинг"]
            for col, header in enumerate(headers_tmn, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = bold_font
                cell.alignment = center_align
                cell.border = thin_border
            
            row += 1
            tmn_start_row = row
            for rest in tmn_list:
                rest_data = stats_df[stats_df["Ресторан"] == rest]
                if rest_data.empty:
                    continue
                rest_data = rest_data.iloc[0]
                ws.cell(row=row, column=1, value=rest)
                for i in range(1, 6):
                    ws.cell(row=row, column=i+1, value=int(rest_data[str(i)]))
                ws.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
                ws.cell(row=row, column=8, value=f'=IF(G{row}=0,0,SUMPRODUCT(B{row}:F{row},{{1,2,3,4,5}})/G{row})')
                ws.cell(row=row, column=8).number_format = '0.00'
                
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=col).alignment = center_align
                row += 1
            
            tmn_end_row = row - 1
            
            # Итого Тюмень
            ws.cell(row=row, column=1, value="Итого Тюмень:")
            ws.cell(row=row, column=1).font = bold_font
            for i in range(2, 7):
                col_letter = get_column_letter(i)
                ws.cell(row=row, column=i, value=f"=SUM({col_letter}{tmn_start_row}:{col_letter}{tmn_end_row})")
            ws.cell(row=row, column=7, value=f"=SUM(G{tmn_start_row}:G{tmn_end_row})")
            ws.cell(row=row, column=8, value=f'=IF(G{row}=0,0,SUMPRODUCT(B{row}:F{row},{{1,2,3,4,5}})/G{row})')
            ws.cell(row=row, column=8).number_format = '0.00'
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).font = bold_font
                ws.cell(row=row, column=col).border = double_border
                ws.cell(row=row, column=col).alignment = center_align
            
            return row + 2
        
        # Добавляем блоки
        current_row = add_data_block(ws, 3, "Сайт/приложение", site_fill, stats1, spb, tmn)
        current_row = add_data_block(ws, current_row, "Агрегаторы", agg_fill, stats2, spb, tmn)
        current_row = add_data_block(ws, current_row, "Геосервисы (Я.Карты, 2ГИС, Гугл карты)", geo_fill, stats3, spb, tmn)
        current_row = add_data_block(ws, current_row, "Общий (сайт+агрегаторы+геосервисы)", total_fill, stats_all, spb, tmn)
        
        # Настройка ширины столбцов
        ws.column_dimensions['A'].width = 20
        for col in range(2, 9):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # ==========================================================
        # ЛИСТ 2: АНАЛИЗ ОТЗЫВОВ
        # ==========================================================
        ws2 = wb.create_sheet("Анализ отзывов", 1)
        
        # Заголовок
        ws2.merge_cells('A1:G1')
        ws2['A1'] = "Сводная по жалобам"
        ws2['A1'].font = Font(bold=True, size=14)
        ws2['A1'].alignment = center_align
        
        # --- СПб ---
        row = 3
        ws2.merge_cells(f'A{row}:G{row}')
        ws2[f'A{row}'] = "1-26 Апреля"
        ws2[f'A{row}'].font = Font(bold=True, size=12)
        ws2[f'A{row}'].alignment = center_align
        
        row += 1
        headers_complaints = ["Ресторан", "Жалоба на продукт", "Ошибки приготовления", 
                             "Перепутанные/недоложенные позиции", "Жалобы на сервис", "Опоздание", "Всего:"]
        for col, header in enumerate(headers_complaints, 1):
            cell = ws2.cell(row=row, column=col, value=header)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        row += 1
        start_row_comp = row
        for rest in spb:
            rest_data = complaints_all[complaints_all["Ресторан"] == rest]
            if rest_data.empty:
                continue
            rest_data = rest_data.iloc[0]
            ws2.cell(row=row, column=1, value=rest)
            col = 2
            for category in COMPLAINT_KEYWORDS.keys():
                ws2.cell(row=row, column=col, value=int(rest_data[category]))
                col += 1
            ws2.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
            
            for col in range(1, 8):
                ws2.cell(row=row, column=col).border = thin_border
                ws2.cell(row=row, column=col).alignment = center_align
            row += 1
        
        # Итого СПб
        ws2.cell(row=row, column=1, value="Всего:")
        ws2.cell(row=row, column=1).font = bold_font
        for i in range(2, 7):
            col_letter = get_column_letter(i)
            ws2.cell(row=row, column=i, value=f"=SUM({col_letter}{start_row_comp}:{col_letter}{row-1})")
        ws2.cell(row=row, column=7, value=f"=SUM(G{start_row_comp}:G{row-1})")
        
        for col in range(1, 8):
            ws2.cell(row=row, column=col).font = bold_font
            ws2.cell(row=row, column=col).border = double_border
            ws2.cell(row=row, column=col).alignment = center_align
        
        # --- Тюмень ---
        row += 3
        ws2.merge_cells(f'A{row}:G{row}')
        ws2[f'A{row}'] = "Тюмень"
        ws2[f'A{row}'].font = Font(bold=True, size=12)
        ws2[f'A{row}'].alignment = center_align
        
        row += 1
        ws2.merge_cells(f'A{row}:G{row}')
        ws2[f'A{row}'] = "1-26 Апреля"
        ws2[f'A{row}'].font = Font(bold=True, size=12)
        ws2[f'A{row}'].alignment = center_align
        
        row += 1
        for col, header in enumerate(headers_complaints, 1):
            cell = ws2.cell(row=row, column=col, value=header)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        row += 1
        start_row_tmn = row
        for rest in tmn:
            rest_data = complaints_all[complaints_all["Ресторан"] == rest]
            if rest_data.empty:
                continue
            rest_data = rest_data.iloc[0]
            ws2.cell(row=row, column=1, value=rest)
            col = 2
            for category in COMPLAINT_KEYWORDS.keys():
                ws2.cell(row=row, column=col, value=int(rest_data[category]))
                col += 1
            ws2.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})")
            
            for col in range(1, 8):
                ws2.cell(row=row, column=col).border = thin_border
                ws2.cell(row=row, column=col).alignment = center_align
            row += 1
        
        # Итого Тюмень
        ws2.cell(row=row, column=1, value="Всего:")
        ws2.cell(row=row, column=1).font = bold_font
        for i in range(2, 7):
            col_letter = get_column_letter(i)
            ws2.cell(row=row, column=i, value=f"=SUM({col_letter}{start_row_tmn}:{col_letter}{row-1})")
        ws2.cell(row=row, column=7, value=f"=SUM(G{start_row_tmn}:G{row-1})")
        
        for col in range(1, 8):
            ws2.cell(row=row, column=col).font = bold_font
            ws2.cell(row=row, column=col).border = double_border
            ws2.cell(row=row, column=col).alignment = center_align
        
        # Настройка ширины столбцов
        ws2.column_dimensions['A'].width = 20
        for col in range(2, 8):
            ws2.column_dimensions[get_column_letter(col)].width = 18
    
    output.seek(0)
    return output

# ==========================================================
# ПРИМЕНЯЕМ СТИЛИ
# ==========================================================
apply_dynamic_styles()
apply_file_uploader_styles()

# ==========================================================
# ИНТЕРФЕЙС
# ==========================================================

st.markdown(f"""
<div class="header-block">
    <h1>📈 КР неделя</h1>
    <p>{greeting_by_time()}</p>
    <p style="margin-top:10px; margin-bottom:0; font-size:16px;">Еженедельный отчёт по отзывам из трёх источников</p>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="content-block">', unsafe_allow_html=True)
st.markdown("###  Загрузка файлов")
col1, col2, col3 = st.columns(3)

with col1:
    st.markdown("#### 📱 Сайт / приложение")
    file1 = st.file_uploader("Загрузите Excel", type=["xlsx", "xls"], key="site", label_visibility="collapsed")
with col2:
    st.markdown("####  Агрегаторы")
    file2 = st.file_uploader("Загрузите Excel", type=["xlsx", "xls"], key="agg", label_visibility="collapsed")
with col3:
    st.markdown("#### 🗺 Геосервисы")
    file3 = st.file_uploader("Загрузите Excel", type=["xlsx", "xls"], key="geo", label_visibility="collapsed")
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="content-block">', unsafe_allow_html=True)
generate_report = st.button("🚀 Сформировать отчёт", use_container_width=True, type="primary")
st.markdown('</div>', unsafe_allow_html=True)

# ==========================================================
# ОБРАБОТКА ДАННЫХ
# ==========================================================

if generate_report:
    if not (file1 and file2 and file3):
        st.error("⚠️ Пожалуйста, загрузите все 3 файла!")
        st.stop()

    with st.spinner("⏳ Обработка данных..."):
        try:
            # --- Сайт ---
            df1 = pd.read_excel(file1)
            df1["Ресторан"] = df1["Ресторан"].apply(map_restaurant_file1)
            df1["Текст"] = df1["Комментарий"].fillna("")
            df1["Рейтинг"] = pd.to_numeric(df1["Рейтинг"], errors="coerce")
            df1 = df1[["Ресторан", "Рейтинг", "Текст"]].dropna(subset=["Ресторан", "Рейтинг"])

            # --- Агрегаторы ---
            df2 = pd.read_excel(file2)
            df2["Ресторан"] = df2["Адрес"].apply(map_restaurant_address)
            df2["Текст"] = df2["Отзыв"].fillna("")
            df2["Рейтинг"] = pd.to_numeric(df2["Оценка"], errors="coerce")
            df2 = df2[["Ресторан", "Рейтинг", "Текст"]].dropna(subset=["Ресторан", "Рейтинг"])

            # --- Геосервисы (без фильтрации удалённых) ---
            df3 = pd.read_excel(file3)
            
            if "Название филиала" in df3.columns and "Адрес филиала" in df3.columns:
                df3["Ресторан"] = df3.apply(
                    lambda r: map_restaurant_address(str(r["Название филиала"])) or map_restaurant_address(str(r["Адрес филиала"])),
                    axis=1
                )
            elif "Адрес филиала" in df3.columns:
                df3["Ресторан"] = df3["Адрес филиала"].apply(map_restaurant_address)
            else:
                df3["Ресторан"] = None

            df3["Текст"] = df3["Текст отзыва"].fillna("")
            df3["Рейтинг"] = pd.to_numeric(df3["Оценка"], errors="coerce")
            df3 = df3[["Ресторан", "Рейтинг", "Текст"]].dropna(subset=["Ресторан", "Рейтинг"])

            # --- Общие данные ---
            df_all = pd.concat([df1, df2, df3], ignore_index=True)
            spb = sorted([r for r in df_all["Ресторан"].unique() if "ТМН" not in str(r)])
            tmn = sorted([r for r in df_all["Ресторан"].unique() if "ТМН" in str(r)])

            stats1 = calc_stats(df1)
            stats2 = calc_stats(df2)
            stats3 = calc_stats(df3)
            stats_all = calc_stats(df_all)

            # --- Анализ отзывов ---
            complaints_all = calc_summary_fast(df_all, COMPLAINT_KEYWORDS)
            positives_all = calc_summary_fast(df_all, POSITIVE_KEYWORDS)

            # --- Создание Excel с формулами ---
            output = create_excel_with_formulas(
                stats1, stats2, stats3, stats_all,
                complaints_all, positives_all, spb, tmn
            )
            
            st.success("✅ Отчёт успешно сформирован!")
            st.download_button(
                " Скачать Excel", output,
                file_name="КР_неделя_отчет.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.markdown('<div class="content-block">', unsafe_allow_html=True)
            st.subheader("📊 Превью: Общий итог")
            st.dataframe(stats_all, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        except Exception as e:
            st.error(f"❌ Произошла ошибка при обработке: {e}")
            st.exception(e)
